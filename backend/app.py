import os, io, uuid, logging, smtplib, json, base64, time, threading, queue, hashlib, functools, random, ipaddress, re
import requests

# ── Base Directory setup (resolves paths absolute to app.py) ──────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── DMVIC Certificate & Key Path Configuration ────────────────────────────────
DMVIC_CERT_PATH = os.path.join(BASE_DIR, 'certs', 'dmvic', 'dmvic_prod_cert.pem')
DMVIC_KEY_PATH = os.path.join(BASE_DIR, 'certs', 'dmvic', 'dmvic_prod_key.pem')

# Convenience tuple if passing mTLS to requests (cert, key)
DMVIC_CLIENT_CERT = (DMVIC_CERT_PATH, DMVIC_KEY_PATH) if (
    os.path.exists(DMVIC_CERT_PATH) and os.path.exists(DMVIC_KEY_PATH)
) else None

# ── Load .env (if present) into os.environ BEFORE any require_env() calls ────
try:
    from dotenv import load_dotenv
    # Look for .env in the same directory as app.py
    load_dotenv(os.path.join(BASE_DIR, '.env'))
except ImportError:
    logging.warning("python-dotenv not installed — .env file will not be loaded automatically. "
                    "Run `pip install python-dotenv`, or export the required env vars manually.")

from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from functools import wraps

from flask import (Flask, render_template, request, jsonify, Response,
                   session, redirect, send_file, abort)
from flask_cors import CORS
from flask_wtf.csrf import CSRFProtect, CSRFError, generate_csrf
from mongo_store import MongoStore
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

# DMVIC product classification (psv/commercial/private/motorcycle buckets).
try:
    from dmvic_mapping import PRODUCT_TO_CERT_TYPE, PRODUCT_TO_VEHICLE_TYPE_B
    DMVIC_MAPPING_AVAILABLE = True
except ImportError:
    PRODUCT_TO_CERT_TYPE = {}
    PRODUCT_TO_VEHICLE_TYPE_B = {}
    DMVIC_MAPPING_AVAILABLE = False
    logging.warning("dmvic_mapping.py not found next to app.py — DMVIC issuance "
                    "will mark every policy as 'unsupported' until it's added.")

# ── Rate limiting (flask-limiter) ─────────────────────────────────────────────
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    LIMITER_AVAILABLE = True
except ImportError:
    LIMITER_AVAILABLE = False
    logging.warning("flask-limiter not installed — rate limiting disabled")

# ── Redis caching (optional — app degrades gracefully if no Redis server) ────
try:
    import redis
    REDIS_LIB_AVAILABLE = True
except ImportError:
    REDIS_LIB_AVAILABLE = False
    logging.warning("redis package not installed — caching disabled")

# ── PDF generation (reportlab) ────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.units import cm
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("reportlab not installed — PDF generation disabled")

# ── Excel generation (openpyxl) ───────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not installed — Excel export disabled")

# ── File-type sniffing for upload validation ─────────────────────────────────
try:
    import filetype  # pip install filetype
    FILETYPE_AVAILABLE = True
except ImportError:
    FILETYPE_AVAILABLE = False
    logging.warning("filetype package not installed — upload content validation degraded to extension-only")


# ─────────────────────────────────────────────────────────────────────────────
# ENVIRONMENT / SECURITY BOOTSTRAP
# ─────────────────────────────────────────────────────────────────────────────

ENV = os.environ.get('FLASK_ENV', 'production').lower()
IS_PRODUCTION = ENV == 'production'
DEBUG_MODE = os.environ.get('FLASK_DEBUG', '0') == '1' and not IS_PRODUCTION


def require_env(name, allow_dev_fallback=None):
    """SECURITY: Load a required secret from the environment. In production
    there is NO fallback value — the app refuses to start rather than run
    with a known/hardcoded credential. In non-production, an explicit dev
    fallback may be supplied (never anything resembling a real secret)."""
    val = os.environ.get(name)
    if val:
        return val
    if not IS_PRODUCTION and allow_dev_fallback is not None:
        logging.warning("%s not set — using LOCAL DEV placeholder. Never do this in production.", name)
        return allow_dev_fallback
    raise RuntimeError(
        f"Required environment variable '{name}' is not set. Refusing to start "
        f"with a hardcoded credential. Set it in your environment/.env before running."
    )


logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# APP SETUP
# ─────────────────────────────────────────────────────────────────────────────

app = Flask(
    __name__,
    template_folder='../templates',
    static_folder='../css',
    static_url_path='/css'
)

app.secret_key = require_env('SECRET_KEY', allow_dev_fallback=os.urandom(32).hex())

ALLOWED_ORIGINS = [o.strip() for o in os.environ.get('ALLOWED_ORIGINS', '').split(',') if o.strip()]
if not ALLOWED_ORIGINS:
    if IS_PRODUCTION:
        raise RuntimeError("ALLOWED_ORIGINS must be set in production (comma-separated list of trusted origins).")
    ALLOWED_ORIGINS = ["http://localhost:3000", "http://127.0.0.1:3000"]
    log.warning("ALLOWED_ORIGINS not set — defaulting to local dev origins: %s", ALLOWED_ORIGINS)

CORS(app, supports_credentials=True, origins=ALLOWED_ORIGINS)

# ── CSRF protection (Flask-WTF) ──────────────────────────────────────────────
# csrf_token() is NOT automatically available in templates just by
# instantiating CSRFProtect — that only happens for WTForms-rendered forms.
# Since our templates call {{ csrf_token() }} directly in a <meta> tag (not
# via a WTForm), we must register it ourselves as a Jinja global here.
# WITHOUT this block, every page render raises
# jinja2.exceptions.UndefinedError: 'csrf_token' is undefined.
csrf = CSRFProtect(app)


@app.errorhandler(CSRFError)
def handle_csrf_error(error):
    return jsonify({"error": "Your form session expired. Refresh the page and try again."}), 400

@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=generate_csrf)

app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE']   = os.environ.get('SESSION_COOKIE_SECURE', '1' if IS_PRODUCTION else '0') == '1'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['MAX_CONTENT_LENGTH']      = 16 * 1024 * 1024  # 16 MB max upload

UPLOAD_FOLDER   = os.path.join(os.path.dirname(__file__), 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}
ALLOWED_MIME_BY_EXT = {
    'pdf':  {'application/pdf'},
    'png':  {'image/png'},
    'jpg':  {'image/jpeg'},
    'jpeg': {'image/jpeg'},
}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# RATE LIMITING
# ─────────────────────────────────────────────────────────────────────────────

if LIMITER_AVAILABLE:
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        default_limits=["200 per minute"],
        storage_uri=os.environ.get('RATELIMIT_STORAGE_URI', 'memory://'),
    )
else:
    class _NullLimiter:
        def limit(self, *a, **k):
            def deco(f):
                return f
            return deco
    limiter = _NullLimiter()
    if IS_PRODUCTION:
        log.error("flask-limiter is not installed but this is a production environment — "
                  "brute-force protection on login/register is DISABLED. Install flask-limiter.")


# ─────────────────────────────────────────────────────────────────────────────
# REDIS CACHE
# ─────────────────────────────────────────────────────────────────────────────

REDIS_URL = os.environ.get('REDIS_URL', 'redis://localhost:6379/0')
redis_client = None
if REDIS_LIB_AVAILABLE:
    try:
        redis_client = redis.from_url(REDIS_URL, socket_connect_timeout=1, decode_responses=True)
        redis_client.ping()
        log.info("Redis cache connected at %s", REDIS_URL)
    except Exception as e:
        log.warning("Redis not reachable (%s) — running without cache", e)
        redis_client = None


def cache_get(key):
    if not redis_client:
        return None
    try:
        val = redis_client.get(key)
        return json.loads(val) if val else None
    except Exception as e:
        log.warning("Cache read error: %s", e)
        return None


def cache_set(key, value, ttl=60):
    if not redis_client:
        return
    try:
        redis_client.setex(key, ttl, json.dumps(value, default=str))
    except Exception as e:
        log.warning("Cache write error: %s", e)


def cache_delete_prefix(prefix):
    if not redis_client:
        return
    try:
        for k in redis_client.scan_iter(f"{prefix}*"):
            redis_client.delete(k)
    except Exception as e:
        log.warning("Cache invalidation error: %s", e)


def cached_response(prefix, ttl=60, key_fn=None):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            cache_key = key_fn() if key_fn else \
                f"{prefix}:{session.get('user_id')}:{session.get('role')}:{request.query_string.decode()}"
            hit = cache_get(cache_key)
            if hit is not None:
                return jsonify(hit)
            resp = f(*args, **kwargs)
            try:
                body = resp.get_json()
                if body is not None:
                    cache_set(cache_key, body, ttl)
            except Exception:
                pass
            return resp
        return wrapper
    return decorator


# ─────────────────────────────────────────────────────────────────────────────
# BACKGROUND TASK QUEUE
# ─────────────────────────────────────────────────────────────────────────────

task_queue = queue.Queue()
NUM_WORKERS = int(os.environ.get('BACKGROUND_WORKERS', 2))


def _background_worker(worker_id):
    while True:
        job_name, fn, args, kwargs = task_queue.get()
        try:
            log.info("[worker-%s] running background job: %s", worker_id, job_name)
            fn(*args, **kwargs)
        except Exception as e:
            log.error("[worker-%s] background job '%s' failed: %s", worker_id, job_name, e)
        finally:
            task_queue.task_done()


for i in range(NUM_WORKERS):
    threading.Thread(target=_background_worker, args=(i,), daemon=True).start()


def enqueue(job_name, fn, *args, **kwargs):
    task_queue.put((job_name, fn, args, kwargs))


# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────

MONGODB_URI = require_env('MONGODB_URI', allow_dev_fallback='mongodb://127.0.0.1:27017/westlake_insurance')
MONGODB_DB_NAME = os.environ.get('MONGODB_DB_NAME', 'westlake_insurance')

mongo_store = MongoStore(
    MONGODB_URI,
    MONGODB_DB_NAME,
    admin_email=os.environ.get('ADMIN_EMAIL'),
    admin_password=os.environ.get('ADMIN_PASSWORD'),
)
log.info("MongoDB store ready for database '%s'", MONGODB_DB_NAME)


def query(sql, params=(), fetchone=False, commit=False):
    return mongo_store.query(sql, params, fetchone=fetchone, commit=commit)


# ─────────────────────────────────────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────────────────────────────────────

COMPANY_EMAIL = os.environ.get('COMPANY_EMAIL', 'westlakeagencyltd@gmail.com')
SMTP_EMAIL    = os.environ.get('SMTP_EMAIL', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_HOST     = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT     = int(os.environ.get('SMTP_PORT', 587))
BREVO_API_KEY = os.environ.get('BREVO_API_KEY', '').strip()
BREVO_SENDER_EMAIL = os.environ.get('BREVO_SENDER_EMAIL', 'noreply@zeelineinsurance.tech').strip()
BREVO_SENDER_NAME = os.environ.get('BREVO_SENDER_NAME', 'Zee Line Risk Solutions').strip()
BREVO_VERIFICATION_TEMPLATE_ID = os.environ.get('BREVO_VERIFICATION_TEMPLATE_ID', '').strip()
BREVO_PASSWORD_RESET_TEMPLATE_ID = os.environ.get('BREVO_PASSWORD_RESET_TEMPLATE_ID', '').strip()
PUBLIC_SITE_URL = os.environ.get('PUBLIC_SITE_URL', 'https://zeelineinsurance.tech').rstrip('/')

if IS_PRODUCTION and not BREVO_API_KEY and (not SMTP_EMAIL or not SMTP_PASSWORD):
    log.warning("No Brevo or SMTP credentials are configured — outbound email is disabled.")


# ─────────────────────────────────────────────────────────────────────────────
# ARCHPAY PAYMENT CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

ARCHPAY_MODE            = os.environ.get('ARCHPAY_MODE', 'live')
ARCHPAY_BASE_URL        = os.environ.get('ARCHPAY_BASE_URL', 'https://pay.archietech.app/api/v1').rstrip('/')
ARCHPAY_API_KEY         = require_env('ARCHPAY_API_KEY', allow_dev_fallback='')
ARCHPAY_CHANNEL_ID      = os.environ.get('ARCHPAY_CHANNEL_ID', '').strip()
ARCHPAY_CALLBACK_SECRET = require_env('ARCHPAY_CALLBACK_SECRET', allow_dev_fallback=os.urandom(16).hex())
ARCHPAY_WEBHOOK_URL     = os.environ.get('ARCHPAY_WEBHOOK_URL', '').strip()

# Backwards-compatible names for existing routes/templates that still say mpesa.
MPESA_ENV = ARCHPAY_MODE
MPESA_CALLBACK_SECRET = ARCHPAY_CALLBACK_SECRET

def mpesa_format_phone(phone):
    phone = str(phone).strip().replace(' ','').replace('-','').replace('+','')
    if phone.startswith('0') and len(phone) == 10:
        phone = '254' + phone[1:]
    if phone.startswith('254') and len(phone) == 12:
        return phone
    return None

def mpesa_stk_push(phone, amount, account_ref, description='Insurance Premium'):
    if not ARCHPAY_API_KEY:
        return {'success': False, 'error': 'ArchPay API key is not configured.'}
    phone_fmt = mpesa_format_phone(phone)
    if not phone_fmt:
        return {'success': False, 'error': f'Invalid phone number: {phone}'}
    amount_int = max(1, int(round(float(amount))))
    payload = {
        "phone": phone_fmt,
        "amount": amount_int,
        "accountReference": str(account_ref)[:12],
        "description": str(description)[:20],
    }
    if ARCHPAY_CHANNEL_ID:
        payload["channelId"] = ARCHPAY_CHANNEL_ID
    if ARCHPAY_WEBHOOK_URL:
        # ArchPay uses this per-request URL for the final M-Pesa result.  A
        # channel-level default is not reliable when several applications use
        # the same ArchPay account.
        payload["callbackUrl"] = ARCHPAY_WEBHOOK_URL

    try:
        res  = requests.post(
            f"{ARCHPAY_BASE_URL}/stkpush",
            json=payload,
            headers={
                "x-api-key": ARCHPAY_API_KEY,
                "Content-Type": "application/json",
            },
            timeout=20
        )
        data = res.json()
        log.info("ArchPay STK push status=%s success=%s", res.status_code, data.get('success'))
        if res.ok and data.get('success'):
            return {
                'success':             True,
                'checkout_request_id': data.get('checkoutRequestId'),
                'merchant_request_id': data.get('merchantRequestId'),
                'customer_message':    data.get('message') or 'Check your phone for the M-Pesa prompt.',
                'credits_remaining':   data.get('creditsRemaining'),
                'channel_id':          data.get('channelId'),
            }
        return {
            'success': False,
            'error': data.get('error') or data.get('message') or 'STK push failed',
            'code': data.get('code'),
        }
    except Exception as e:
        log.error("ArchPay STK push error: %s", type(e).__name__)
        return {'success': False, 'error': 'Could not reach ArchPay. Please try again.'}

def mpesa_query_status(checkout_request_id):
    if not ARCHPAY_API_KEY:
        return {'success': False, 'error': 'ArchPay API key is not configured.'}
    try:
        res  = requests.post(
            f"{ARCHPAY_BASE_URL}/verify",
            json={"checkoutRequestId": checkout_request_id},
            headers={
                "x-api-key": ARCHPAY_API_KEY,
                "Content-Type": "application/json",
            },
            timeout=15
        )
        data = res.json()
        status = (data.get('status') or '').strip().lower()
        receipt = data.get('mpesaReceiptNumber')
        terminal_failures = {'cancelled', 'canceled', 'timeout', 'reversed'}
        paid = bool(res.ok and (
            (data.get('success') and status in {'completed', 'success', 'successful', 'paid'})
            # ArchPay has returned `success: false, status: failed` for
            # completed collections while still supplying the Safaricom
            # receipt.  A receipt is the payment proof, except for an explicit
            # reversal/cancellation/timeout.
            or (receipt and status not in terminal_failures)
        ))
        if paid and receipt and status == 'failed':
            log.warning("ArchPay verify returned a receipt with failed status; accepting receipt as settled")
        log.info("ArchPay verify status=%s payment_status=%s", res.status_code, status)
        return {
            'success': paid,
            'status': status,
            'mpesa_receipt_number': receipt,
            'amount': data.get('amount'),
            'phone': data.get('phone'),
            'transaction_date': data.get('transactionDate'),
            'result_desc': data.get('error') or status or 'Payment not completed',
            'code': data.get('code'),
        }
    except Exception as e:
        log.error("ArchPay verify error: %s", type(e).__name__)
        return {'success': False, 'error': 'Could not verify payment status. Please try again.'}


def send_brevo_email(to, subject=None, html_body=None, attachments=None, *, template_id=None, params=None, tags=None):
    """Send transactional mail through Brevo's HTTPS API without exposing credentials in logs."""
    if not BREVO_API_KEY:
        return False

    payload = {
        "sender": {"email": BREVO_SENDER_EMAIL, "name": BREVO_SENDER_NAME},
        "to": [{"email": to}],
    }
    if template_id:
        try:
            payload["templateId"] = int(template_id)
        except (TypeError, ValueError):
            log.error("Brevo template ID is invalid")
            return False
        payload["params"] = params or {}
    else:
        payload["subject"] = subject
        payload["htmlContent"] = html_body
    if tags:
        payload["tags"] = tags
    if attachments:
        payload["attachment"] = [
            {"name": filename, "content": base64.b64encode(content).decode("ascii")}
            for filename, content in attachments
        ]

    try:
        response = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            json=payload,
            headers={"api-key": BREVO_API_KEY, "Content-Type": "application/json"},
            timeout=15,
        )
    except requests.RequestException as exc:
        log.error("Brevo email request failed: %s", type(exc).__name__)
        return False

    if response.ok:
        log.info("Brevo transactional email accepted for delivery")
        return True
    log.error("Brevo email rejected with HTTP %s", response.status_code)
    return False


def send_email(to, subject, html_body, attachments=None):
    if BREVO_API_KEY:
        return send_brevo_email(to, subject, html_body, attachments)
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        log.warning("SMTP not configured — email not sent")
        return False
    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From']    = SMTP_EMAIL
        msg['To']      = to
        msg.attach(MIMEText(html_body, 'html'))

        if attachments:
            for fname, data in attachments:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(data)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition',
                                f'attachment; filename="{fname}"')
                msg.attach(part)

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            s.login(SMTP_EMAIL, SMTP_PASSWORD)
            s.send_message(msg)
        return True
    except Exception as e:
        log.error("Email error: %s", type(e).__name__)
        return False


def notify_new_agent(full_name, username, email):
    html = f"""
    <h2>New Agent Registration — Westlake Insurance</h2>
    <p><b>Name:</b> {full_name}<br>
       <b>Username:</b> {username}<br>
       <b>Email:</b> {email}<br>
       <b>Status:</b> PENDING — please review in the Admin Dashboard.</p>
    """
    send_email(COMPANY_EMAIL, "New Agent Registration", html)


def notify_quotation(agent, client, vehicle, insurance, quote, pdf_bytes=None):
    html = f"""
    <h2>New Quotation Generated — {quote['id']}</h2>

    <h3>Agent Details</h3>
    <p><b>Name:</b> {agent['name']}<br>
       <b>Email:</b> {agent['email']}<br>
       <b>Date:</b> {quote['generated_at']}</p>

    <h3>Client Details</h3>
    <p><b>Name:</b> {client['policy_holder_name']}<br>
       <b>Email:</b> {client.get('email','—')}<br>
       <b>Phone:</b> {client['phone']}<br>
       <b>KRA PIN:</b> {client['kra_pin']}</p>

    <h3>Vehicle Details</h3>
    <p><b>Reg:</b> {vehicle['vehicle_reg']}<br>
       <b>Make:</b> {vehicle['make']}<br>
       <b>Chassis:</b> {vehicle['chassis_number']}<br>
       <b>Body:</b> {vehicle['vehicle_body_type']}<br>
       <b>Seats:</b> {vehicle['seats']}</p>

    <h3>Insurance Details</h3>
    <p><b>Insurer:</b> {insurance['company']}<br>
       <b>Cover:</b> {insurance['type_of_cover']}<br>
       <b>Certificate:</b> {insurance['type_of_certificate']}<br>
       <b>Commencing:</b> {insurance['commencing_date']}<br>
       <b>Expiry:</b> {insurance['expiry_date']}</p>

    <h3>Quotation</h3>
    <p><b>Quotation No:</b> {quote['id']}<br>
       <b>Base Premium:</b> KES {quote['base_premium']:,.0f}<br>
       <b>Levies & Taxes:</b> KES {quote['levies_and_taxes']:,.0f}<br>
       <b>Total Payable:</b> KES {quote['total_payable']:,.0f}</p>
    """
    attachments = []
    if pdf_bytes:
        attachments.append((f"Quote_{quote['id']}.pdf", pdf_bytes))
    send_email(COMPANY_EMAIL,
               f"Quotation {quote['id']} — {client['policy_holder_name']}",
               html, attachments or None)


def notify_underpayment_attempt(agent, policy_no, quoted_amount, attempted_amount, phone):
    shortfall = quoted_amount - attempted_amount
    html = f"""
    <h2 style="color:#d97706;">⚠️ Underpayment Warning — {policy_no}</h2>
    <p><b>Agent:</b> {agent.get('name','')}<br>
       <b>Agent Email:</b> {agent.get('email','')}<br>
       <b>Policy No:</b> {policy_no}<br>
       <b>Quoted/Due Amount:</b> KES {quoted_amount:,.0f}<br>
       <b>Amount Sent:</b> KES {attempted_amount:,.0f}<br>
       <b>Shortfall:</b> KES {shortfall:,.0f}<br>
       <b>M-Pesa Phone Used:</b> {phone}<br>
       <b>Time:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    <p>This M-Pesa prompt was sent to the client for less than the quoted/outstanding
       balance. The payment was NOT blocked — this is a notification only. This
       agent's account has been flagged for admin review (see Agents in the admin
       dashboard).</p>
    """
    send_email(COMPANY_EMAIL, f"⚠️ Underpayment Warning — {policy_no}", html)
    log.warning("Underpayment warning sent: policy=%s agent=%s quoted=%s sent=%s",
                policy_no, agent.get('email'), quoted_amount, attempted_amount)


# ─────────────────────────────────────────────────────────────────────────────
# DMVIC CONFIGURATION (Motor certificate issuance — UAT)
#
# FIXED in v2.5: this entire block used to be accidentally indented inside
# notify_underpayment_attempt() above, which made every name in it
# (DMVIC_CERT_TYPES, dmvic_get_token, dmvic_issue_certificate, ...)
# invisible everywhere else in the module — buy_cover() or anything else
# trying to call dmvic_get_token()/dmvic_issue_certificate() would have
# raised NameError. It now lives at module level, its own section, same
# indentation as every other top-level config block in this file (compare
# to the M-PESA DARAJA CONFIGURATION section above).
#
# SECURITY: no hardcoded fallback credentials in production, same pattern as
# M-Pesa above. Cert/key are mTLS client credentials issued by DMVIC.
# ─────────────────────────────────────────────────────────────────────────────

DMVIC_ENV          = os.environ.get('DMVIC_ENV', 'uat')
DMVIC_BASE_URL      = os.environ.get('DMVIC_BASE_URL', 'https://uat-api.dmvic.com')
DMVIC_API_USERNAME  = require_env('DMVIC_API_USERNAME', allow_dev_fallback='')
DMVIC_API_PASSWORD  = require_env('DMVIC_API_PASSWORD', allow_dev_fallback='')
DMVIC_CLIENT_ID     = require_env('DMVIC_CLIENT_ID', allow_dev_fallback='')
DMVIC_IRA_NUMBER    = os.environ.get('DMVIC_IRA_NUMBER', '')
DMVIC_CERT_PATH     = os.environ.get('DMVIC_CERT_PATH', 'backend/certs/dmvic/dmvic_uat_cert.pem')
DMVIC_KEY_PATH      = os.environ.get('DMVIC_KEY_PATH', 'backend/certs/dmvic/dmvic_uat_key.pem')

def dmvic_cert_tuple():
    """requests' cert= param expects (certfile, keyfile) for mTLS.
    Resolves via DMVIC_CERT_PATH/DMVIC_KEY_PATH (which default to the UAT
    cert/key, matching DMVIC_ENV/DMVIC_BASE_URL's UAT default) so the cert
    pair always matches whichever DMVIC environment we're actually calling.
    Returns None if the files aren't present so callers can fail loudly
    instead of requests silently sending no client cert."""
    cert_path = DMVIC_CERT_PATH if os.path.isabs(DMVIC_CERT_PATH) else os.path.join(BASE_DIR, DMVIC_CERT_PATH)
    key_path  = DMVIC_KEY_PATH if os.path.isabs(DMVIC_KEY_PATH) else os.path.join(BASE_DIR, DMVIC_KEY_PATH)

    if os.path.exists(cert_path) and os.path.exists(key_path):
        return (cert_path, key_path)

    cert_pem = os.environ.get('DMVIC_CERT_PEM', '')
    key_pem = os.environ.get('DMVIC_KEY_PEM', '')
    if cert_pem and key_pem:
        try:
            pem_dir = '/tmp/westlake-dmvic'
            os.makedirs(pem_dir, mode=0o700, exist_ok=True)
            cert_path = os.path.join(pem_dir, 'client-cert.pem')
            key_path = os.path.join(pem_dir, 'client-key.pem')
            for path, pem in ((cert_path, cert_pem), (key_path, key_pem)):
                with open(path, 'w', encoding='utf-8') as pem_file:
                    pem_file.write(pem)
                os.chmod(path, 0o600)
            return (cert_path, key_path)
        except OSError:
            log.error("Could not prepare DMVIC client certificate files")

    log.warning("DMVIC mTLS cert or key missing at: %s / %s", cert_path, key_path)
    return None

# Confirmed against DMVIC's Intermediary Issuance API doc (section 4.12.1.1.1).
# NOTE: DMVIC support confirmed (2026-07-11) that intermediaries are ONLY
# permitted to issue Type A certificates for these two vehicle categories —
# Type A Bus and Type A Matatu are NOT issuable via the Intermediary API,
# even though those codes exist in the Member Company API docs.
DMVIC_CERT_TYPES = {
    'psv_unmarked': 1, 'type_a_taxi': 8,
}
# Cover type codes for the INTERMEDIARY endpoint are 100/200/300 — these are
# DIFFERENT from the Member Company endpoint's 1/2/3 codes. Confirmed against
# DMVIC's Intermediary doc (section 4.12.1.1.1).
DMVIC_COVER_TYPES = {
    'comprehensive': 100, 'third_party_only': 200, 'third_party_fire_theft': 300,
}
# Type B (Motor Commercial) VehicleType codes — confirmed against DMVIC's
# Intermediary Issuance API doc v1.8.2, section 4.12.2.
DMVIC_VEHICLE_TYPE_B = {
    'own_goods': 1, 'general_cartage': 2, 'institutional': 3,
    'special_vehicles': 4, 'tankers': 5, 'motor_trade': 6,
}
# MemberCompanyID per insurer, confirmed by DMVIC support:
#   Definite   = 49  (2026-07-11)
#   Monarch    = 43
#   Directline = 18
# Each can still be overridden via env var if DMVIC ever reissues IDs.
DMVIC_MEMBER_COMPANY_IDS = {
    'definite':   int(os.environ.get('DMVIC_MEMBER_COMPANY_ID_DEFINITE', 49)),
    'monarch':    int(os.environ.get('DMVIC_MEMBER_COMPANY_ID_MONARCH', 43)),
    'directline': int(os.environ.get('DMVIC_MEMBER_COMPANY_ID_DIRECTLINE', 18)),
}

def dmvic_member_company_id(company):
    """Resolve the correct DMVIC MemberCompanyID for the insurer this
    quote/policy was written under. Returns None if the company isn't
    mapped yet, so callers can hold the policy as pending_manual instead
    of silently issuing under the wrong insurer's ID."""
    return DMVIC_MEMBER_COMPANY_IDS.get((company or '').lower())

DMVIC_ERROR_MESSAGES = {
    'ER001': 'Certificate request format was invalid.',
    'ER002': 'DMVIC returned an unknown error.',
    'ER003': 'A required field was missing.',
    'ER004': 'One of the submitted values was invalid.',
    'ER005': 'This vehicle already has an active policy — cover cannot start before it expires.',
    'ER006': 'DMVIC has no certificate inventory available right now.',
    'ER007': 'Vehicle details changed from a previously issued certificate — manual review needed.',
}
DMVIC_LOGIN_ERRORS = {
    -2: "DMVIC account password is not set. Please activate your account.",
    -3: "DMVIC username or password is incorrect.",
    -4: "DMVIC account is locked by admin.",
    -5: "DMVIC account is blocked.",
    -6: "DMVIC username doesn't exist.",
    -7: "DMVIC entity is suspended.",
    -8: "DMVIC entity is deactivated.",
}

_dmvic_token_cache = {"token": None, "expires": None}
_dmvic_token_lock = threading.Lock()


def _dmvic_clear_token_cache():
    """Discard a token DMVIC has explicitly rejected.

    DMVIC permits one active token per account.  Retaining a rejected token
    until its advertised expiry makes every later request fail in exactly the
    same way, even after the account itself has been restored.
    """
    _dmvic_token_cache["token"] = None
    _dmvic_token_cache["expires"] = None


def _dmvic_clear_token_cache():
    """Discard a token DMVIC has explicitly rejected.

    DMVIC permits one active token per account.  Retaining a rejected token
    until its advertised expiry makes every later request fail in exactly the
    same way, even after the account itself has been restored.
    """
    _dmvic_token_cache["token"] = None
    _dmvic_token_cache["expires"] = None

# ── Existing DMVIC functions below ──────────────────────────────────────────
def dmvic_get_token(force_refresh=False):
    """POST /api/v1/Account/Login. Caches the token in-process until
    shortly before its stated expiry so we don't log in on every single
    certificate issuance. DMVIC keeps one active token per account, so the
    cache refresh is serialized to prevent concurrent workers from invalidating
    each other's token. Set force_refresh after an ER001 response.
    """
    with _dmvic_token_lock:
        if force_refresh:
            _dmvic_clear_token_cache()

        cached = _dmvic_token_cache.get("token")
        expires = _dmvic_token_cache.get("expires")
        if not force_refresh and cached and expires and datetime.now() < expires - timedelta(minutes=5):
            return cached

        try:
            res = requests.post(
                f"{DMVIC_BASE_URL}/api/v1/Account/Login",
                json={
                    "Username": DMVIC_API_USERNAME,
                    "Password": DMVIC_API_PASSWORD,
                    "ClientID": DMVIC_CLIENT_ID,
                },
                cert=dmvic_cert_tuple(),
                timeout=15,
            )
            data = res.json()
        except Exception as e:
            log.error("DMVIC login error: %s: %s", type(e).__name__, e)
            return None

        if data.get("code") == 1 and data.get("token"):
            _dmvic_token_cache["token"] = data["token"]
            try:
                _dmvic_token_cache["expires"] = datetime.fromisoformat(
                    data["expires"].replace("Z", "+00:00")
                ).replace(tzinfo=None)
            except Exception:
                _dmvic_token_cache["expires"] = datetime.now() + timedelta(hours=1)
            return data["token"]

        code = data.get("code")
        log.error("DMVIC login failed (code=%s): %s", code, DMVIC_LOGIN_ERRORS.get(code, data.get("message")))
        return None


def _dmvic_is_token_error(result):
    if result.get("success"):
        return False
    if "ER001" in (result.get("error_codes") or []):
        return True
    error = (result.get("error") or "").lower()
    return "token" in error and ("expired" in error or "invalid" in error)


def dmvic_issue_with_retry(issue_fn, token, **kwargs):
    """Retry one issuance only when DMVIC explicitly rejects its token."""
    result = issue_fn(token, **kwargs)
    if not _dmvic_is_token_error(result):
        return result

    log.warning("DMVIC issuance returned ER001; refreshing the token for one retry")
    fresh_token = dmvic_get_token(force_refresh=True)
    return issue_fn(fresh_token, **kwargs) if fresh_token else result


def dmvic_issuance_request_id(response_data):
    """Find a DMVIC policy-alert request ID regardless of response nesting.

    DMVIC Support confirmed on 2026-07-24 that a policy alert must be completed
    through /api/v6/Integration/ConfirmCertificateIssuance with the exact
    IssuanceRequestID produced by the original issuance request. Their alert
    payloads are not consistently top-level, so do not discard an alert merely
    because the ID is nested under a callback or error object.
    """
    def walk(value):
        if isinstance(value, dict):
            for key, candidate in value.items():
                normalized = re.sub(r"[^a-z0-9]", "", str(key).lower())
                if normalized == 'issuancerequestid' and candidate:
                    return str(candidate).strip()
            for candidate in value.values():
                found = walk(candidate)
                if found:
                    return found
        elif isinstance(value, list):
            for candidate in value:
                found = walk(candidate)
                if found:
                    return found
        return None

    return walk(response_data)


def dmvic_validate_double_insurance(token, *, policy_start_date, policy_end_date,
                                     vehicle_reg=None, chassis_number=None):
    """Check DMVIC for certificates overlapping a proposed cover period.

    The check is advisory and is deliberately isolated from policy issuance:
    DMVIC or network failures must not make the local policy search fail.
    """
    if not vehicle_reg and not chassis_number:
        return {"success": False, "error": "A registration or chassis number is required."}

    payload = {
        "policystartdate": policy_start_date,
        "policyenddate": policy_end_date,
    }
    if vehicle_reg:
        payload["vehicleregistrationnumber"] = vehicle_reg
    if chassis_number:
        payload["chassisnumber"] = chassis_number

    try:
        res = requests.post(
            f"{DMVIC_BASE_URL}/api/v6/Integration/ValidateDoubleInsurance",
            json=payload,
            headers={
                "Authorization": f"Bearer {token}",
                "ClientID": DMVIC_CLIENT_ID,
            },
            cert=dmvic_cert_tuple(),
            timeout=20,
        )
        data = res.json()
    except Exception as e:
        log.error("DMVIC double-insurance lookup error: %s", type(e).__name__)
        return {"success": False, "error": "Could not reach DMVIC."}

    if data.get("success"):
        matches = (data.get("callbackObj") or {}).get("DoubleInsurance") or []
        return {"success": True, "matches": matches if isinstance(matches, list) else [matches]}

    errors = data.get("Error") or data.get("Errors") or []
    if isinstance(errors, dict):
        errors = [errors]
    # DMVIC uses ER0016 for a clear vehicle lookup, which is a successful
    # business result rather than a technical failure.
    if any((error.get("errorCode") or error.get("code")) == "ER0016" for error in errors):
        return {"success": True, "matches": []}

    error_codes = [error.get("errorCode") or error.get("code") for error in errors]
    messages = [error.get("errorText") or error.get("message") or "" for error in errors]
    log.warning("DMVIC double-insurance lookup failed (%s): %s",
                data.get("APIRequestNumber"), messages)
    return {
        "success": False,
        "error": "; ".join(message for message in messages if message) or "DMVIC lookup failed.",
        "error_codes": error_codes,
    }


def dmvic_confirm_certificate_issuance(token, issuance_request_id, *, is_approved,
                                        is_logbook_verified, is_vehicle_inspected,
                                        additional_comments='', usernames=''):
    """Complete a DMVIC policy-alert issuance after an admin review."""
    payload = {
        "IssuanceRequestID": issuance_request_id,
        "IsApproved": bool(is_approved),
        "IsLogBookVerified": bool(is_logbook_verified),
        "IsVehicleInspected": bool(is_vehicle_inspected),
        "AdditionalComments": additional_comments or "",
        "Usernames": usernames or "",
    }
    try:
        res = requests.post(
            f"{DMVIC_BASE_URL}/api/v6/Integration/ConfirmCertificateIssuance",
            json=payload,
            headers={
                "Authorization": f"Bearer {token}",
                "ClientID": DMVIC_CLIENT_ID,
            },
            cert=dmvic_cert_tuple(),
            timeout=25,
        )
        data = res.json()
    except Exception as e:
        log.error("DMVIC confirm-issuance error: %s", type(e).__name__)
        return {"success": False, "error": "Could not reach DMVIC. Please try again."}

    if data.get("success"):
        txn = (data.get("callbackObj") or {}).get("issueCertificate") or data.get("Transaction", {})
        return {
            "success": True,
            "transaction_no": txn.get("TransactionNo"),
            "certificate_no": txn.get("actualCNo") or txn.get("CertificateNo"),
            "api_request_number": data.get("APIRequestNumber"),
        }

    errors = data.get("Errors") or data.get("Error") or []
    if isinstance(errors, dict):
        errors = [errors]
    messages = []
    for error in errors:
        code = error.get("code") or error.get("errorCode", "")
        message = error.get("message") or error.get("errorText", "")
        messages.append(DMVIC_ERROR_MESSAGES.get(code, message) or message)
    log.warning("DMVIC confirm-issuance failed (%s): %s", data.get("APIRequestNumber"), messages)
    return {
        "success": False,
        "error": "; ".join(messages) or "Certificate confirmation failed.",
        "error_codes": [error.get("code") or error.get("errorCode") for error in errors],
    }
def dmvic_issue_certificate(token, *, member_company_id, cert_type, cover_type, policyholder, policy_number,
                             commencing_date, expiring_date, chassis_number, phone_number,
                             body_type, licensed_to_carry, email, insured_pin,
                             year_of_registration=None, registration_number=None, vehicle_make=None, vehicle_model=None,
                             engine_number=None, sum_insured=None, year_of_manufacture=None,
                             huduma_number=None, color=None, passenger_capacity=None,
                             cubic_capacity=None):
    """POST /api/v7/IntermediaryIntegration/IssuanceTypeACertificate
    Dates must already be formatted DD/MM/YYYY before calling this."""
    payload = {
        "MemberCompanyID": member_company_id,   # was: DMVIC_MEMBER_COMPANY_ID
        "TypeOfCertificate": cert_type,
        "Typeofcover": cover_type,
        "Policyholder": policyholder,
        "policynumber": policy_number,
        "Commencingdate": commencing_date,
        "Expiringdate": expiring_date,
        "Registrationnumber": registration_number or "",
        "Chassisnumber": chassis_number,
        "Phonenumber": phone_number,
        "Bodytype": body_type,
        "Licensedtocarry": licensed_to_carry,
        "Email": email,
        "InsuredPIN": insured_pin,
        # DMVIC marks this mandatory (format YYYY); fall back to year_of_manufacture
        # if a distinct registration year wasn't supplied, since in practice they're
        # usually the same for a new policy.
        "Yearofregistration": year_of_registration or year_of_manufacture,
    }
    if sum_insured is not None:
        payload["SumInsured"] = sum_insured
    if vehicle_make and vehicle_model:
        payload["Vehiclemake"] = vehicle_make
        payload["Vehiclemodel"] = vehicle_model
    if engine_number:
        payload["Enginenumber"] = engine_number
    if year_of_manufacture is not None:
        payload["Yearofmanufacture"] = year_of_manufacture
    if huduma_number:
        payload["HudumaNumber"] = huduma_number
    if color:
        payload["color"] = color
    if passenger_capacity is not None:
        payload["passengerCapacity"] = passenger_capacity
    if cubic_capacity is not None:
        payload["cubicCapacity"] = cubic_capacity

    try:
        res = requests.post(
            f"{DMVIC_BASE_URL}/api/v7/IntermediaryIntegration/IssuanceTypeACertificate",
            json=payload,
            headers={
                "Authorization": f"Bearer {token}",
                "ClientID": DMVIC_CLIENT_ID,
            },
            cert=dmvic_cert_tuple(),
            timeout=25,
        )
        data = res.json()
    except Exception as e:
        log.error("DMVIC issuance error: %s", type(e).__name__)
        return {"success": False, "error": "Could not reach DMVIC. Please try again."}

    if data.get("success"):
        # Confirmed via live UAT test (2026-07-11, txn UAT-TAE4170): the
        # Intermediary endpoint nests results under callbackObj.issueCertificate,
        # using actualCNo for the certificate number — NOT a top-level
        # "Transaction" object with "CertificateNo" as the Member Company docs
        # describe. Keeping both lookups so this degrades gracefully if DMVIC's
        # response shape ever changes back.
        txn = (data.get("callbackObj") or {}).get("issueCertificate") or data.get("Transaction", {})
        return {
            "success": True,
            "transaction_no": txn.get("TransactionNo"),
            "certificate_no": txn.get("actualCNo") or txn.get("CertificateNo"),
            "api_request_number": data.get("APIRequestNumber"),
        }

    errors = data.get("Errors") or data.get("Error") or []
    messages = []
    for err in errors:
        code = err.get("code") or err.get("errorCode", "")
        msg  = err.get("message") or err.get("errorText", "")
        messages.append(DMVIC_ERROR_MESSAGES.get(code, msg) or msg)

    log.warning("DMVIC issuance failed (%s): %s | raw_errors=%s | payload_sent=%s",
                data.get("APIRequestNumber"), messages, errors, payload)
    return {
        "success": False,
        "error": "; ".join(messages) or "Certificate issuance failed.",
        "error_codes": [e.get("code") or e.get("errorCode") for e in errors],
        "issuance_request_id": dmvic_issuance_request_id(data),
    }


def dmvic_issue_certificate_type_b(token, *, member_company_id, cover_type, vehicle_type, policyholder, policy_number,
                                    commencing_date, expiring_date, chassis_number, phone_number,
                                    body_type, tonnage_carrying_capacity, licensed_to_carry, email, insured_pin,
                                    year_of_registration, registration_number=None, vehicle_make=None,
                                    vehicle_model=None, engine_number=None, sum_insured=None,
                                    year_of_manufacture=None, huduma_number=None, color=None,
                                    passenger_capacity=None, cubic_capacity=None):
    """POST /api/v7/IntermediaryIntegration/IssuanceTypeBCertificate (Motor Commercial)
    Confirmed against DMVIC's Intermediary Issuance API doc v1.8.2, section 4.12.2.
    Dates must already be formatted DD/MM/YYYY before calling this.
    Yearofregistration and Tonnagecarryingcapacity are mandatory for this endpoint
    (unlike Type A, where Yearofregistration falls back to year_of_manufacture)."""
    payload = {
        "MemberCompanyID": member_company_id,   # was: DMVIC_MEMBER_COMPANY_ID
        "Typeofcover": cover_type,
        "VehicleType": vehicle_type,
        "Policyholder": policyholder,
        "policynumber": policy_number,
        "Commencingdate": commencing_date,
        "Expiringdate": expiring_date,
        "Registrationnumber": registration_number or "",
        "Chassisnumber": chassis_number,
        "Phonenumber": phone_number,
        "Bodytype": body_type,
        # DMVIC's parameter table labels this "TonnageCarryingCapacity", but its own
        # worked request/response examples (see spec 4.12.2.1.1 / 4.12.2.3.1) use the
        # field name "Tonnage" — sending the table's name gets silently dropped and
        # DMVIC reports it as missing (ER003 "Tonnage is required") even with a value set.
        "Tonnage": tonnage_carrying_capacity,
        "Licensedtocarry": licensed_to_carry,
        "Email": email,
        "InsuredPIN": insured_pin,
        "Yearofregistration": year_of_registration,
    }
    if sum_insured is not None:
        payload["SumInsured"] = sum_insured
    if vehicle_make and vehicle_model:
        payload["Vehiclemake"] = vehicle_make
        payload["Vehiclemodel"] = vehicle_model
    if engine_number:
        payload["Enginenumber"] = engine_number
    if year_of_manufacture is not None:
        payload["Yearofmanufacture"] = year_of_manufacture
    if huduma_number:
        payload["HudumaNumber"] = huduma_number
    if color:
        payload["color"] = color
    if passenger_capacity is not None:
        payload["passengerCapacity"] = passenger_capacity
    if cubic_capacity is not None:
        payload["cubicCapacity"] = cubic_capacity

    try:
        res = requests.post(
            f"{DMVIC_BASE_URL}/api/v7/IntermediaryIntegration/IssuanceTypeBCertificate",
            json=payload,
            headers={
                "Authorization": f"Bearer {token}",
                "ClientID": DMVIC_CLIENT_ID,
            },
            cert=dmvic_cert_tuple(),
            timeout=25,
        )
        data = res.json()
    except Exception as e:
        log.error("DMVIC Type B issuance error: %s", type(e).__name__)
        return {"success": False, "error": "Could not reach DMVIC. Please try again."}

    if data.get("success"):
        # Same nesting pattern confirmed for Type A — callbackObj.issueCertificate,
        # actualCNo for the certificate number. Doc's own example response uses
        # this shape too (see section 4.12.2.3.1), so no fallback lookup needed
        # here the way Type A's comment describes, but kept for safety.
        txn = (data.get("callbackObj") or {}).get("issueCertificate") or data.get("Transaction", {})
        return {
            "success": True,
            "transaction_no": txn.get("TransactionNo"),
            "certificate_no": txn.get("actualCNo") or txn.get("CertificateNo"),
            "api_request_number": data.get("APIRequestNumber"),
        }

    errors = data.get("Errors") or data.get("Error") or []
    messages = []
    for err in errors:
        code = err.get("code") or err.get("errorCode", "")
        msg  = err.get("message") or err.get("errorText", "")
        messages.append(DMVIC_ERROR_MESSAGES.get(code, msg) or msg)

    log.warning("DMVIC Type B issuance failed (%s): %s | raw_errors=%s | payload_sent=%s",
                data.get("APIRequestNumber"), messages, errors, payload)
    return {
        "success": False,
        "error": "; ".join(messages) or "Certificate issuance failed.",
        "error_codes": [e.get("code") or e.get("errorCode") for e in errors],
        "issuance_request_id": dmvic_issuance_request_id(data),
    }


def dmvic_issue_certificate_type_c(token, *, member_company_id, cover_type, policyholder, policy_number,
                                    commencing_date, expiring_date, chassis_number, phone_number,
                                    body_type, email, insured_pin, year_of_registration,
                                    registration_number=None, vehicle_make=None, vehicle_model=None,
                                    engine_number=None, sum_insured=None, year_of_manufacture=None,
                                    huduma_number=None, color=None, passenger_capacity=None,
                                    cubic_capacity=None):
    """POST /api/v7/IntermediaryIntegration/IssuanceTypeCCertificate (Motor Private)
    Confirmed against DMVIC's Intermediary Issuance API doc v1.8.2, section 4.12.3.
    Dates must already be formatted DD/MM/YYYY before calling this.
    No VehicleType/Tonnagecarryingcapacity here — those are Type B only.
    Yearofregistration is mandatory (no year_of_manufacture fallback, unlike Type A)."""
    payload = {
        "MemberCompanyID": member_company_id,   # was: "Membercompanyid" — casing bug, now matches Type A/B
        "Typeofcover": cover_type,
        "Policyholder": policyholder,
        "policynumber": policy_number,
        "Commencingdate": commencing_date,
        "Expiringdate": expiring_date,
        "Registrationnumber": registration_number or "",
        "Chassisnumber": chassis_number,
        "Phonenumber": phone_number,
        "Bodytype": body_type,
        "Email": email,
        "InsuredPIN": insured_pin,
        "Yearofregistration": year_of_registration,
    }
    if sum_insured is not None:
        payload["SumInsured"] = sum_insured
    if vehicle_make and vehicle_model:
        payload["Vehiclemake"] = vehicle_make
        payload["Vehiclemodel"] = vehicle_model
    if engine_number:
        payload["Enginenumber"] = engine_number
    if year_of_manufacture is not None:
        payload["Yearofmanufacture"] = year_of_manufacture
    if huduma_number:
        payload["HudumaNumber"] = huduma_number
    if color:
        payload["color"] = color
    if passenger_capacity is not None:
        payload["passengerCapacity"] = passenger_capacity
    if cubic_capacity is not None:
        payload["cubicCapacity"] = cubic_capacity

    try:
        res = requests.post(
            f"{DMVIC_BASE_URL}/api/v7/IntermediaryIntegration/IssuanceTypeCCertificate",
            json=payload,
            headers={
                "Authorization": f"Bearer {token}",
                "ClientID": DMVIC_CLIENT_ID,
            },
            cert=dmvic_cert_tuple(),
            timeout=60,
        )
        data = res.json()
    except Exception as e:
        log.error("DMVIC Type C issuance error: %s", type(e).__name__)
        return {"success": False, "error": "Could not reach DMVIC. Please try again."}

    if data.get("success"):
        txn = (data.get("callbackObj") or {}).get("issueCertificate") or data.get("Transaction", {})
        return {
            "success": True,
            "transaction_no": txn.get("TransactionNo"),
            "certificate_no": txn.get("actualCNo") or txn.get("CertificateNo"),
            "api_request_number": data.get("APIRequestNumber"),
        }

    errors = data.get("Errors") or data.get("Error") or []
    messages = []
    for err in errors:
        code = err.get("code") or err.get("errorCode", "")
        msg  = err.get("message") or err.get("errorText", "")
        messages.append(DMVIC_ERROR_MESSAGES.get(code, msg) or msg)

    # DEBUG: ER003/ER004 map to generic text and don't say WHICH field DMVIC
    # rejected. Log the raw error objects (may contain a field/property name)
    # plus the payload we sent (minus nothing sensitive here) so failures are
    # actually diagnosable instead of just "something was missing/invalid".
    log.warning("DMVIC Type C issuance failed (%s): %s | raw_errors=%s | payload_sent=%s",
                data.get("APIRequestNumber"), messages, errors, payload)
    return {
        "success": False,
        "error": "; ".join(messages) or "Certificate issuance failed.",
        "error_codes": [e.get("code") or e.get("errorCode") for e in errors],
        "issuance_request_id": dmvic_issuance_request_id(data),
    }


# DMVIC's TypeOfCertificate codes for Type D (Motor Cycle), per Intermediary
# Issuance API doc v1.8.2, section 4.12.4.1.1. Only 'motorcycle' (private,
# non-PSV) is currently offered by Westlake — the PSV and Commercial motorcycle
# variants are kept here for completeness but are not wired into any live
# product yet (see issue_dmvic_certificate()'s bucket == 'motorcycle' branch).
DMVIC_CERT_TYPE_D = {
    'motorcycle':            4,   # Type D Motor Cycle (private)
    'motorcycle_psv':        9,   # Type D PSV Motor Cycle — NOT currently offered
    'motorcycle_commercial': 10,  # Type D Motor Cycle Commercial — NOT currently offered
}


def dmvic_issue_certificate_type_d(token, *, member_company_id, type_of_certificate, cover_type, policyholder,
                                    policy_number, commencing_date, expiring_date, chassis_number, phone_number,
                                    body_type, email, insured_pin, year_of_registration,
                                    licensed_to_carry=None, tonnage=None,
                                    registration_number=None, vehicle_make=None, vehicle_model=None,
                                    engine_number=None, sum_insured=None, year_of_manufacture=None,
                                    huduma_number=None, color=None, passenger_capacity=None,
                                    cubic_capacity=None):
    """POST /api/v7/IntermediaryIntegration/IssuanceTypeDCertificate (Motor Cycle)
    Confirmed against DMVIC's Intermediary Issuance API doc v1.8.2, section 4.12.4.
    Dates must already be formatted DD/MM/YYYY before calling this.
    Per the spec, Licensedtocarry is mandatory for TypeOfCertificate 4 (Motor Cycle)
    and 9 (PSV Motor Cycle); Tonnage is mandatory for TypeOfCertificate 10 (Motor
    Cycle Commercial) instead — callers pass whichever one applies and leave the
    other as None. Yearofregistration is mandatory with no fallback, same as Type C."""
    payload = {
        "TypeOfCertificate": type_of_certificate,
        "MemberCompanyID": member_company_id,
        "Typeofcover": cover_type,
        "Policyholder": policyholder,
        "policynumber": policy_number,
        "Commencingdate": commencing_date,
        "Expiringdate": expiring_date,
        "Registrationnumber": registration_number or "",
        "Chassisnumber": chassis_number,
        "Phonenumber": phone_number,
        "Bodytype": body_type,
        "Email": email,
        "InsuredPIN": insured_pin,
        "Yearofregistration": year_of_registration,
    }
    if licensed_to_carry is not None:
        payload["Licensedtocarry"] = licensed_to_carry
    if vehicle_make and vehicle_model:
        payload["Vehiclemake"] = vehicle_make
        payload["Vehiclemodel"] = vehicle_model
    if engine_number:
        payload["Enginenumber"] = engine_number
    if tonnage is not None:
        payload["Tonnage"] = tonnage
    if sum_insured is not None:
        payload["SumInsured"] = sum_insured
    if year_of_manufacture is not None:
        payload["Yearofmanufacture"] = year_of_manufacture
    if huduma_number:
        payload["HudumaNumber"] = huduma_number
    if color:
        payload["color"] = color
    if passenger_capacity is not None:
        payload["passengerCapacity"] = passenger_capacity
    if cubic_capacity is not None:
        payload["cubicCapacity"] = cubic_capacity

    try:
        res = requests.post(
            f"{DMVIC_BASE_URL}/api/v7/IntermediaryIntegration/IssuanceTypeDCertificate",
            json=payload,
            headers={
                "Authorization": f"Bearer {token}",
                "ClientID": DMVIC_CLIENT_ID,
            },
            cert=dmvic_cert_tuple(),
            timeout=60,
        )
        data = res.json()
    except Exception as e:
        log.error("DMVIC Type D issuance error: %s", type(e).__name__)
        return {"success": False, "error": "Could not reach DMVIC. Please try again."}

    if data.get("success"):
        # Same callbackObj.issueCertificate / actualCNo nesting confirmed for
        # Type D's own sample success response (section 4.12.4.3.1).
        txn = (data.get("callbackObj") or {}).get("issueCertificate") or data.get("Transaction", {})
        return {
            "success": True,
            "transaction_no": txn.get("TransactionNo"),
            "certificate_no": txn.get("actualCNo") or txn.get("CertificateNo"),
            "api_request_number": data.get("APIRequestNumber"),
        }

    errors = data.get("Errors") or data.get("Error") or []
    messages = []
    for err in errors:
        code = err.get("code") or err.get("errorCode", "")
        msg  = err.get("message") or err.get("errorText", "")
        messages.append(DMVIC_ERROR_MESSAGES.get(code, msg) or msg)

    log.warning("DMVIC Type D issuance failed (%s): %s | raw_errors=%s | payload_sent=%s",
                data.get("APIRequestNumber"), messages, errors, payload)
    return {
        "success": False,
        "error": "; ".join(messages) or "Certificate issuance failed.",
        "error_codes": [e.get("code") or e.get("errorCode") for e in errors],
        "issuance_request_id": dmvic_issuance_request_id(data),
    }


def _dmvic_fmt_phone(p):
    """DMVIC's Intermediary API wants a bare 9-digit Kenyan number (no
    leading 0, no +254/254 country code) — e.g. '794399514'. Strips
    whatever format the agent typed it in (0794399514, +254794399514,
    254794399514, with spaces/dashes) down to that."""
    if not p:
        return ''
    digits = ''.join(ch for ch in str(p) if ch.isdigit())
    if digits.startswith('254') and len(digits) == 12:
        digits = digits[3:]
    elif digits.startswith('0') and len(digits) == 10:
        digits = digits[1:]
    return digits


def _dmvic_fmt_date(d):
    if not d:
        return ''
    
    # Handle date or datetime objects
    if isinstance(d, (date, datetime)):
        return d.strftime('%d/%m/%Y')
        
    if isinstance(d, str):
        d_str = d.strip()
        # Handle full datetime ISO strings like "2026-07-21 16:30:00"
        if ' ' in d_str:
            d_str = d_str.split(' ')[0]
            
        # If it's already DD/MM/YYYY
        if '/' in d_str and len(d_str.split('/')[2]) == 4:
            return d_str

        # Try standard YYYY-MM-DD format
        try:
            return datetime.strptime(d_str, '%Y-%m-%d').strftime('%d/%m/%Y')
        except ValueError:
            log.warning("Could not parse DMVIC date string: %s", d)
            return d_str

    return str(d)


def dmvic_vehicle_identity(quote_row):
    """Return the vehicle identifiers that are safe to send to DMVIC.

    Registration and chassis are required for the quotation. Engine, make and
    model are optional in the user workflow. DMVIC compares supplied values
    with previous certificates, so partial or guessed optional values must be
    omitted rather than represented as empty strings or an invented split.
    """
    identity = {
        "registration_number": str(quote_row.get("vehicle_reg") or "").strip().upper(),
        "chassis_number": str(quote_row.get("chassis_number") or "").strip().upper(),
        "engine_number": str(quote_row.get("engine_number") or "").strip().upper(),
        "vehicle_make": str(quote_row.get("vehicle_make") or "").strip(),
        "vehicle_model": str(quote_row.get("vehicle_model") or "").strip(),
    }
    labels = {
        "registration_number": "registration number",
        "chassis_number": "chassis number",
    }
    missing = [labels[key] for key in labels if not identity[key]]

    # DMVIC expects make and model to be separate values. Never treat the old
    # combined display field as a reliable source, and omit either field unless
    # the agent supplied both exact logbook values.
    if not (identity["vehicle_make"] and identity["vehicle_model"]):
        identity["vehicle_make"] = None
        identity["vehicle_model"] = None
    if not identity["engine_number"]:
        identity["engine_number"] = None
    return identity, missing


def issue_dmvic_certificate(policy_no, quote_row):
    """
    Attempts DMVIC certificate issuance for a newly-created policy. Called
    from buy_cover() on a background worker so the agent isn't blocked
    waiting on DMVIC's response.

    Confirmed with the business (2026-07-22): four product categories are
    now issued through DMVIC's Intermediary endpoints —
      - PSV (tuktuk_psv, motorcycle_psv only — NOT bus/matatu)  -> Type A
      - Motor Commercial                                        -> Type B
      - Motor Private                                           -> Type C
      - Motorcycle, private/non-PSV only ('motorcycle' product) -> Type D
    motorcycle_psv is intentionally NOT routed through Type D — Westlake
    doesn't offer PSV motorcycle cover, so that product is held for manual
    review instead (see the bucket == 'motorcycle' branch below).

    Requires the migration in migrations_add_dmvic_columns.sql to have
    been run (adds dmvic_status/dmvic_transaction_no/dmvic_certificate_no/
    dmvic_api_request_no/dmvic_cert_type/dmvic_error/dmvic_issued_at to
    the policies table) — every UPDATE below will fail with an
    unknown-column error until that migration is applied.
    """
    bucket = PRODUCT_TO_CERT_TYPE.get(quote_row.get('product'))
    product = quote_row.get('product')
    company = quote_row.get('company')

    identity, missing_identity = dmvic_vehicle_identity(quote_row)
    if missing_identity:
        msg = ("DMVIC issuance held: verify and re-quote the "
               f"{', '.join(missing_identity)} from the logbook. DMVIC compares "
               "these values with previously issued certificates; the legacy "
               "combined Make / Model field is not sent as a substitute.")
        log.warning("DMVIC identity preflight held %s: %s", policy_no, missing_identity)
        query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                  WHERE policy_no=%s""", (msg, policy_no), commit=True)
        return

    identity, missing_identity = dmvic_vehicle_identity(quote_row)
    if missing_identity:
        msg = ("DMVIC issuance held: verify and re-quote the "
               f"{', '.join(missing_identity)} from the logbook. DMVIC compares "
               "these values with previously issued certificates; the legacy "
               "combined Make / Model field is not sent as a substitute.")
        log.warning("DMVIC identity preflight held %s: %s", policy_no, missing_identity)
        query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                  WHERE policy_no=%s""", (msg, policy_no), commit=True)
        return

    member_company_id = dmvic_member_company_id(company)
    if member_company_id is None:
        msg = (f"DMVIC issuance held: no MemberCompanyID mapped for insurer "
               f"'{company}'. Add it to DMVIC_MEMBER_COMPANY_IDS.")
        log.warning(msg)
        query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                  WHERE policy_no=%s""", (msg, policy_no), commit=True)
        return

    cover_type = DMVIC_COVER_TYPES.get(quote_row.get('type_of_cover'))
    if cover_type is None:
        err = f"No DMVIC cover-type mapping for '{quote_row.get('type_of_cover')}'"
        log.error("DMVIC issuance skipped for %s: %s", policy_no, err)
        query("""UPDATE policies SET dmvic_status='failed', dmvic_error=%s
                  WHERE policy_no=%s""", (err, policy_no), commit=True)
        return

    # ── PSV -> Type A ───────────────────────────────────────────────────
    if bucket == 'psv':
        # DMVIC support confirmed (2026-07-11) that this Intermediary account
        # can ONLY issue 'psv_unmarked' (1) or 'type_a_taxi' (8) — Matatu and
        # Bus are NOT issuable via this endpoint at all. Business confirmed
        # (2026-07-21) that only tuktuk_psv/motorcycle_psv are actually
        # issued anyway, so the old broader 'psv' bucket products (bus,
        # matatu, asset-only variants) are out of scope entirely.
        PSV_TAXI_PRODUCTS = set()   # none of Westlake's PSV products are taxis
        PSV_UNMARKED_PRODUCTS = {'tuktuk_psv', 'motorcycle_psv'}

        if product in PSV_TAXI_PRODUCTS:
            cert_type = DMVIC_CERT_TYPES.get('type_a_taxi', 8)
        elif product in PSV_UNMARKED_PRODUCTS:
            cert_type = DMVIC_CERT_TYPES.get('psv_unmarked', 1)
        else:
            msg = (f"DMVIC issuance held: product '{product}' is bucketed as PSV "
                   f"but isn't tuktuk_psv or motorcycle_psv — only those two are "
                   f"confirmed issuable. Needs manual issuance or confirmation.")
            log.warning(msg)
            query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                      WHERE policy_no=%s""", (msg, policy_no), commit=True)
            return

        token = dmvic_get_token()
        if not token:
            query("""UPDATE policies SET dmvic_status='failed',
                      dmvic_error='Could not obtain DMVIC auth token'
                      WHERE policy_no=%s""", (policy_no,), commit=True)
            return

        query("""UPDATE policies SET dmvic_status='pending', dmvic_cert_type=%s
                  WHERE policy_no=%s""", (str(cert_type), policy_no), commit=True)

        result = dmvic_issue_with_retry(
            dmvic_issue_certificate,
            token,
            member_company_id=member_company_id,
            cert_type=cert_type,
            cover_type=cover_type,
            policyholder=quote_row.get('policy_holder_name', ''),
            policy_number=policy_no,
            commencing_date=_dmvic_fmt_date(quote_row.get('commencing_date')),
            expiring_date=_dmvic_fmt_date(quote_row.get('expiry_date')),
            chassis_number=identity['chassis_number'],
            phone_number=_dmvic_fmt_phone(quote_row.get('phone', '')),
            body_type=quote_row.get('vehicle_body_type', ''),
            licensed_to_carry=quote_row.get('seats', 0),
            email=quote_row.get('email', '') or '',
            insured_pin=quote_row.get('kra_pin', ''),
            registration_number=identity['registration_number'],
            vehicle_make=identity['vehicle_make'],
            vehicle_model=identity['vehicle_model'],
            engine_number=identity['engine_number'],
            sum_insured=float(quote_row.get('vehicle_value') or 0) or None,
            year_of_manufacture=quote_row.get('year_of_manufacture') or None,
            year_of_registration=quote_row.get('year_of_registration') or None,
        )

    # ── Motor Commercial -> Type B ──────────────────────────────────────
    elif bucket == 'commercial':
        # commercial_hybrid is deliberately absent from PRODUCT_TO_VEHICLE_TYPE_B
        # (see dmvic_mapping.py docstring) — it's Definite's combined product and
        # has to be split into own_goods vs general_cartage using the sub_type
        # captured on the quotation, same distinction calculate_premium()'s
        # get_definite_tp_commercial() makes for pricing. Resolve it here before
        # falling through to the normal per-product lookup below.
        #
        # UNCONFIRMED: 'prime_mover' sub_type is mapped to general_cartage as a
        # best guess (a prime mover hauling a trailer is closer to "general
        # cartage" than "own goods" in DMVIC's categorisation) — this has not
        # been confirmed with DMVIC and should be verified before relying on it.
        if product == 'commercial_hybrid':
            sub_type = quote_row.get('sub_type')
            HYBRID_SUBTYPE_TO_VEHICLE_B = {
                'own_goods':      'own_goods',
                'general_cartage': 'general_cartage',
                'prime_mover':     'general_cartage',   # UNCONFIRMED — best guess, see note above
            }
            resolved = HYBRID_SUBTYPE_TO_VEHICLE_B.get(sub_type)
            if resolved is None:
                msg = (f"DMVIC issuance held: commercial_hybrid requires sub_type "
                       f"'own_goods', 'general_cartage', or 'prime_mover' to resolve its "
                       f"DMVIC VehicleType, but got '{sub_type}'. Re-quote with a sub_type set.")
                log.warning(msg)
                query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                          WHERE policy_no=%s""", (msg, policy_no), commit=True)
                return
            vehicle_type = DMVIC_VEHICLE_TYPE_B.get(resolved)
        else:
            vehicle_type = DMVIC_VEHICLE_TYPE_B.get(PRODUCT_TO_VEHICLE_TYPE_B.get(product))

        if vehicle_type is None:
            msg = (f"DMVIC issuance held: product '{product}' has no confirmed "
                   f"VehicleType mapping for Type B — see dmvic_mapping.py.")
            log.warning(msg)
            query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                      WHERE policy_no=%s""", (msg, policy_no), commit=True)
            return

        year_of_registration = quote_row.get('year_of_registration') or quote_row.get('year_of_manufacture')
        if not year_of_registration:
            msg = "DMVIC issuance held: Yearofregistration is mandatory for Type B and neither year_of_registration nor year_of_manufacture was set."
            log.warning(msg)
            query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                      WHERE policy_no=%s""", (msg, policy_no), commit=True)
            return

        token = dmvic_get_token()
        if not token:
            query("""UPDATE policies SET dmvic_status='failed',
                      dmvic_error='Could not obtain DMVIC auth token'
                      WHERE policy_no=%s""", (policy_no,), commit=True)
            return

        query("""UPDATE policies SET dmvic_status='pending', dmvic_cert_type=%s
                  WHERE policy_no=%s""", (f"type_b_{vehicle_type}", policy_no), commit=True)

        result = dmvic_issue_with_retry(
            dmvic_issue_certificate_type_b,
            token,
            member_company_id=member_company_id,
            cover_type=cover_type,
            vehicle_type=vehicle_type,
            policyholder=quote_row.get('policy_holder_name', ''),
            policy_number=policy_no,
            commencing_date=_dmvic_fmt_date(quote_row.get('commencing_date')),
            expiring_date=_dmvic_fmt_date(quote_row.get('expiry_date')),
            chassis_number=identity['chassis_number'],
            phone_number=_dmvic_fmt_phone(quote_row.get('phone', '')),
            body_type=quote_row.get('vehicle_body_type', ''),
            tonnage_carrying_capacity=float(quote_row.get('tonnage') or 0),
            licensed_to_carry=quote_row.get('seats', 0),
            email=quote_row.get('email', '') or '',
            insured_pin=quote_row.get('kra_pin', ''),
            year_of_registration=year_of_registration,
            registration_number=identity['registration_number'],
            vehicle_make=identity['vehicle_make'],
            vehicle_model=identity['vehicle_model'],
            engine_number=identity['engine_number'],
            sum_insured=float(quote_row.get('vehicle_value') or 0) or None,
            year_of_manufacture=quote_row.get('year_of_manufacture') or None,
        )

    # ── Motor Private -> Type C ─────────────────────────────────────────
    elif bucket == 'general':
        year_of_registration = quote_row.get('year_of_registration') or quote_row.get('year_of_manufacture')
        if not year_of_registration:
            msg = "DMVIC issuance held: Yearofregistration is mandatory for Type C and neither year_of_registration nor year_of_manufacture was set."
            log.warning(msg)
            query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                      WHERE policy_no=%s""", (msg, policy_no), commit=True)
            return

        token = dmvic_get_token()
        if not token:
            query("""UPDATE policies SET dmvic_status='failed',
                      dmvic_error='Could not obtain DMVIC auth token'
                      WHERE policy_no=%s""", (policy_no,), commit=True)
            return

        query("""UPDATE policies SET dmvic_status='pending', dmvic_cert_type='type_c'
                  WHERE policy_no=%s""", (policy_no,), commit=True)

        result = dmvic_issue_with_retry(
            dmvic_issue_certificate_type_c,
            token,
            member_company_id=member_company_id,
            cover_type=cover_type,
            policyholder=quote_row.get('policy_holder_name', ''),
            policy_number=policy_no,
            commencing_date=_dmvic_fmt_date(quote_row.get('commencing_date')),
            expiring_date=_dmvic_fmt_date(quote_row.get('expiry_date')),
            chassis_number=identity['chassis_number'],
            phone_number=_dmvic_fmt_phone(quote_row.get('phone', '')),
            body_type=quote_row.get('vehicle_body_type', ''),
            email=quote_row.get('email', '') or '',
            insured_pin=quote_row.get('kra_pin', ''),
            year_of_registration=year_of_registration,
            registration_number=identity['registration_number'],
            vehicle_make=identity['vehicle_make'],
            vehicle_model=identity['vehicle_model'],
            engine_number=identity['engine_number'],
            sum_insured=float(quote_row.get('vehicle_value') or 0) or None,
            year_of_manufacture=quote_row.get('year_of_manufacture') or None,
        )

    # ── Motorcycle -> Type D ─────────────────────────────────────────────
    # Confirmed with the business (2026-07-22): only private (non-PSV) motorcycle
    # cover is offered — motorcycle_psv stays held for manual issuance since
    # Westlake doesn't currently do PSV. Falls into this branch either via
    # dmvic_mapping.py's PRODUCT_TO_CERT_TYPE bucket ('motorcycle'), or as a
    # fallback directly on product name in case that mapping file hasn't been
    # updated yet — bucket is None until it is.
    elif bucket == 'motorcycle' or (bucket is None and product in DMVIC_CERT_TYPE_D):
        if product == 'motorcycle':
            type_of_certificate = DMVIC_CERT_TYPE_D['motorcycle']
        else:
            msg = (f"DMVIC issuance held: product '{product}' is a motorcycle variant "
                   f"(PSV/Commercial) that Westlake doesn't currently offer — only plain "
                   f"'motorcycle' (private, Type D code 4) is wired up.")
            log.warning(msg)
            query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                      WHERE policy_no=%s""", (msg, policy_no), commit=True)
            return

        year_of_registration = quote_row.get('year_of_registration') or quote_row.get('year_of_manufacture')
        if not year_of_registration:
            msg = "DMVIC issuance held: Yearofregistration is mandatory for Type D and neither year_of_registration nor year_of_manufacture was set."
            log.warning(msg)
            query("""UPDATE policies SET dmvic_status='pending_manual', dmvic_error=%s
                      WHERE policy_no=%s""", (msg, policy_no), commit=True)
            return

        token = dmvic_get_token()
        if not token:
            query("""UPDATE policies SET dmvic_status='failed',
                      dmvic_error='Could not obtain DMVIC auth token'
                      WHERE policy_no=%s""", (policy_no,), commit=True)
            return

        query("""UPDATE policies SET dmvic_status='pending', dmvic_cert_type='type_d'
                  WHERE policy_no=%s""", (policy_no,), commit=True)

        result = dmvic_issue_with_retry(
            dmvic_issue_certificate_type_d,
            token,
            member_company_id=member_company_id,
            type_of_certificate=type_of_certificate,
            cover_type=cover_type,
            policyholder=quote_row.get('policy_holder_name', ''),
            policy_number=policy_no,
            commencing_date=_dmvic_fmt_date(quote_row.get('commencing_date')),
            expiring_date=_dmvic_fmt_date(quote_row.get('expiry_date')),
            chassis_number=identity['chassis_number'],
            phone_number=_dmvic_fmt_phone(quote_row.get('phone', '')),
            body_type=quote_row.get('vehicle_body_type', ''),
            # Licensedtocarry is mandatory for TypeOfCertificate 4/9 (this branch);
            # Tonnage (mandatory for code 10, Motor Cycle Commercial) isn't applicable
            # here since that variant isn't wired up — left as None.
            licensed_to_carry=quote_row.get('seats', 0),
            email=quote_row.get('email', '') or '',
            insured_pin=quote_row.get('kra_pin', ''),
            year_of_registration=year_of_registration,
            registration_number=identity['registration_number'],
            vehicle_make=identity['vehicle_make'],
            vehicle_model=identity['vehicle_model'],
            engine_number=identity['engine_number'],
            sum_insured=float(quote_row.get('vehicle_value') or 0) or None,
            year_of_manufacture=quote_row.get('year_of_manufacture') or None,
        )

    # ── Anything else (still-unmapped products) ──────────────────────────
    else:
        msg = (f"DMVIC issuance skipped: product '{product}' isn't PSV/Commercial/"
               f"Private (bucket={bucket}) — no endpoint wired up for this yet.")
        log.info(msg)
        query("""UPDATE policies SET dmvic_status='unsupported', dmvic_error=%s
                  WHERE policy_no=%s""", (msg, policy_no), commit=True)
        return

    if result.get('success'):
        query("""UPDATE policies
                  SET dmvic_status='issued',
                      dmvic_transaction_no=%s,
                      dmvic_certificate_no=%s,
                      dmvic_api_request_no=%s,
                      dmvic_issued_at=NOW(),
                      dmvic_error=NULL
                  WHERE policy_no=%s""",
              (result.get('transaction_no'), result.get('certificate_no'),
               result.get('api_request_number'), policy_no), commit=True)
        log.info("DMVIC certificate issued for %s: %s",
                  policy_no, result.get('certificate_no'))
    else:
        # DMVIC Support confirmed on 2026-07-25 that policy alerts are produced
        # by its record comparison and that this account has no manual policy
        # confirmation path.  An IssuanceRequestID is therefore retained for
        # traceability only; it must not be presented as an approval bypass.
        issuance_request_id = result.get('issuance_request_id')
        error_codes = {str(code).upper() for code in (result.get('error_codes') or []) if code}
        is_policy_alert = bool(issuance_request_id or error_codes.intersection({'ER005', 'ER007'}))
        if is_policy_alert:
            reason = result.get('error', 'DMVIC policy alert')
            msg = (f"DMVIC data review required: {reason} Verify the certificate type, "
                   "engine number, vehicle make and vehicle model against the logbook and "
                   "the insurer's existing record, then re-quote. DMVIC does not provide a "
                   "manual confirmation route for this alert.")
            query("""UPDATE policies
                      SET dmvic_status='pending_manual',
                          dmvic_issuance_request_id=%s,
                          dmvic_error=%s
                      WHERE policy_no=%s""",
                  (issuance_request_id, msg, policy_no), commit=True)
            log.warning("DMVIC policy alert for %s requires data correction; codes=%s", policy_no, error_codes)
        else:
            query("""UPDATE policies SET dmvic_status='failed', dmvic_error=%s
                      WHERE policy_no=%s""",
                  (result.get('error', 'Unknown DMVIC error'), policy_no), commit=True)
            log.warning("DMVIC issuance failed for %s: %s", policy_no, result.get('error'))
# ─────────────────────────────────────────────────────────────────────────────
# MONARCH POLICY NUMBER SEQUENCES
#
# Monarch issues policy numbers as HDO|<class code>|<sequence>|<year>:
#   - Private:    HDO|0700|533144|2026, incrementing by 1 per policy
#   - Commercial: HDO|0800|012718|2026, incrementing by 1 per policy
# Confirmed with Monarch (2026-07-30). Requires
# migrations_add_monarch_sequences.sql to have been run.
# ─────────────────────────────────────────────────────────────────────────────

MONARCH_CLASS_CODES = {
    'private':    '0700',
    'commercial': '0800',
}

_monarch_seq_lock = threading.Lock()


def monarch_policy_class(product):
    """Maps a product to Monarch's private/commercial policy-number series
    using the same bucket classification DMVIC issuance uses. Returns None
    for products that aren't private or commercial (e.g. PSV, motorcycle) —
    callers must handle that case."""
    bucket = PRODUCT_TO_CERT_TYPE.get(product)
    if bucket == 'general':
        return 'private'
    if bucket == 'commercial':
        return 'commercial'
    return None


def next_monarch_policy_no(policy_class):
    """Atomically increments and returns the next Monarch policy number for
    the given class ('private' or 'commercial'), formatted as
    HDO|<class_code>|<6-digit seq>|<year>."""
    class_code = MONARCH_CLASS_CODES[policy_class]
    with _monarch_seq_lock:
        query("""UPDATE monarch_policy_sequences
                  SET last_seq = last_seq + 1
                  WHERE policy_class=%s""", (policy_class,), commit=True)
        row = query("""SELECT last_seq FROM monarch_policy_sequences
                        WHERE policy_class=%s""", (policy_class,), fetchone=True)
    seq = row['last_seq']
    year = datetime.now().year
    return f"HDO|{class_code}|{seq:06d}|{year}"

# ─────────────────────────────────────────────────────────────────────────────
# EMAIL OTP VERIFICATION (registration email confirmation)
# ─────────────────────────────────────────────────────────────────────────────

OTP_LENGTH          = 6
OTP_EXPIRY_MINUTES   = 10
OTP_MAX_ATTEMPTS     = 5
OTP_RESEND_COOLDOWN  = 60


def generate_otp():
    return f"{random.SystemRandom().randint(0, 10**OTP_LENGTH - 1):0{OTP_LENGTH}d}"


def otp_email_content(full_name, code, purpose):
    recipient_name = full_name or "there"
    if purpose == 'password_reset':
        return (
            "Reset your password — Zee Line Risk Solutions",
            f"""
            <h2>Zee Line Risk Solutions — Reset Your Password</h2>
            <p>Hi {recipient_name},</p>
            <p>Use the code below to reset your password:</p>
            <h1 style=\"letter-spacing:6px;font-size:32px;\">{code}</h1>
            <p>This code expires in {OTP_EXPIRY_MINUTES} minutes. If you did not request a password reset, you can safely ignore this email.</p>
            """,
        )
    return (
        "Verify your email — Zee Line Risk Solutions",
        f"""
        <h2>Zee Line Risk Solutions — Verify Your Email</h2>
        <p>Hi {recipient_name},</p>
        <p>Thanks for registering as an agent. Use the code below to verify your email address:</p>
        <h1 style=\"letter-spacing:6px;font-size:32px;\">{code}</h1>
        <p>This code expires in {OTP_EXPIRY_MINUTES} minutes. If you did not request this, you can safely ignore this email.</p>
        """,
    )


def send_otp_email(email, full_name, code, purpose):
    subject, html = otp_email_content(full_name, code, purpose)
    template_id = (
        BREVO_PASSWORD_RESET_TEMPLATE_ID
        if purpose == 'password_reset'
        else BREVO_VERIFICATION_TEMPLATE_ID
    )
    if BREVO_API_KEY and template_id:
        return send_brevo_email(
            email,
            template_id=template_id,
            params={
                "NAME": full_name or "there",
                "OTP": code,
                "EXPIRES_MINUTES": OTP_EXPIRY_MINUTES,
            },
            tags=[f"otp-{purpose}"],
        )
    return send_email(email, subject, html)


def create_and_send_otp(user_id, email, full_name, purpose):
    code    = generate_otp()
    expires = datetime.now() + timedelta(minutes=OTP_EXPIRY_MINUTES)

    query("""UPDATE verification_codes
             SET used=1
             WHERE user_id=%s AND purpose=%s AND used=0""",
          (user_id, purpose), commit=True)

    query("""INSERT INTO verification_codes
             (user_id, code, purpose, expires_at, attempts)
             VALUES (%s,%s,%s,%s,0)""",
          (user_id, code, purpose, expires), commit=True)

    if not send_otp_email(email, full_name, code, purpose):
        log.error("OTP email delivery request failed (user_id=%s, purpose=%s)", user_id, purpose)
        return False
    log.info("OTP email accepted for delivery (user_id=%s, purpose=%s)", user_id, purpose)
    return True


def create_and_send_verification_otp(user_id, email, full_name):
    return create_and_send_otp(user_id, email, full_name, 'register')


def create_and_send_password_reset_otp(user_id, email, full_name):
    return create_and_send_otp(user_id, email, full_name, 'password_reset')


def verify_otp(user_id, submitted_code, purpose):
    row = query("""SELECT * FROM verification_codes
                    WHERE user_id=%s AND purpose=%s AND used=0
                    ORDER BY created_at DESC LIMIT 1""",
                (user_id, purpose), fetchone=True)
    if not row:
        return False, "No pending code. Please request a new one."

    if datetime.now() > row['expires_at']:
        return False, "That code has expired. Please request a new one."

    if row['attempts'] >= OTP_MAX_ATTEMPTS:
        query("UPDATE verification_codes SET used=1 WHERE id=%s", (row['id'],), commit=True)
        return False, "Too many incorrect attempts. Please request a new code."

    import hmac
    if not hmac.compare_digest(str(row['code']), submitted_code.strip()):
        query("UPDATE verification_codes SET attempts=attempts+1 WHERE id=%s",
              (row['id'],), commit=True)
        remaining = OTP_MAX_ATTEMPTS - (row['attempts'] + 1)
        return False, f"Incorrect code. {max(remaining,0)} attempt(s) remaining."

    query("UPDATE verification_codes SET used=1 WHERE id=%s", (row['id'],), commit=True)
    return True, "Code verified."


def verify_registration_otp(user_id, submitted_code):
    return verify_otp(user_id, submitted_code, 'register')


# ─────────────────────────────────────────────────────────────────────────────
# AUTH DECORATORS
# ─────────────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({"error": "unauthorized"}), 401
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated


def clear_session_preserving_csrf():
    """Rotate application session state without invalidating the page's CSRF form token."""
    csrf_token = session.get('csrf_token')
    session.clear()
    if csrf_token:
        session['csrf_token'] = csrf_token

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            return jsonify({"error": "forbidden"}), 403
        return f(*args, **kwargs)
    return decorated

def approved_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('status') != 'approved' and session.get('role') != 'admin':
            return jsonify({"error": "account pending approval"}), 403
        return f(*args, **kwargs)
    return decorated


# ─────────────────────────────────────────────────────────────────────────────
# INSURER PRODUCT CAPABILITY MAP
# ─────────────────────────────────────────────────────────────────────────────

INSURER_PRODUCTS = {
    'monarch': {
        'private':              {'label': 'Motor Private',                 'icon': 'fa-car',            'covers': ['comprehensive', 'third_party_only']},
        'commercial_own_goods': {'label': 'Commercial — Own Goods',        'icon': 'fa-truck',          'covers': ['comprehensive', 'third_party_only']},
        'general_cartage':      {'label': 'Commercial — General Cartage', 'icon': 'fa-truck-loading',  'covers': ['comprehensive', 'third_party_only']},
        'institutional':        {'label': 'Institutional',                 'icon': 'fa-school',         'covers': ['comprehensive', 'third_party_only']},
        'agriculture_forestry': {'label': 'Agriculture & Forestry',       'icon': 'fa-tractor',        'covers': ['comprehensive', 'third_party_only']},
        'special_vehicles':     {'label': 'Special Vehicles',              'icon': 'fa-cogs',           'covers': ['comprehensive']},
        'driving_school':       {'label': 'Driving School',                'icon': 'fa-id-card',        'covers': ['comprehensive', 'third_party_only']},
        'asset_finance':        {'label': 'Asset Finance',                 'icon': 'fa-file-invoice',   'covers': ['comprehensive']},
        'psv':                  {'label': 'PSV Chauffeur Driven',          'icon': 'fa-bus',            'covers': ['comprehensive', 'third_party_only']},
        'tour_service':         {'label': 'Tour Service Vehicles',        'icon': 'fa-route',          'covers': ['comprehensive', 'third_party_only']},
        'motorcycle':           {'label': 'Motorcycle — Private',         'icon': 'fa-motorcycle',     'covers': ['comprehensive', 'third_party_only']},
        'motorcycle_psv':       {'label': 'Motorcycle — PSV',              'icon': 'fa-motorcycle',     'covers': ['comprehensive', 'third_party_only']},
        'tuktuk_commercial':    {'label': 'TukTuk — Commercial',          'icon': 'fa-shuttle-van',    'covers': ['comprehensive', 'third_party_only']},
        'tuktuk_psv':           {'label': 'TukTuk — PSV',                  'icon': 'fa-shuttle-van',    'covers': ['comprehensive', 'third_party_only']},
    },
    'directline': {
        'private':              {'label': 'Motor Private',                 'icon': 'fa-car',            'covers': ['comprehensive', 'third_party_only']},
        'commercial_own_goods': {'label': 'Commercial — Own Goods',       'icon': 'fa-truck',          'covers': ['third_party_only']},
        'general_cartage':      {'label': 'Commercial — General Cartage','icon': 'fa-truck-loading',  'covers': ['third_party_only']},
        'institutional':        {'label': 'Institutional',                 'icon': 'fa-school',         'covers': ['third_party_only']},
        'agriculture_forestry': {'label': 'Agriculture & Forestry',       'icon': 'fa-tractor',        'covers': ['third_party_only']},
        'special_vehicles':     {'label': 'Special Vehicles',              'icon': 'fa-cogs',           'covers': ['third_party_only']},
        'motorcycle':           {'label': 'Motorcycle — Private',         'icon': 'fa-motorcycle',     'covers': ['third_party_only']},
        'motorcycle_psv':       {'label': 'Motorcycle — PSV/Boda',        'icon': 'fa-motorcycle',     'covers': ['third_party_only']},
        'psv':                  {'label': 'PSV Matatu/Bus',                'icon': 'fa-bus',            'covers': ['third_party_only']},
    },
    'definite': {
        'private':                {'label': 'Motor Private (Individual)',           'icon': 'fa-car',              'covers': ['comprehensive', 'third_party_only']},
        'private_fleet':          {'label': 'Motor Private (Fleet)',                'icon': 'fa-car-side',         'covers': ['comprehensive', 'third_party_only']},
        'commercial_hybrid':      {'label': 'Commercial (Own Goods/Cartage)',      'icon': 'fa-truck',            'covers': ['comprehensive', 'third_party_only']},
        'tanker':                 {'label': 'Tankers — Flammable Liquids',         'icon': 'fa-gas-pump',         'covers': ['comprehensive']},
        'motor_trade':            {'label': 'Motor Trade — Road Risks',            'icon': 'fa-store',            'covers': ['comprehensive', 'third_party_only']},
        'private_hire_self':      {'label': 'Private Hire — Self Drive',           'icon': 'fa-key',              'covers': ['comprehensive', 'third_party_only']},
        'private_hire_chauffeur': {'label': 'Private Hire — Chauffeur/Taxi',       'icon': 'fa-user-tie',         'covers': ['comprehensive', 'third_party_only']},
        'driving_school_car':     {'label': 'Driving School — Cars',              'icon': 'fa-id-card',          'covers': ['comprehensive', 'third_party_only']},
        'driving_school_heavy':   {'label': 'Driving School — Heavy Vehicles',    'icon': 'fa-truck-monster',    'covers': ['comprehensive', 'third_party_only']},
        'institutional':          {'label': 'Institutional (School/Hotel/Office bus)', 'icon': 'fa-school',      'covers': ['comprehensive', 'third_party_only']},
        'ambulance_fire':         {'label': 'Ambulance / Firefighter',             'icon': 'fa-truck-medical',    'covers': ['comprehensive', 'third_party_only']},
        'agriculture_forestry':   {'label': 'Agricultural & Forestry',            'icon': 'fa-tractor',          'covers': ['comprehensive', 'third_party_only']},
        'special_types':          {'label': 'Special Types (Crane/Forklift/etc.)', 'icon': 'fa-cogs',             'covers': ['comprehensive', 'third_party_only']},
        'motorcycle':             {'label': 'Motorcycle (Non-PSV)',               'icon': 'fa-motorcycle',       'covers': ['comprehensive', 'third_party_only']},
        'motorcycle_psv':         {'label': 'Motorcycle — PSV',                    'icon': 'fa-motorcycle',       'covers': ['comprehensive', 'third_party_only']},
        'electric_motorbike':     {'label': 'Electric Motorbike',                  'icon': 'fa-charging-station', 'covers': ['third_party_only']},
        'tuktuk_commercial':      {'label': 'TukTuk — Commercial',                'icon': 'fa-shuttle-van',      'covers': ['comprehensive', 'third_party_only']},
        'tuktuk_psv':             {'label': 'TukTuk — PSV',                        'icon': 'fa-shuttle-van',      'covers': ['comprehensive', 'third_party_only']},
        'psv':                    {'label': 'PSV Matatu (7–35 pax)',              'icon': 'fa-bus',              'covers': ['comprehensive', 'third_party_only']},
        'psv_bus':                {'label': 'PSV Bus (Above 35 pax)',             'icon': 'fa-bus-alt',          'covers': ['comprehensive', 'third_party_only']},
        'psv_electric_bus':       {'label': 'PSV Electric Bus',                    'icon': 'fa-bus-alt',          'covers': ['comprehensive', 'third_party_only']},
        'asset_only_matatu':      {'label': 'Asset Only — Matatu',                'icon': 'fa-shield-alt',       'covers': ['comprehensive']},
        'asset_only_bus':         {'label': 'Asset Only — Bus',                    'icon': 'fa-shield-alt',       'covers': ['comprehensive']},
        'tour_service':           {'label': 'Tour Service Vehicles',              'icon': 'fa-route',            'covers': ['comprehensive', 'third_party_only']},
    },
}


def insurer_offers(company, product, cover):
    company = (company or '').lower()
    catalog = INSURER_PRODUCTS.get(company)
    if not catalog:
        return False
    entry = catalog.get(product)
    if not entry:
        return False
    return cover in entry['covers']


@app.route('/api/insurers/products')
@login_required
def api_insurer_products():
    return jsonify({"insurers": INSURER_PRODUCTS})


# ─────────────────────────────────────────────────────────────────────────────
# PREMIUM CALCULATOR
# ─────────────────────────────────────────────────────────────────────────────

PSV_RATE_TABLE = {
    7:  (6300,3781,2145,70020,23761,12147,8281),
    8:  (6570,3958,2195,72899,24751,12687,8638),
    9:  (6840,4137,2295,75868,25737,13139,8909),
    10: (7109,4319,2395,78750,26729,13681,9269),
    11: (7215,4310,2495,79817,27101,13815,9416),
    12: (7650,4588,2545,84597,28710,14667,9991),
    13: (7917,4770,2645,87478,29698,15118,10258),
    14: (7920,4840,2745,88352,30007,15311,10382),
    15: (8368,5040,2795,93327,31678,16110,10980),
    16: (8639,5220,2895,96210,32668,16650,11339),
    17: (8910,5398,2995,99180,33660,17187,11700),
    18: (9180,5488,3095,102057,34650,17640,11968),
    19: (9450,5670,3195,104938,35640,18179,12330),
    20: (9721,5851,3245,107910,36627,18629,12690),
    21: (9897,6029,3345,110788,37620,19171,13049),
    22: (10169,6117,3395,113671,38610,19621,13321),
    23: (10439,6300,3495,116637,39600,20160,13677),
    24: (10710,6477,3595,119518,40587,20700,14040),
    25: (10980,6660,3695,122488,41577,21148,14400),
    26: (11136,6699,3895,124324,42196,21487,14526),
    27: (11389,6797,4045,127161,43178,21928,14873),
    28: (11901,7137,4255,132939,45135,22947,15553),
    29: (11970,7215,4395,133821,45345,23122,15661),
    30: (12917,7733,4595,144498,48957,24904,16914),
    31: (13430,8075,4795,150278,50912,25923,17593),
    32: (13855,8328,4945,155547,52698,26861,18189),
    33: (13940,8365,5145,156127,52891,26893,18285),
    34: (14960,9007,5345,167618,56777,28898,19635),
    35: (15468,9263,5495,173398,58733,29919,20230),
    36: (17423,10456,6195,195330,66213,33659,22779),
    37: (17705,10685,6895,199287,67545,34320,23241),
    38: (18749,11248,7595,210972,71472,36298,24597),
    39: (19109,11479,8295,214897,72798,37028,25057),
    40: (19145,11570,8995,216318,73252,37244,25217),
    41: (19235,11515,9095,216254,73214,37246,25215),
    42: (19558,11710,9245,219516,74367,37757,25597),
    43: (19812,11903,9395,222780,75452,38332,25982),
    44: (20133,12094,9545,226110,76542,38910,26366),
    45: (20393,12220,9695,229372,77693,39487,26750),
    46: (20710,12410,9795,232638,78782,40063,27133),
    47: (20968,12606,9945,235902,79870,40572,27519),
    48: (21291,12732,10095,239167,81023,41150,27903),
    49: (21544,12925,10195,242431,82109,41727,28220),
    50: (21867,13118,10345,245756,83199,42302,28606),
    51: (21995,13185,10395,247295,83710,42559,28797),
    52: (22725,14695,10495,256995,90915,50055,30035),
    53: (23455,14995,10545,266495,93995,51145,30685),
    54: (24185,15195,10595,275295,96995,52055,31235),
    55: (24915,15395,10695,284095,99995,52975,31785),
    56: (25645,15695,10745,292795,102995,54055,32435),
    57: (26375,15895,10845,301555,105515,55055,33035),
    58: (27105,16145,10895,310395,106495,55975,33585),
    59: (27835,16395,10995,319075,107395,56475,33885),
    60: (28565,16615,11045,327835,108665,56475,33885),
    61: (29295,16845,11095,336595,109795,57145,34285),
    62: (29995,16995,11145,339995,109995,57225,34335),
    63: (31595,19195,11195,344895,114955,57475,34485),
    64: (31795,19295,11245,346895,115625,57805,34685),
    65: (31925,19445,11345,349895,116625,58305,34985),
    66: (31965,19445,11345,349895,116625,58305,34985),
    67: (32445,19645,11445,353895,117955,58975,35385),
    68: (32705,19745,11545,357895,119295,59645,35785),
    69: (32835,19895,11595,358395,119455,59725,35835),
    70: (33035,19995,11645,359895,119955,59975,35985),
    71: (33235,20095,11745,363895,121295,60645,36385),
    72: (33435,20195,11795,365895,121955,60975,36585),
    73: (33625,20345,11845,367895,122625,61305,36785),
    74: (33825,20445,11895,370395,123455,61725,37035),
    75: (34025,20545,11995,372895,124295,62145,37285),
    76: (34225,20645,12045,374395,124795,62395,37435),
    77: (34425,20745,12095,376895,125625,62805,37685),
    78: (34615,20895,12195,378395,126125,63055,37835),
    79: (34815,20995,12245,379895,126625,63305,37985),
    80: (35015,21095,12295,381395,127125,63555,38135),
    81: (35215,21195,12395,385395,128455,64225,38535),
    82: (35415,21345,12445,386895,128955,64475,38685),
    83: (35605,21455,12495,388395,129455,64725,38835),
    84: (35805,21545,12595,392395,130795,65395,39235),
    85: (36005,21645,12645,393895,131295,65645,39385),
    86: (36205,21795,12695,395395,131795,65895,39535),
    87: (36405,21895,12795,396895,132295,66145,39685),
    88: (36595,21995,12845,400895,133625,66805,40085),
    89: (36795,22095,12895,402395,134125,67055,40235),
    90: (36995,22245,12945,403895,134625,67305,40385),
    91: (37195,22355,13045,405395,135125,67555,40535),
    92: (37395,22445,13095,409395,136455,68225,40935),
    93: (37585,22545,13145,410895,136955,68475,41085),
    94: (37785,22695,13245,412395,137455,68725,41235),
    95: (37985,22795,13295,413895,137955,68975,41385),
    96: (38185,22895,13345,417895,139295,69645,41785),
    97: (38385,22995,13395,419395,139795,69895,41935),
    98: (38575,23095,13495,420895,140295,70145,42085),
    99: (38775,23245,13545,424895,141625,70805,42485),
    100:(38975,23345,13595,426395,142125,71055,42635),
    101:(39175,23495,13695,427895,142625,71305,42785),
    102:(39375,23545,13745,429395,143125,71555,42935),
    103:(39565,23695,13795,433395,144455,72225,43335),
    104:(39765,23795,13895,434895,144955,72475,43485),
    105:(39965,23895,13945,436895,145625,72805,43685),
}
PSV_COL = {
    '30_days': 0, '14_days': 1, '7_days': 2,
    'annual':  3, 'inst_3':  4, 'inst_6': 5, 'inst_9': 6,'inst_2': 7,
}

def get_psv_premium(seats, certificate):
    seats = int(seats or 0)
    if seats < 7:   seats = 7
    if seats > 105: seats = 105
    if seats not in PSV_RATE_TABLE:
        for cap in sorted(PSV_RATE_TABLE.keys()):
            if cap >= seats:
                seats = cap
                break
    row = PSV_RATE_TABLE[seats]
    col = PSV_COL.get(certificate, PSV_COL['annual'])
    if certificate == 'inst_10':
        base_amt = row[PSV_COL['inst_9']] * 1.02
    elif certificate == 'inst_2':
        base_amt = _period_base(row[PSV_COL['annual']], 'inst_2')
    else:
        base_amt = row[col]

    breakdown = {}
    for cert, idx in PSV_COL.items():
        breakdown[cert] = row[idx]
    breakdown['inst_10'] = round(row[PSV_COL['inst_9']] * 1.02)
    breakdown['inst_2']  = round(_period_base(row[PSV_COL['annual']], 'inst_2'))

    return round(base_amt), breakdown


TOTAL_LEVY_RATE = 0.0045
STAMP_DUTY      = 40

PERIOD_FACTORS = {
    'annual':  1.00,
    '30_days': 0.125,
    '14_days': 0.075,
    '7_days':  0.050,
}

INSTALLMENT_COUNTS = {
    'inst_2':  2,
    'inst_3':  3,
    'inst_6':  6,
    'inst_9':  9,
    'inst_10': 10,
}

NO_SHORT_TERM = {
    ('directline', 'private',        'third_party_only'),
    ('directline', 'motorcycle',     'third_party_only'),
    ('directline', 'motorcycle_psv', 'third_party_only'),
}

def _period_base(annual_base, certificate):
    if certificate in INSTALLMENT_COUNTS:
        n = INSTALLMENT_COUNTS[certificate]
        _, annual_total = _add_levies(annual_base)
        installment_total = annual_total / n
        return (installment_total - STAMP_DUTY) / (1 + TOTAL_LEVY_RATE)
    factor = PERIOD_FACTORS.get(certificate, 1.0)
    return annual_base * factor

def available_periods(company, product, cover):
    all_periods = ['annual', '30_days', '14_days', '7_days',
                    'inst_2', 'inst_3', 'inst_6', 'inst_9', 'inst_10']
    key = ((company or '').lower(), product, cover)
    if key in NO_SHORT_TERM:
        return [p for p in all_periods if p not in ('30_days', '14_days', '7_days')]
    return all_periods

@app.route('/api/insurers/periods')
@login_required
def api_available_periods():
    company = request.args.get('company', '')
    product = request.args.get('product', '')
    cover   = request.args.get('cover', '')
    return jsonify({"periods": available_periods(company, product, cover)})

DEFINITE_COMP_RATES = {
    'private': [(500_000, 1_000_000, 0.045), (1_000_001, 2_000_000, 0.035), (2_000_001, None, 0.030)],
    'private_fleet':             0.040,
    'commercial_hybrid':         0.045,
    'tanker':                    0.080,
    'motor_trade':               0.045,
    'private_hire_self':         0.075,
    'private_hire_chauffeur':    0.055,
    'driving_school_car':        0.050,
    'driving_school_heavy':      0.055,
    'institutional':             0.035,
    'ambulance_fire':            0.040,
    'agriculture_forestry':      0.035,
    'special_types':             0.030,
    'motorcycle':                0.030,
    'motorcycle_psv':            0.040,
    'motorcycle_psv_individual': 0.050,
    'tuktuk_commercial':         0.040,
    'tuktuk_psv':                0.040,
    'psv':                       0.040,
    'psv_bus':                   0.045,
    'psv_electric_bus':          0.050,
    'asset_only_matatu':         0.040,
    'asset_only_bus':            0.045,
    'tour_service':              0.045,
}

DEFINITE_COMP_MINIMUMS = {
    'private':                   30_000,
    'private_fleet':             30_000,
    'commercial_hybrid':         35_000,
    'tanker':                   100_000,
    'motor_trade':               35_000,
    'private_hire_self':         45_000,
    'private_hire_chauffeur':    37_500,
    'driving_school_car':        40_000,
    'driving_school_heavy':      40_000,
    'institutional':             35_000,
    'ambulance_fire':            40_000,
    'agriculture_forestry':      20_000,
    'special_types':             40_000,
    'motorcycle':                 5_000,
    'motorcycle_psv':             6_500,
    'motorcycle_psv_individual':  7_500,
    'tuktuk_commercial':         15_000,
    'tuktuk_psv':                21_500,
    'psv':                       30_000,
    'psv_bus':                   30_000,
    'psv_electric_bus':          50_000,
    'asset_only_matatu':         40_000,
    'asset_only_bus':            50_000,
    'tour_service':               40_000,
}

DEFINITE_TP_FLAT = {
    'private':               4_500,
    'private_fleet':          4_500,
    'motor_trade':           12_500,
    'driving_school_car':     7_500,
    'agriculture_forestry':   3_000,
    'special_types':          5_000,
    'motorcycle':             2_000,
    'tuktuk_commercial':      4_000,
    'tuktuk_psv':            21_500,
    'ambulance_fire':         7_500,
    'electric_motorbike':     5_000,
}

DEFINITE_TP_TONNAGE = [
    (0,    3,   5_500),
    (3.1,  8,   7_500),
    (8.1, None, 9_500),
]
DEFINITE_TP_PRIME_MOVER = 15_000

DEFINITE_TP_PAX_SCALE = {
    'private_hire_self':      [(0, 9, 12_500)],
    'private_hire_chauffeur': [(0, 9, 5_500), (10, 17, 8_500), (18, 25, 12_500), (26, None, 15_500)],
    'driving_school_heavy':   [(0, 15, 15_000), (15.1, None, 20_000)],
    'institutional':          [(0, 9, 7_500), (10, 25, 15_000), (26, None, 20_000)],
    'tour_service':           [(0, 9, 7_500), (10, 25, 12_500), (26, None, 15_000)],
}

DEFINITE_MOTORCYCLE_PSV_TP = 3_500


def get_definite_pax_band(scale_key, count):
    count = int(count or 0)
    bands = DEFINITE_TP_PAX_SCALE.get(scale_key, [])
    for lo, hi, amt in bands:
        if hi is None and count >= lo:
            return amt
        if hi is not None and lo <= count <= hi:
            return amt
    return bands[-1][2] if bands else 0


def get_definite_tonnage_tp(tonnage):
    t = float(tonnage or 0)
    for lo, hi, amt in DEFINITE_TP_TONNAGE:
        if hi is None and t >= lo:
            return amt
        if hi is not None and lo <= t <= hi:
            return amt
    return DEFINITE_TP_TONNAGE[-1][2]


def get_definite_comp_rate(product, value, sub_type=None):
    if product == 'motorcycle_psv' and sub_type == 'individual':
        return DEFINITE_COMP_RATES['motorcycle_psv_individual']
    entry = DEFINITE_COMP_RATES.get(product)
    if entry is None:
        return 0.040
    if isinstance(entry, list):
        for lo, hi, rate in entry:
            if hi is None and value >= lo:
                return rate
            if hi is not None and lo <= value <= hi:
                return rate
        return entry[0][2]
    return entry


def get_definite_comp_minimum(product, sub_type=None):
    if product == 'motorcycle_psv' and sub_type == 'individual':
        return DEFINITE_COMP_MINIMUMS['motorcycle_psv_individual']
    return DEFINITE_COMP_MINIMUMS.get(product, 15_000)


def get_definite_tp_commercial(tonnage=0, prime_mover=False):
    if prime_mover:
        return DEFINITE_TP_PRIME_MOVER
    return get_definite_tonnage_tp(tonnage)


MONARCH_COMP_TIERS = {
    'private':              [(500_000,1_500_000,0.0400),(1_500_001,2_000_000,0.0375),(2_000_001,2_500_000,0.0350),(2_500_001,None,0.0300)],
    'commercial_own_goods': [(500_000,1_500_000,0.0400),(1_500_001,2_000_000,0.0400),(2_000_001,2_500_000,0.0375),(2_500_001,None,0.0400)],
    'general_cartage':      [(500_000,1_500_000,0.0400),(1_500_001,2_000_000,0.0400),(2_000_001,2_500_000,0.0400),(2_500_001,None,0.0375)],
    'institutional':        [(500_000,1_500_000,0.0400),(1_500_001,2_000_000,0.0375),(2_000_001,2_500_000,0.0350),(2_500_001,None,0.0300)],
    'agriculture_forestry': [(500_000,1_500_000,0.0350),(1_500_001,2_000_000,0.0325),(2_000_001,None,0.0300)],
    'special_vehicles':     [(500_000,1_500_000,0.0400),(1_500_001,2_000_000,0.0375),(2_000_001,2_500_000,0.0350),(2_500_001,None,0.0300)],
    'driving_school':       [(500_000,1_500_000,0.0400),(1_500_001,2_000_000,0.0375),(2_000_001,2_500_000,0.0350),(2_500_001,None,0.0300)],
    'asset_finance':        [(500_000,1_500_000,0.0400),(1_500_001,2_000_000,0.0375),(2_000_001,2_500_000,0.0350),(2_500_001,None,0.0300)],
    'psv':                  [(500_000,None,0.0550)],
    'tour_service':         [(500_000,None,0.0400)],
    'motorcycle':           [(80_000, None,0.0300)],
    'motorcycle_psv':       [(80_000, None,0.0400)],
    'tuktuk_commercial':    [(200_000,None,0.0400)],
    'tuktuk_psv':           [(200_000,None,0.0500)],
    'commercial_vehicle':   [(500_000,1_500_000,0.0400),(1_500_001,2_000_000,0.0400),(2_000_001,2_500_000,0.0375),(2_500_001,None,0.0400)],
}

MONARCH_MINIMUMS = {
    'private':              27_500,
    'commercial_own_goods': 30_000,
    'general_cartage':      30_000,
    'institutional':        30_000,
    'agriculture_forestry': 30_000,
    'special_vehicles':     30_000,
    'driving_school':       30_000,
    'asset_finance':        40_000,
    'psv':                  35_000,
    'tour_service':         35_000,
    'motorcycle':            5_000,
    'motorcycle_psv':        6_000,
    'tuktuk_commercial':    10_000,
    'tuktuk_psv':           10_000,
    'commercial_vehicle':   30_000,
}

MONARCH_TONNAGE_TP = [
    (0,   3,    4_500),
    (3.1, 8,    5_500),
    (8.1, 12,   6_500),
    (12.1,15,   7_500),
    (15.1,20,  10_000),
    (20.1,None,15_000),
]

MONARCH_TP_FLAT = {
    'psv':               5_500,
    'tour_service':      5_500,
    'motorcycle_psv':    3_000,
    'motorcycle':        2_000,
    'tuktuk_commercial': 3_000,
    'tuktuk_psv':        3_000,
    'private':           3_700,
    'institutional':     5_000,
    'driving_school':    5_000,
    'agriculture_forestry': 3_000,
    'special_vehicles':  7_500,
    'commercial_vehicle':None,
    'commercial_own_goods':None,
    'general_cartage':   None,
    'asset_finance':     None,
}

def get_monarch_comp_rate(product, value):
    tiers = MONARCH_COMP_TIERS.get(product, MONARCH_COMP_TIERS['private'])
    for lo, hi, rate in tiers:
        if hi is None and value >= lo:   return rate
        if hi is not None and lo <= value <= hi: return rate
    return tiers[0][2]

def get_monarch_tonnage_tp(tonnage):
    t = float(tonnage or 0)
    for lo, hi, amt in MONARCH_TONNAGE_TP:
        if hi is None and t >= lo:             return amt
        if hi is not None and lo <= t <= hi:   return amt
    return 15_000

def get_monarch_tp_flat(product, seats=0):
    seats = int(seats or 0)
    if product in ('institutional', 'driving_school'):
        return 5_000 if seats <= 14 else 7_500
    if product == 'agriculture_forestry':
        return 3_000
    return MONARCH_TP_FLAT.get(product)


DIRECTLINE_COMP_TIERS = {
    'private': [(0, 1_500_000, 0.0400), (1_500_001, 3_000_000, 0.0375),
                (3_000_001, 5_000_000, 0.0350), (5_000_001, None, 0.0300)],
}

DIRECTLINE_MINIMUMS = {
    'private':              35_000,
    'commercial_own_goods': 40_000,
    'general_cartage':      45_000,
    'institutional':        40_000,
    'agriculture_forestry': 40_000,
    'special_vehicles':     40_000,
    'commercial_vehicle':   40_000,
}

DIRECTLINE_TONNAGE_TP = [
    (0,    10,   5_665),
    (10.1, 15,  15_100),
    (15.1, 20,  20_100),
    (20.1, None,25_200),
]

DIRECTLINE_TP_FLAT = {
    'private':         4_580,
    'motorcycle':      3_194,
    'motorcycle_psv':  3_651,
}

def get_directline_comp_rate(product, value):
    tiers = DIRECTLINE_COMP_TIERS.get(product, DIRECTLINE_COMP_TIERS['private'])
    for lo, hi, rate in tiers:
        if hi is None and value >= lo:           return rate
        if hi is not None and lo <= value <= hi: return rate
    return tiers[0][2]

def get_directline_tonnage_tp(tonnage):
    t = float(tonnage or 0)
    for lo, hi, amt in DIRECTLINE_TONNAGE_TP:
        if hi is None and t >= lo:           return amt
        if hi is not None and lo <= t <= hi: return amt
    return 25_200


GENERIC_RATES = {
    'comprehensive': {
        'private':            0.0400,
        'commercial_vehicle': 0.0450,
        'motorcycle':         0.0400,
        'psv':                0.0495,
        'tuktuk_commercial':  0.0400,
        'tuktuk_psv':         0.0500,
    },
    'third_party_only': {
        'private':            7_500,
        'commercial_vehicle': 12_000,
        'psv':                15_000,
        'motorcycle':         1_500,
        'tuktuk_commercial':  3_000,
        'tuktuk_psv':         3_000,
    },
    'third_party_fire_theft': {
        'private':            0.0150,
        'commercial_vehicle': 0.0200,
        'motorcycle':         0.0150,
    },
}

GENERIC_MINIMUMS = {
    'private':            15_000,
    'commercial_vehicle': 20_000,
    'motorcycle':          3_000,
    'psv':                30_000,
    'tuktuk_commercial':  10_000,
    'tuktuk_psv':         10_000,
}


def _add_levies(base):
    levies = base * TOTAL_LEVY_RATE + STAMP_DUTY
    return round(levies), round(base + levies)


def _build_breakdown(rate_fn, certificate, company='', product='', cover=''):
    annual_base = rate_fn(1.0)
    allowed = set(available_periods(company, product, cover))
    breakdown = {}
    for period in ('annual', '30_days', '14_days', '7_days',
                    'inst_2', 'inst_3', 'inst_6', 'inst_9', 'inst_10'):
        if period not in allowed:
            continue
        base = _period_base(annual_base, period)
        _, total = _add_levies(base)
        breakdown[period] = total
    return breakdown


def _flat_quote(annual_amount, certificate, minimum_floor=500):
    annual_base = max(annual_amount, minimum_floor)
    base = _period_base(annual_base, certificate)
    def rate_fn(f, amt=annual_base):
        return amt
    return base, rate_fn


class UnsupportedInsurerProductError(Exception):
    def __init__(self, company, product, cover):
        self.company, self.product, self.cover = company, product, cover
        super().__init__(
            f"{company or 'this insurer'} does not offer {cover.replace('_',' ')} "
            f"cover for {product.replace('_',' ')}."
        )


def calculate_premium(cover, product, value, certificate, seats=0, company='',
                       tonnage=0, sub_type=None, pax=0, enforce_catalog=True):
    company  = (company or '').lower()
    seats    = int(seats or 0)
    value    = float(value or 0)
    tonnage  = float(tonnage or 0)
    pax      = int(pax or 0)

    if enforce_catalog and company in INSURER_PRODUCTS:
        if not insurer_offers(company, product, cover):
            raise UnsupportedInsurerProductError(company, product, cover)

    if product == 'psv' and cover == 'third_party_only' and company != 'definite':
        if seats < 7:
            seats = 7
        base_amt, raw_breakdown = get_psv_premium(seats, certificate)
        levies, total = _add_levies(base_amt)
        breakdown = {}
        for period, raw_base in raw_breakdown.items():
            _, period_total = _add_levies(raw_base)
            breakdown[period] = period_total
        return {
            'base_premium':     round(base_amt),
            'levies_and_taxes': levies,
            'total_payable':    total,
            'period_breakdown': breakdown,
            'psv_table':        True,
            'seats_used':       seats,
        }

    if company == 'definite':
        minimum = get_definite_comp_minimum(product, sub_type)

        if cover == 'comprehensive':
            rate = get_definite_comp_rate(product, value, sub_type)
            annual_base = max(value * rate, minimum)
            base = _period_base(annual_base, certificate)
            def rate_fn(f, amt=annual_base):
                return amt
            levies, total = _add_levies(base)
            breakdown = _build_breakdown(rate_fn, certificate, company, product, cover)
            return {
                'base_premium':     round(base),
                'levies_and_taxes': levies,
                'total_payable':    total,
                'period_breakdown': breakdown,
                'psv_table':        False,
                'insurer':          'Definite Assurance',
                'rate_applied':     f"{rate*100:.1f}%",
            }

        elif cover == 'third_party_only':
            if product in ('commercial_hybrid',):
                flat = get_definite_tp_commercial(tonnage, prime_mover=(sub_type == 'prime_mover'))
            elif product in DEFINITE_TP_PAX_SCALE:
                flat = get_definite_pax_band(product, pax or seats)
            elif product == 'motorcycle_psv':
                flat = DEFINITE_MOTORCYCLE_PSV_TP
            elif product == 'psv':
                if seats < 7:
                    seats = 7
                base_amt, raw_breakdown = get_psv_premium(seats, certificate)
                levies, total = _add_levies(base_amt)
                breakdown = {}
                for period, raw_base in raw_breakdown.items():
                    _, period_total = _add_levies(raw_base)
                    breakdown[period] = period_total
                return {
                    'base_premium': round(base_amt), 'levies_and_taxes': levies,
                    'total_payable': total, 'period_breakdown': breakdown,
                    'psv_table': True, 'seats_used': seats, 'insurer': 'Definite Assurance',
                }
            elif product == 'psv_bus':
                flat = 30_000
            else:
                flat = DEFINITE_TP_FLAT.get(product)
                if flat is None:
                    flat = GENERIC_RATES['third_party_only'].get(product, 7_500)
            base, rate_fn = _flat_quote(flat, certificate)
            levies, total = _add_levies(base)
            breakdown = _build_breakdown(rate_fn, certificate, company, product, cover)
            return {
                'base_premium': round(base), 'levies_and_taxes': levies,
                'total_payable': total, 'period_breakdown': breakdown,
                'psv_table': False, 'insurer': 'Definite Assurance',
            }

    if company == 'monarch':
        mprod   = product if product in MONARCH_COMP_TIERS else 'private'
        minimum = MONARCH_MINIMUMS.get(mprod, 27_500)

        if cover == 'comprehensive':
            rate = get_monarch_comp_rate(mprod, value)
            annual_base = max(value * rate, minimum)
            base = _period_base(annual_base, certificate)
            def rate_fn(f, amt=annual_base):
                return amt
            levies, total = _add_levies(base)
            breakdown = _build_breakdown(rate_fn, certificate, company, product, cover)
            return {
                'base_premium': round(base), 'levies_and_taxes': levies,
                'total_payable': total, 'period_breakdown': breakdown,
                'psv_table': False, 'insurer': 'Monarch',
                'rate_applied': f"{rate*100:.2f}%",
            }

        elif cover == 'third_party_only':
            if mprod in ('commercial_vehicle', 'commercial_own_goods', 'general_cartage'):
                flat = get_monarch_tonnage_tp(tonnage)
            else:
                flat = get_monarch_tp_flat(mprod, seats)
                if flat is None:
                    flat = GENERIC_RATES['third_party_only'].get(product, 7_500)
            base, rate_fn = _flat_quote(flat, certificate)
            levies, total = _add_levies(base)
            breakdown = _build_breakdown(rate_fn, certificate, company, product, cover)
            return {
                'base_premium': round(base), 'levies_and_taxes': levies,
                'total_payable': total, 'period_breakdown': breakdown,
                'psv_table': False, 'insurer': 'Monarch',
            }

    if company == 'directline':
        dprod   = product
        minimum = DIRECTLINE_MINIMUMS.get(dprod, 35_000)

        if cover == 'comprehensive' and dprod in DIRECTLINE_COMP_TIERS:
            rate = get_directline_comp_rate(dprod, value)
            annual_base = max(value * rate, minimum)
            base = _period_base(annual_base, certificate)
            def rate_fn(f, amt=annual_base):
                return amt
            levies, total = _add_levies(base)
            breakdown = _build_breakdown(rate_fn, certificate, company, product, cover)
            return {
                'base_premium': round(base), 'levies_and_taxes': levies,
                'total_payable': total, 'period_breakdown': breakdown,
                'psv_table': False, 'insurer': 'Directline',
                'rate_applied': f"{rate*100:.2f}%",
            }

        elif cover == 'third_party_only':
            if dprod in ('commercial_vehicle', 'commercial_own_goods', 'general_cartage',
                         'institutional', 'agriculture_forestry', 'special_vehicles'):
                flat = get_directline_tonnage_tp(tonnage)
            else:
                flat = DIRECTLINE_TP_FLAT.get(dprod)
                if flat is None:
                    flat = GENERIC_RATES['third_party_only'].get(product, 7_500)
            base, rate_fn = _flat_quote(flat, certificate)
            levies, total = _add_levies(base)
            breakdown = _build_breakdown(rate_fn, certificate, company, product, cover)
            return {
                'base_premium': round(base), 'levies_and_taxes': levies,
                'total_payable': total, 'period_breakdown': breakdown,
                'psv_table': False, 'insurer': 'Directline',
            }

    rate_table = GENERIC_RATES.get(cover, GENERIC_RATES['third_party_only'])
    rate       = rate_table.get(product, rate_table.get('private', 7500))
    minimum    = GENERIC_MINIMUMS.get(product, 5000)

    if isinstance(rate, float):
        annual_base = max(value * rate, minimum)
        base = _period_base(annual_base, certificate)
        def rate_fn(f, amt=annual_base):
            return amt
    else:
        base, rate_fn = _flat_quote(rate, certificate)

    levies, total = _add_levies(base)
    breakdown = _build_breakdown(rate_fn, certificate, company, product, cover)
    return {
        'base_premium':     round(base),
        'levies_and_taxes': levies,
        'total_payable':    total,
        'period_breakdown': breakdown,
        'psv_table':        False,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PDF QUOTATION GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def generate_quote_pdf(data, quote_id, agent):
    if not PDF_AVAILABLE:
        return None
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             rightMargin=2*cm, leftMargin=2*cm,
                             topMargin=2*cm,   bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    blue   = colors.HexColor('#1d4ed8')
    gray   = colors.HexColor('#64748b')
    light  = colors.HexColor('#eff6ff')

    title_style = ParagraphStyle('title', parent=styles['Normal'],
                                  fontSize=18, textColor=blue,
                                  spaceAfter=4, fontName='Helvetica-Bold')
    sub_style   = ParagraphStyle('sub',   parent=styles['Normal'],
                                  fontSize=10, textColor=gray)
    head_style  = ParagraphStyle('head',  parent=styles['Normal'],
                                  fontSize=11, textColor=blue,
                                  spaceBefore=12, spaceAfter=4,
                                  fontName='Helvetica-Bold')
    body_style  = ParagraphStyle('body',  parent=styles['Normal'],
                                  fontSize=10, leading=16)

    def section(title, rows):
        elems = [Paragraph(title, head_style), HRFlowable(width='100%', thickness=0.5, color=blue)]
        tdata = [[Paragraph(f"<b>{k}</b>", body_style),
                  Paragraph(str(v), body_style)] for k,v in rows]
        t = Table(tdata, colWidths=[6*cm, 10*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.white),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ('TEXTCOLOR',  (0,0), (-1,-1), colors.HexColor('#1e293b')),
            ('FONTSIZE',   (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#e2e8f0')),
        ]))
        elems.append(t)
        return elems

    story = []
    story.append(Paragraph("WESTLAKE INSURANCE", title_style))
    story.append(Paragraph("Insurance Quotation", sub_style))
    story.append(Spacer(1, 0.3*cm))

    summary = Table([[
        Paragraph(f"<b>Quotation No:</b> {quote_id}", body_style),
        Paragraph(f"<b>Date:</b> {date.today().strftime('%d %b %Y')}", body_style),
        Paragraph(f"<b>Total Payable:</b> KES {data['total_payable']:,.0f}", body_style),
    ]], colWidths=[6*cm, 5*cm, 5*cm])
    summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), light),
        ('TEXTCOLOR',  (0,0), (-1,-1), blue),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('BOX', (0,0), (-1,-1), 0.5, blue),
    ]))
    story.append(summary)
    story.append(Spacer(1, 0.4*cm))

    story += section("Insurance Details", [
        ("Insurance Company",  data.get('company','').title()),
        ("Type of Cover",      data.get('type_of_cover','').replace('_',' ').title()),
        ("Certificate Type",   data.get('type_of_certificate','').replace('_',' ').title()),
        ("Commencing Date",    data.get('commencing_date','')),
        ("Expiry Date",        data.get('expiry_date','')),
    ])

    story += section("Insured Details", [
        ("Full Name",    data.get('policy_holder_name','')),
        ("KRA PIN",      data.get('kra_pin','')),
        ("Phone",        data.get('phone','')),
        ("Email",        data.get('email','—')),
        ("ID Number",    data.get('id_number','—')),
        ("Postal Address", data.get('postal_address','—')),
    ])

    story += section("Vehicle Details", [
        ("Registration No",    data.get('vehicle_reg','')),
        ("Make / Model",       data.get('make','')),
        ("Year of Manufacture", data.get('year_of_manufacture','—')),
        ("Chassis Number",     data.get('chassis_number','')),
        ("Engine Number",      data.get('engine_number','—')),
        ("Body Type",          data.get('vehicle_body_type','')),
        ("Seats / Passengers", str(data.get('seats',''))),
        ("Vehicle Value",      f"KES {float(data.get('vehicle_value',0)):,.0f}" if data.get('vehicle_value') else '—'),
        ("Tonnage",            f"{data.get('tonnage')} T" if data.get('tonnage') else '—'),
    ])

    story += section("Premium Breakdown", [
        ("Base Premium",     f"KES {data['base_premium']:,.2f}"),
        ("Levies & Taxes",   f"KES {data['levies_and_taxes']:,.2f}"),
        ("Total Payable",    f"KES {data['total_payable']:,.2f}"),
    ])

    story += section("Prepared By", [
        ("Agent Name",  agent.get('name','')),
        ("Agent Email", agent.get('email','')),
        ("Date",        date.today().strftime('%d %b %Y')),
    ])

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        "This quotation is valid for 30 days from the date of issue. "
        "Westlake Insurance reserves the right to amend or withdraw this quotation.",
        ParagraphStyle('disc', parent=styles['Normal'], fontSize=8, textColor=gray)))

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# REPORT GENERATORS
# ─────────────────────────────────────────────────────────────────────────────

def build_report_data(period):
    today = date.today()
    if period == 'daily':
        start = today
    elif period == 'weekly':
        start = today - timedelta(days=7)
    elif period == 'monthly':
        start = today.replace(day=1)
    else:
        start = today.replace(month=1, day=1)

    quotations = query("""
        SELECT q.*, u.full_name AS agent_name, u.email AS agent_email
        FROM   quotations q
        LEFT JOIN users u ON u.id = q.agent_id
        WHERE  DATE(q.created_at) >= %s
        ORDER BY q.created_at DESC
    """, (start,))

    policies = query("""
        SELECT p.*, u.full_name AS agent_name, c.first_name, c.last_name, c.vehicle_reg
        FROM   policies p
        LEFT JOIN users u ON u.id = p.agent_id
        LEFT JOIN clients c ON c.id = p.client_id
        WHERE  DATE(p.created_at) >= %s
    """, (start,))

    agents = query("""
        SELECT u.id, u.full_name, u.email,
               COUNT(DISTINCT q.id)  AS total_quotes,
               COUNT(DISTINCT p.id)  AS total_policies,
               COALESCE(SUM(q.total_payable),0) AS total_premium
        FROM   users u
        LEFT JOIN quotations q ON q.agent_id = u.id AND DATE(q.created_at) >= %s
        LEFT JOIN policies   p ON p.agent_id = u.id AND DATE(p.created_at) >= %s
        WHERE  u.role = 'agent'
        GROUP BY u.id
    """, (start, start))

    return {
        'period':     period,
        'start_date': str(start),
        'end_date':   str(today),
        'quotations': quotations,
        'policies':   policies,
        'agents':     agents,
        'totals': {
            'quotes':    len(quotations),
            'policies':  len(policies),
            'premium':   sum(float(q.get('total_payable', 0) or 0) for q in quotations),
        }
    }


def generate_report_pdf(rdata):
    if not PDF_AVAILABLE:
        return None
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm,   bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    blue   = colors.HexColor('#1d4ed8')
    story  = []

    title_s = ParagraphStyle('t', parent=styles['Normal'],
                              fontSize=16, textColor=blue, fontName='Helvetica-Bold')
    head_s  = ParagraphStyle('h', parent=styles['Normal'],
                              fontSize=12, textColor=blue, fontName='Helvetica-Bold',
                              spaceBefore=14, spaceAfter=4)
    body_s  = ParagraphStyle('b', parent=styles['Normal'], fontSize=9, leading=14)

    story.append(Paragraph(f"Westlake Insurance — {rdata['period'].title()} Report", title_s))
    story.append(Paragraph(f"Period: {rdata['start_date']} to {rdata['end_date']}", body_s))
    story.append(Spacer(1, 0.3*cm))

    summary_data = [
        ["Quotations Generated", str(rdata['totals']['quotes'])],
        ["Policies Sold",        str(rdata['totals']['policies'])],
        ["Total Premium",        f"KES {rdata['totals']['premium']:,.0f}"],
    ]
    t = Table(summary_data, colWidths=[8*cm, 8*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#eff6ff')),
        ('TEXTCOLOR',  (0,0), (-1,-1), blue),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('FONTNAME',   (1,0), (1,-1), 'Helvetica-Bold'),
        ('GRID',       (0,0), (-1,-1), 0.25, colors.HexColor('#bfdbfe')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)

    story.append(Paragraph("Agent Performance", head_s))
    story.append(HRFlowable(width='100%', thickness=0.5, color=blue))
    agent_rows = [["Agent", "Quotes", "Policies", "Premium (KES)"]]
    for a in rdata['agents']:
        agent_rows.append([
            a['full_name'],
            str(a['total_quotes']),
            str(a['total_policies']),
            f"{float(a['total_premium']):,.0f}",
        ])
    at = Table(agent_rows, colWidths=[7*cm, 3*cm, 3*cm, 4*cm])
    at.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), blue),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(at)

    story.append(Paragraph("Quotations", head_s))
    story.append(HRFlowable(width='100%', thickness=0.5, color=blue))
    qrows = [["Quote ID", "Agent", "Client", "Vehicle", "Premium"]]
    for q in rdata['quotations'][:50]:
        qrows.append([
            q.get('id',''),
            q.get('agent_name',''),
            q.get('policy_holder_name',''),
            q.get('vehicle_reg',''),
            f"KES {float(q.get('total_payable',0) or 0):,.0f}",
        ])
    qt = Table(qrows, colWidths=[3.5*cm, 3.5*cm, 3.5*cm, 3*cm, 3.5*cm])
    qt.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), blue),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(qt)

    doc.build(story)
    return buf.getvalue()


def generate_report_excel(rdata):
    if not EXCEL_AVAILABLE:
        return None
    wb = openpyxl.Workbook()

    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='1D4ED8')
    alt_fill  = PatternFill('solid', fgColor='EFF6FF')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols+1):
            cell = ws.cell(row=row, column=col)
            cell.font   = hdr_font
            cell.fill   = hdr_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def style_row(ws, row, cols, alt=False):
        for col in range(1, cols+1):
            cell = ws.cell(row=row, column=col)
            if alt:
                cell.fill = alt_fill
            cell.border = thin_border

    ws = wb.active
    ws.title = 'Summary'
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws['A1'] = f'Westlake Insurance — {rdata["period"].title()} Report'
    ws['A1'].font = Font(bold=True, size=14, color='1D4ED8')
    ws['A2'] = f'Period: {rdata["start_date"]} to {rdata["end_date"]}'

    rows = [
        ('Metric', 'Value'),
        ('Quotations Generated', rdata['totals']['quotes']),
        ('Policies Sold',        rdata['totals']['policies']),
        ('Total Premium (KES)',  f"{rdata['totals']['premium']:,.0f}"),
    ]
    for i, row in enumerate(rows, 4):
        ws.cell(i, 1, row[0])
        ws.cell(i, 2, row[1])
        if i == 4:
            style_header(ws, i, 2)
        else:
            style_row(ws, i, 2, alt=(i % 2 == 1))

    wa = wb.create_sheet('Agent Performance')
    wa.column_dimensions['A'].width = 25
    wa.column_dimensions['B'].width = 12
    wa.column_dimensions['C'].width = 12
    wa.column_dimensions['D'].width = 20
    agent_headers = ['Agent Name', 'Quotes', 'Policies', 'Premium (KES)']
    for col, h in enumerate(agent_headers, 1):
        wa.cell(1, col, h)
    style_header(wa, 1, 4)
    for r, a in enumerate(rdata['agents'], 2):
        wa.cell(r, 1, a['full_name'])
        wa.cell(r, 2, a['total_quotes'])
        wa.cell(r, 3, a['total_policies'])
        wa.cell(r, 4, f"{float(a['total_premium']):,.0f}")
        style_row(wa, r, 4, alt=(r % 2 == 0))

    wq = wb.create_sheet('Quotations')
    q_headers = ['Quote ID', 'Agent', 'Client', 'Vehicle Reg', 'Cover', 'Premium', 'Date']
    for col, h in enumerate(q_headers, 1):
        wq.cell(1, col, h)
        wq.column_dimensions[wq.cell(1, col).column_letter].width = 18
    style_header(wq, 1, 7)
    for r, q in enumerate(rdata['quotations'], 2):
        wq.cell(r, 1, q.get('id',''))
        wq.cell(r, 2, q.get('agent_name',''))
        wq.cell(r, 3, q.get('policy_holder_name',''))
        wq.cell(r, 4, q.get('vehicle_reg',''))
        wq.cell(r, 5, q.get('type_of_cover',''))
        wq.cell(r, 6, float(q.get('total_payable',0) or 0))
        wq.cell(r, 7, str(q.get('created_at','')))
        style_row(wq, r, 7, alt=(r % 2 == 0))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# GENERIC ERROR HANDLING
# ─────────────────────────────────────────────────────────────────────────────

def safe_error_response(exc, message="Something went wrong. Please try again.", status=500):
    ref = uuid.uuid4().hex[:10]
    log.error("[err-%s] %s: %s", ref, type(exc).__name__, exc)
    return jsonify({"error": message, "reference": ref}), status


@app.errorhandler(500)
def handle_500(e):
    ref = uuid.uuid4().hex[:10]
    log.error("[err-%s] Unhandled 500: %s", ref, e)
    return jsonify({"error": "maintenance on the way be patient.", "reference": ref}), 500


# ─────────────────────────────────────────────────────────────────────────────
# CERTIFICATE DECLARATIONS
#
# Admin-facing feature: certificates that DMVIC has issued (dmvic_status=
# 'issued') sit here, grouped by insurer, until an admin declares them —
# i.e. reports them to the insurer with proof of the declaration-fee
# M-Pesa payment. Requires migrations_add_declarations.sql to have been
# run (adds declared_at/declaration_id to policies, and a new
# `declarations` table).
#
# NOTE: this section is deliberately placed here — AFTER login_required /
# admin_required (defined in AUTH DECORATORS above) and AFTER
# safe_error_response (defined just above in GENERIC ERROR HANDLING) —
# rather than up near the DMVIC config block. @app.route decorators run
# at import time, so if this section were pasted any earlier than both of
# those definitions, Flask would raise NameError on startup the moment it
# hit the first @login_required / @admin_required / call to
# safe_error_response.
# ─────────────────────────────────────────────────────────────────────────────

DECLARATIONS_FOLDER = os.path.join(os.path.dirname(__file__), 'declarations')
os.makedirs(DECLARATIONS_FOLDER, exist_ok=True)

# Fill in the real recipient inbox for each insurer once you have them.
# Falls back to COMPANY_EMAIL (with a warning) if a company has no entry yet,
# so nothing silently fails to send — it just goes to the wrong place until
# you fill this in, which the log line below will make obvious.
INSURER_DECLARATION_EMAILS = {
    'monarch':     os.environ.get('MONARCH_DECLARATION_EMAIL', 'certificate@monarchinsur'),
    'directline':  os.environ.get('DIRECTLINE_DECLARATION_EMAIL', ''),
    'definite':    os.environ.get('DEFINITE_DECLARATION_EMAIL', ''),
}

# Full legal names as they should appear in the "Insurance Company" column —
# matches the wording in the sample template exactly.
INSURER_FULL_NAMES = {
    'monarch':    'The Monarch Insurance Company Ltd.',
    'directline':  'Directline Assurance Company Ltd.',
    'definite':    'Definite Assurance Company Ltd.',
}

INTERMEDIARY_NAME = 'Westlake Insurance Agency'

COVER_TYPE_LABELS = {
    'comprehensive':          'COMP',
    'third_party_only':       'TPO',
    'third_party_fire_theft': 'TPFT',
}

DECLARATION_COLUMNS = [
    'Certificate #', 'Ref #', 'Certificate Type', 'Insurance Company',
    'Intermediary', 'Issued By', 'Insured Person', 'Issuance On', 'Policy #',
    'Vehicle #', 'Chassis #', 'Start Date', 'End Date', 'Vehicle Make',
    'Vehicle Model', 'Body Type', 'Year of Manufacture', 'Engine Number',
    'Insured PIN', 'Sum Insured', 'Email', 'Phone Number',
    'Certificate Status', 'PREMIUM', 'COVER TYPE',
]


def _fetch_pending_certificates(company=None):
    """Policies that have an issued DMVIC certificate and have not yet been
    declared to the insurer. Joined against quotations/users for everything
    the declaration sheet needs."""
    sql = """
        SELECT p.policy_no, p.vehicle_reg, p.type_of_cover,
               p.commencing_date, p.expiry_date, p.total_payable,
               p.dmvic_certificate_no, p.dmvic_transaction_no,
               p.dmvic_cert_type, p.dmvic_issued_at,
               q.company, q.policy_holder_name, q.chassis_number,
               q.make, q.vehicle_body_type, q.year_of_manufacture,
               q.engine_number, q.kra_pin, q.vehicle_value, q.email, q.phone,
               u.full_name AS agent_name
        FROM   policies p
        LEFT JOIN quotations q ON q.id = p.quote_id
        LEFT JOIN users u       ON u.id = p.agent_id
        WHERE  p.dmvic_status = 'issued' AND p.declared_at IS NULL
    """
    params = ()
    if company:
        sql += " AND q.company = %s"
        params = (company,)
    sql += " ORDER BY q.company, p.dmvic_issued_at"
    return query(sql, params)


@app.route('/api/admin/declarations/pending')
@login_required
@admin_required
def list_pending_declarations():
    rows = _fetch_pending_certificates()
    grouped = {}
    for r in rows:
        grouped.setdefault(r['company'] or 'unknown', []).append(r)
    summary = {
        company: {
            "count": len(items),
            "total_premium": sum(float(i['total_payable'] or 0) for i in items),
            "certificates": items,
        }
        for company, items in grouped.items()
    }
    return jsonify({"pending": summary})


def build_declaration_excel(company, rows, mpesa_message):
    if not EXCEL_AVAILABLE:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    for col, header in enumerate(DECLARATION_COLUMNS, 1):
        ws.cell(1, col, header).font = Font(bold=True)

    company_label = INSURER_FULL_NAMES.get(company, company.title())

    r = 2
    for row in rows:
        cert_type_label = row.get('dmvic_cert_type') or ''
        ws.cell(r, 1,  row.get('dmvic_certificate_no') or '')
        ws.cell(r, 2,  row.get('dmvic_transaction_no') or '')
        ws.cell(r, 3,  cert_type_label)
        ws.cell(r, 4,  company_label)
        ws.cell(r, 5,  INTERMEDIARY_NAME)
        ws.cell(r, 6,  row.get('agent_name') or '')
        ws.cell(r, 7,  row.get('policy_holder_name') or '')
        ws.cell(r, 8,  row['dmvic_issued_at'].strftime('%d/%m/%Y') if row.get('dmvic_issued_at') else '')
        ws.cell(r, 9,  row.get('policy_no') or '')
        ws.cell(r, 10, row.get('vehicle_reg') or '')
        ws.cell(r, 11, row.get('chassis_number') or '')
        ws.cell(r, 12, row['commencing_date'].strftime('%d/%m/%Y') if row.get('commencing_date') else '')
        ws.cell(r, 13, row['expiry_date'].strftime('%d/%m/%Y') if row.get('expiry_date') else '')
        ws.cell(r, 14, row.get('make') or '')
        ws.cell(r, 15, row.get('vehicle_model') or '')  # NOTE: not currently stored — see caveat below
        ws.cell(r, 16, row.get('vehicle_body_type') or '')
        ws.cell(r, 17, row.get('year_of_manufacture') or '')
        ws.cell(r, 18, row.get('engine_number') or '')
        ws.cell(r, 19, row.get('kra_pin') or '')
        ws.cell(r, 20, float(row.get('vehicle_value') or 0))
        ws.cell(r, 21, row.get('email') or '')
        ws.cell(r, 22, row.get('phone') or '')
        ws.cell(r, 23, 'Active')
        ws.cell(r, 24, float(row.get('total_payable') or 0))
        ws.cell(r, 25, COVER_TYPE_LABELS.get(row.get('type_of_cover'), row.get('type_of_cover') or ''))
        r += 1

    total_row = r + 1
    ws.cell(total_row, 24, f"=SUM(X2:X{r-1})")

    msg_row = total_row + 2
    ws.cell(msg_row, 15, mpesa_message)

    widths = [16,14,30,33,26,14,23,18,23,16,21,14,12,14,15,14,11,15,17,13,33,18,10,10,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.route('/api/admin/declarations/send', methods=['POST'])
@login_required
@admin_required
@limiter.limit("10 per minute")
def send_declaration():
    d             = request.get_json() or {}
    company       = (d.get('company') or '').strip().lower()
    policy_nos    = d.get('policy_nos') or []
    mpesa_message = (d.get('mpesa_message') or '').strip()

    if not company or not policy_nos:
        return jsonify({"error": "company and policy_nos are required"}), 400
    if not mpesa_message:
        return jsonify({"error": "Please paste the M-Pesa confirmation message before sending."}), 400

    all_pending = _fetch_pending_certificates(company)
    rows = [r for r in all_pending if r['policy_no'] in policy_nos]
    if not rows:
        return jsonify({"error": "None of the selected certificates are pending declaration for this insurer."}), 400

    excel_bytes = build_declaration_excel(company, rows, mpesa_message)
    if excel_bytes is None:
        return jsonify({"error": "Excel generation unavailable — install openpyxl"}), 503

    decl_id   = f"DECL-{company.upper()}-{date.today().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"
    file_name = f"{decl_id}.xlsx"
    file_path = os.path.join(DECLARATIONS_FOLDER, file_name)
    with open(file_path, 'wb') as f:
        f.write(excel_bytes)

    total_premium = sum(float(r.get('total_payable') or 0) for r in rows)
    recipient = INSURER_DECLARATION_EMAILS.get(company) or COMPANY_EMAIL
    if not INSURER_DECLARATION_EMAILS.get(company):
        log.warning("No declaration email configured for '%s' — falling back to COMPANY_EMAIL. "
                    "Set %s_DECLARATION_EMAIL in the environment.", company, company.upper())

    try:
        query("""INSERT INTO declarations
                    (id, company, created_by, mpesa_message, certificate_count,
                     total_premium, file_path, email_sent_to)
                 VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
              (decl_id, company, session['user_id'], mpesa_message, len(rows),
               total_premium, file_path, recipient), commit=True)
    except Exception as e:
        return safe_error_response(e, "Could not record this declaration.")

    sent = send_email(
        recipient,
        f"Certificate Declaration — {INSURER_FULL_NAMES.get(company, company.title())} — {date.today()}",
        f"""<h2>Certificate Declaration</h2>
            <p>Please find attached {len(rows)} certificate(s) declared by Westlake
               Insurance Agency, totalling KES {total_premium:,.0f}.</p>
            <h3>M-Pesa Payment Confirmation</h3>
            <p style="font-family:monospace;">{mpesa_message}</p>""",
        attachments=[(file_name, excel_bytes)],
    )

    query("""UPDATE policies SET declared_at=NOW(), declaration_id=%s
             WHERE policy_no IN (%s)""" % ('%s', ','.join(['%s'] * len(policy_nos))),
          (decl_id, *policy_nos), commit=True)

    if sent:
        query("UPDATE declarations SET email_sent=1 WHERE id=%s", (decl_id,), commit=True)
    else:
        log.warning("Declaration %s saved but email to %s failed — file is still on disk for manual sending.",
                    decl_id, recipient)

    cache_delete_prefix("cache:dashboard")

    return jsonify({
        "success":        True,
        "declaration_id": decl_id,
        "email_sent":     sent,
        "sent_to":        recipient,
        "certificate_count": len(rows),
        "total_premium":  total_premium,
        "download_url":   f"/api/admin/declarations/{decl_id}/download",
    })


@app.route('/api/admin/declarations/<declaration_id>/download')
@login_required
@admin_required
def download_declaration(declaration_id):
    row = query("SELECT * FROM declarations WHERE id=%s", (declaration_id,), fetchone=True)
    if not row or not os.path.exists(row['file_path']):
        abort(404)
    return send_file(row['file_path'], as_attachment=True,
                      download_name=os.path.basename(row['file_path']))


@app.route('/api/admin/declarations/history')
@login_required
@admin_required
def declarations_history():
    rows = query("""
        SELECT d.*, u.full_name AS sent_by
        FROM   declarations d
        LEFT JOIN users u ON u.id = d.created_by
        ORDER BY d.created_at DESC
        LIMIT 200
    """)
    return jsonify({"declarations": rows})


# ─────────────────────────────────────────────────────────────────────────────
# PAGE ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect('/dashboard')
    return render_template('home.html', site_url=PUBLIC_SITE_URL)


@app.route('/robots.txt')
def robots_txt():
    body = "\n".join([
        "User-agent: *",
        "Allow: /",
        "Disallow: /api/",
        "Disallow: /dashboard",
        "Disallow: /quotation",
        "Disallow: /clients",
        "Disallow: /claims",
        "Disallow: /renewals",
        "Disallow: /pending",
        "Disallow: /login",
        f"Sitemap: {PUBLIC_SITE_URL}/sitemap.xml",
        "",
    ])
    return Response(body, content_type='text/plain; charset=utf-8')


@app.route('/sitemap.xml')
def sitemap_xml():
    pages = ('/', '/privacy', '/terms')
    entries = ''.join(
        f"<url><loc>{PUBLIC_SITE_URL}{path}</loc></url>"
        for path in pages
    )
    body = ('<?xml version="1.0" encoding="UTF-8"?>'
            '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">'
            f'{entries}</urlset>')
    return Response(body, content_type='application/xml; charset=utf-8')


@app.route('/favicon.ico')
def favicon():
    return send_file(
        os.path.join(BASE_DIR, '..', 'css', 'images', 'favicon.ico'),
        mimetype='image/vnd.microsoft.icon',
        max_age=60 * 60 * 24 * 30,
    )

@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect('/dashboard')
    return render_template('login.html')

@app.route('/pending')
@login_required
def pending():
    return render_template('pending.html')

@app.route('/dashboard')
@login_required
def dashboard():
    user = query("SELECT status FROM users WHERE id=%s", (session['user_id'],), fetchone=True)
    if user and user['status'] != 'approved' and session.get('role') != 'admin':
        return redirect('/pending')
    return render_template('dashboard.html')

@app.route('/quotation')
@login_required
def quotation_page():
    return render_template('quotation.html')

@app.route('/clients')
@login_required
def clients_page():
    return render_template('clients.html')

@app.route('/claims')
@login_required
def claims_page():
    return render_template('claims.html')

@app.route('/renewals')
@login_required
def renewals_page():
    return render_template('renewals.html')

@app.route('/terms')
def terms_page():
    return render_template('terms.html')

@app.route('/privacy')
def privacy_page():
    return render_template('privacy.html')


# ─────────────────────────────────────────────────────────────────────────────
# AUTH API
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/login', methods=['POST'])
@limiter.limit("10 per minute")
def api_login():
    d        = request.get_json() or {}
    username = d.get('username', '').strip()
    password = d.get('password', '')

    if not username or not password:
        return jsonify({"error": "Please fill all fields"}), 400

    try:
        user = query("SELECT * FROM users WHERE username=%s", (username,), fetchone=True)
    except Exception as e:
        return safe_error_response(e, "Login is temporarily unavailable.")

    if not user or not check_password_hash(user['password_hash'], password):
        return jsonify({"error": "Invalid username or password"}), 401

    if user['status'] == 'unverified':
        clear_session_preserving_csrf()
        session['unverified_user_id'] = user['id']
        return jsonify({
            "error": "Please verify your email before logging in.",
            "require_verification": True,
            "user_id": user['id'],
            "email": user.get('email', ''),
        }), 403

    session.clear()
    session.regenerate = True
    session['user_id']  = user['id']
    session['username'] = user['username']
    session['role']     = user['role']
    session['status']   = user['status']
    session['email']    = user.get('email', '')
    session['name']     = user.get('full_name', user['username'])

    try:
        query("INSERT INTO audit_log (user_id, action) VALUES (%s, %s)",
              (user['id'], 'login'), commit=True)
    except Exception as e:
        log.warning("Audit log write failed: %s", type(e).__name__)

    if user['role'] == 'agent' and user['status'] != 'approved':
        return jsonify({"redirect": "/pending"})

    return jsonify({"redirect": "/dashboard"})


@app.route('/api/logout')
def logout():
    if 'user_id' in session:
        try:
            query("INSERT INTO audit_log (user_id, action) VALUES (%s, %s)",
                  (session['user_id'], 'logout'), commit=True)
        except Exception as e:
            log.warning("Audit log write failed: %s", type(e).__name__)
    session.clear()
    return jsonify({"success": True})


@app.route('/api/register', methods=['POST'])
@limiter.limit("5 per minute")
def register():
    d                 = request.get_json() or {}
    full_name         = d.get('full_name', '').strip()
    username          = d.get('username', '').strip()
    email             = d.get('email', '').strip().lower()
    phone             = d.get('phone', '').strip()
    id_number         = d.get('id_number', '').strip()
    kra_pin           = d.get('kra_pin', '').strip().upper()
    gender            = d.get('gender', '').strip()
    ira_number        = d.get('ira_number', '').strip() or None
    city              = d.get('city', '').strip()
    county            = d.get('county', '').strip()
    commission_payout = d.get('commission_payout', '').strip().lower()
    password          = d.get('password', '')

    if commission_payout not in ('daily', 'weekly', 'monthly'):
        return jsonify({"error": "Please select a valid commission payout period"}), 400

    if not all([full_name, username, email, phone, id_number, kra_pin,
                gender, city, county, password]):
        return jsonify({"error": "Please fill all required fields"}), 400

    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters long"}), 400

    try:
        existing = query("SELECT id FROM users WHERE username=%s OR email=%s",
                         (username, email), fetchone=True)
    except Exception as e:
        return safe_error_response(e, "Registration is temporarily unavailable.")

    if existing:
        return jsonify({"error": "Username or email already exists"}), 409

    pw_hash = generate_password_hash(password)

    try:
        user_id = query("""
            INSERT INTO users (full_name, username, email, password_hash, role, status,
                               phone, id_number, kra_pin, gender, ira_number,
                               city, county, commission_payout)
            VALUES (%s,%s,%s,%s,'agent','unverified',%s,%s,%s,%s,%s,%s,%s,%s)
        """, (full_name, username, email, pw_hash, phone, id_number, kra_pin,
              gender, ira_number, city, county, commission_payout), commit=True)
    except Exception as e:
        return safe_error_response(e, "Could not create account. Please try again.")

    enqueue("send_verification_otp", create_and_send_verification_otp,
            user_id, email, full_name)

    clear_session_preserving_csrf()
    session['unverified_user_id'] = user_id

    return jsonify({
        "success":              True,
        "require_verification": True,
        "user_id":               user_id,
        "message":              "Account created. Enter the verification code sent to your email."
    })


@app.route('/api/verify-email', methods=['POST'])
@limiter.limit("10 per minute")
def api_verify_email():
    d       = request.get_json() or {}
    code    = d.get('code', '').strip()
    user_id = session.get('unverified_user_id') or d.get('user_id')
    email   = d.get('email', '').strip().lower()

    # The agent may open the code on another device/browser, where the
    # pending-registration session is unavailable. The OTP still protects the
    # action, so resolve a matching unverified account by its registered email.
    if not user_id and email:
        pending_user = query("SELECT id FROM users WHERE email=%s AND status='unverified'",
                             (email,), fetchone=True)
        user_id = pending_user.get('id') if pending_user else None

    if not user_id:
        return jsonify({"error": "Enter the email used for registration, then try the verification code again."}), 400
    if not code:
        return jsonify({"error": "Please enter the verification code."}), 400

    ok, msg = verify_registration_otp(user_id, code)
    if not ok:
        return jsonify({"error": msg}), 400

    user = query("SELECT * FROM users WHERE id=%s", (user_id,), fetchone=True)
    if not user:
        return jsonify({"error": "Account not found."}), 404

    query("UPDATE users SET status='pending' WHERE id=%s", (user_id,), commit=True)

    query("INSERT INTO audit_log (user_id, action) VALUES (%s, %s)",
          (user_id, 'email_verified'), commit=True)

    enqueue("notify_new_agent", notify_new_agent,
            user['full_name'], user['username'], user['email'])

    session.pop('unverified_user_id', None)

    return jsonify({
        "success": True,
        "message": "Email verified! Your account is now awaiting admin approval.",
    })


@app.route('/api/verify-email/resend', methods=['POST'])
@limiter.limit("3 per minute")
def resend_verification_otp():
    d       = request.get_json() or {}
    user_id = session.get('unverified_user_id') or d.get('user_id')
    email   = d.get('email', '').strip().lower()
    if not user_id and email:
        pending_user = query("SELECT id FROM users WHERE email=%s AND status='unverified'",
                             (email,), fetchone=True)
        user_id = pending_user.get('id') if pending_user else None
    if not user_id:
        return jsonify({"error": "Enter the email used for registration to request a new code."}), 400

    user = query("SELECT * FROM users WHERE id=%s AND status='unverified'",
                 (user_id,), fetchone=True)
    if not user:
        return jsonify({"error": "No pending verification found for this account."}), 404

    last = query("""SELECT created_at FROM verification_codes
                     WHERE user_id=%s AND purpose='register'
                     ORDER BY created_at DESC LIMIT 1""",
                (user_id,), fetchone=True)
    if last:
        elapsed = (datetime.now() - last['created_at']).total_seconds()
        if elapsed < OTP_RESEND_COOLDOWN:
            wait = int(OTP_RESEND_COOLDOWN - elapsed)
            return jsonify({"error": f"Please wait {wait}s before requesting another code."}), 429

    enqueue("resend_verification_otp", create_and_send_verification_otp,
            user_id, user['email'], user['full_name'])

    return jsonify({"success": True, "message": "A new verification code has been sent."})


def valid_email_address(value):
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value or ""))


@app.route('/api/password-reset/request', methods=['POST'])
@limiter.limit("3 per minute")
def request_password_reset():
    """Start a password reset without revealing whether an email has an account."""
    d = request.get_json() or {}
    email = d.get('email', '').strip().lower()
    message = "If an account matches that email, a password-reset code has been sent."
    if not valid_email_address(email):
        return jsonify({"success": True, "message": message})

    user = query("SELECT * FROM users WHERE email=%s", (email,), fetchone=True)
    if not user:
        return jsonify({"success": True, "message": message})

    last = query("""SELECT created_at FROM verification_codes
                     WHERE user_id=%s AND purpose='password_reset'
                     ORDER BY created_at DESC LIMIT 1""",
                 (user['id'],), fetchone=True)
    if last:
        elapsed = (datetime.now() - last['created_at']).total_seconds()
        if elapsed < OTP_RESEND_COOLDOWN:
            return jsonify({"success": True, "message": message})

    enqueue("send_password_reset_otp", create_and_send_password_reset_otp,
            user['id'], user['email'], user.get('full_name', ''))
    return jsonify({"success": True, "message": message})


@app.route('/api/password-reset/confirm', methods=['POST'])
@limiter.limit("10 per minute")
def confirm_password_reset():
    d = request.get_json() or {}
    email = d.get('email', '').strip().lower()
    code = d.get('code', '').strip()
    password = d.get('password', '')
    confirm_password = d.get('confirm_password', '')

    if not valid_email_address(email) or not code:
        return jsonify({"error": "Enter your email address and verification code."}), 400
    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters long."}), 400
    if password != confirm_password:
        return jsonify({"error": "Passwords do not match."}), 400

    user = query("SELECT * FROM users WHERE email=%s", (email,), fetchone=True)
    if not user:
        return jsonify({"error": "The reset code is invalid or has expired."}), 400

    ok, message = verify_otp(user['id'], code, 'password_reset')
    if not ok:
        return jsonify({"error": message}), 400

    query("UPDATE users SET password_hash=%s WHERE id=%s",
          (generate_password_hash(password), user['id']), commit=True)
    query("INSERT INTO audit_log (user_id, action) VALUES (%s, %s)",
          (user['id'], 'password_reset'), commit=True)
    session.clear()
    return jsonify({"success": True, "message": "Password reset. You can now sign in."})


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD STATS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/dashboard/stats')
@login_required
@cached_response("cache:dashboard", ttl=30)
def dashboard_stats():
    role    = session.get('role')
    user_id = session.get('user_id')

    try:
        if role == 'admin':
            total_agents = query("SELECT COUNT(*) AS total FROM users WHERE role='agent'", fetchone=True)['total']

            total_clients = query("SELECT COUNT(*) AS total FROM clients", fetchone=True)['total']

            total_quotes = query("SELECT COUNT(*) AS total FROM quotations", fetchone=True)['total']

            active_policies = query("SELECT COUNT(*) AS total FROM policies WHERE status='active'", fetchone=True)['total']

            total_premium = float(query("SELECT COALESCE(SUM(total_payable),0) AS total FROM quotations", fetchone=True)['total'])

            return jsonify({
                "name":            session.get('name'),
                "role":            role,
                "total_clients":   total_clients,
                "total_agents":    total_agents,
                "total_quotes":    total_quotes,
                "active_policies": active_policies,
                "total_premium":   total_premium,
                "total_activity":  total_quotes,
            })

        my_clients = query("SELECT COUNT(*) AS total FROM clients WHERE agent_id=%s", (user_id,), fetchone=True)['total']

        my_quotes = query("SELECT COUNT(*) AS total FROM quotations WHERE agent_id=%s", (user_id,), fetchone=True)['total']

        my_active = query("SELECT COUNT(*) AS total FROM policies WHERE agent_id=%s AND status='active'", (user_id,), fetchone=True)['total']

        my_premium = float(query("SELECT COALESCE(SUM(total_payable),0) AS total FROM quotations WHERE agent_id=%s", (user_id,), fetchone=True)['total'])

        return jsonify({
            "name":        session.get('name'),
            "role":        role,
            "my_clients":  my_clients,
            "my_quotes":   my_quotes,
            "my_active":   my_active,
            "my_premium":  my_premium,
            "my_activity": my_quotes,
        })
    except Exception as e:
        return safe_error_response(e, "Could not load dashboard stats.")


# ─────────────────────────────────────────────────────────────────────────────
# ADMIN — AGENTS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/admin/agents')
@login_required
@admin_required
def list_agents():
    agents = query("""
        SELECT u.id, u.full_name, u.username, u.email, u.status, u.created_at,
               u.flagged, u.flagged_reason, u.flagged_at, u.underpayment_attempts,
               COUNT(DISTINCT q.id)                       AS total_quotes,
               COUNT(DISTINCT p.id)                       AS total_policies,
               COALESCE(SUM(q.total_payable), 0)          AS total_premium,
               COUNT(DISTINCT c.id)                       AS total_clients
        FROM   users u
        LEFT JOIN quotations q ON q.agent_id = u.id
        LEFT JOIN policies   p ON p.agent_id = u.id
        LEFT JOIN clients    c ON c.agent_id = u.id
        WHERE  u.role = 'agent'
        GROUP BY u.id
        ORDER BY u.created_at DESC
    """)
    return jsonify({"agents": agents})


@app.route('/api/admin/agents/<int:agent_id>/status', methods=['POST'])
@login_required
@admin_required
def update_agent_status(agent_id):
    d      = request.get_json() or {}
    status = d.get('status')
    if status not in ('approved', 'rejected', 'suspended'):
        return jsonify({"error": "Invalid status"}), 400
    query("UPDATE users SET status=%s WHERE id=%s", (status, agent_id), commit=True)
    query("INSERT INTO audit_log (user_id, action, detail) VALUES (%s,%s,%s)",
          (session['user_id'], 'agent_status_change',
           f"agent_id={agent_id} status={status}"), commit=True)
    return jsonify({"success": True})


@app.route('/api/admin/agents/<int:agent_id>/unflag', methods=['POST'])
@login_required
@admin_required
def unflag_agent(agent_id):
    """Clears an underpayment-attempt flag once an admin has reviewed it.
    Does NOT reset underpayment_attempts (the running count is kept as a
    permanent history) — only clears the flagged/flagged_reason/flagged_at
    fields so the account stops showing as actively flagged."""
    query("""UPDATE users SET flagged=0, flagged_reason=NULL, flagged_at=NULL
              WHERE id=%s""", (agent_id,), commit=True)
    query("INSERT INTO audit_log (user_id, action, detail) VALUES (%s,%s,%s)",
          (session['user_id'], 'agent_unflagged', f"agent_id={agent_id}"), commit=True)
    return jsonify({"success": True})


@app.route('/api/admin/users/<int:target_user_id>', methods=['DELETE'])
@login_required
@admin_required
@limiter.limit("10 per minute")
def delete_user(target_user_id):
    return jsonify({
        "error": "Permanent user deletion is disabled. Suspend the account instead to preserve its records."
    }), 405


# ─────────────────────────────────────────────────────────────────────────────
# CLIENTS
#
# FIXED in v2.5: add_client() previously had a stray block appended after
# its correct body — leftover code that referenced policy_no/q (which only
# exist in buy_cover()) and an undefined _issue_dmvic_certificate_for_policy
# function. Every call to POST /api/clients/add raised NameError. That
# block has been removed; this function now only creates a client record,
# which is all it was ever meant to do.
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/clients/list')
@login_required
def list_clients():
    if session['role'] == 'admin':
        clients = query("""
            SELECT c.*, u.full_name AS agent_name
            FROM   clients c
            LEFT JOIN users u ON u.id = c.agent_id
            ORDER BY c.created_at DESC
        """)
    else:
        clients = query("""
            SELECT c.*, u.full_name AS agent_name
            FROM   clients c
            LEFT JOIN users u ON u.id = c.agent_id
            WHERE  c.agent_id = %s
            ORDER BY c.created_at DESC
        """, (session['user_id'],))
    return jsonify({"clients": clients})


@app.route('/api/clients/add', methods=['POST'])
@login_required
@approved_required
def add_client():
    d = request.get_json() or {}
    required = ('first_name', 'last_name', 'phone')
    if not all(d.get(k, '').strip() for k in required):
        return jsonify({"error": "First name, last name and phone are required"}), 400

    existing = query(
        "SELECT id FROM clients WHERE phone=%s AND agent_id=%s",
        (d['phone'], session['user_id']), fetchone=True
    )
    if existing:
        return jsonify({"error": "A client with this phone number already exists"}), 409

    try:
        cid = query("""
            INSERT INTO clients
                (agent_id, first_name, last_name, phone,
                 id_number, kra_pin, vehicle_reg, email)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            session['user_id'],
            d['first_name'].strip(), d['last_name'].strip(),
            d['phone'].strip(),
            d.get('id_number','').strip(),
            d.get('kra_pin','').strip().upper(),
            d.get('vehicle_reg','').strip().upper(),
            d.get('email','').strip(),
        ), commit=True)
    except Exception as e:
        return safe_error_response(e, "Could not save this client.")

    cache_delete_prefix("cache:dashboard")
    return jsonify({"success": True, "client_id": cid})


# ─────────────────────────────────────────────────────────────────────────────
# FILE UPLOAD
# ─────────────────────────────────────────────────────────────────────────────

def allowed_file(filename):
    return ('.' in filename and
            filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS)


def content_matches_extension(file_bytes, ext):
    ext = ext.lower()
    allowed_mimes = ALLOWED_MIME_BY_EXT.get(ext, set())
    if not allowed_mimes:
        return False

    if FILETYPE_AVAILABLE:
        kind = filetype.guess(file_bytes)
        if kind is None:
            return False
        return kind.mime in allowed_mimes

    if file_bytes.startswith(b'%PDF-'):
        return 'application/pdf' in allowed_mimes
    if file_bytes.startswith(b'\x89PNG\r\n\x1a\n'):
        return 'image/png' in allowed_mimes
    if file_bytes.startswith(b'\xff\xd8\xff'):
        return 'image/jpeg' in allowed_mimes
    return False


@app.route('/api/upload', methods=['POST'])
@login_required
@approved_required
@limiter.limit("20 per minute")
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400
    uploaded = []
    rejected = []
    for f in request.files.getlist('file'):
        if not f or not f.filename:
            continue
        if not allowed_file(f.filename):
            rejected.append(f.filename)
            continue

        ext = f.filename.rsplit('.', 1)[1].lower()
        head = f.stream.read(2048)
        f.stream.seek(0)

        if not content_matches_extension(head, ext):
            rejected.append(f.filename)
            log.warning("Rejected upload with mismatched content/extension: %s", f.filename)
            continue

        fname = secure_filename(f.filename)
        uid   = str(uuid.uuid4())[:8]
        fname = f"{uid}_{fname}"
        path  = os.path.join(UPLOAD_FOLDER, fname)
        f.save(path)

        try:
            query("""
                INSERT INTO documents (user_id, filename, filepath, doc_type)
                VALUES (%s, %s, %s, %s)
            """, (session['user_id'], f.filename, path, 'claim_evidence'), commit=True)
        except Exception as e:
            log.error("Document DB insert failed: %s", type(e).__name__)
            try:
                os.remove(path)
            except OSError:
                pass
            rejected.append(f.filename)
            continue

        uploaded.append(fname)

    resp = {"success": True, "files": uploaded}
    if rejected:
        resp["rejected"] = rejected
        resp["warning"] = "Some files were rejected (unsupported type or content did not match extension)."
    return jsonify(resp)


# ─────────────────────────────────────────────────────────────────────────────
# QUOTATIONS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/quotations/generate', methods=['POST'])
@login_required
@approved_required
@limiter.limit("30 per minute")
def generate_quotation():
    d = request.get_json() or {}

    # DMVIC stores make and model independently.  Keep the legacy combined
    # ``make`` display field for existing templates/PDFs, but never infer one
    # DMVIC field from the other.
    d['vehicle_make'] = str(d.get('vehicle_make') or '').strip()
    d['vehicle_model'] = str(d.get('vehicle_model') or '').strip()
    d['make'] = str(d.get('make') or f"{d['vehicle_make']} {d['vehicle_model']}").strip()

    required = (
        'company', 'type_of_cover', 'type_of_certificate', 'product',
        'commencing_date', 'expiry_date',
        'policy_holder_name', 'phone', 'kra_pin',
        'vehicle_reg', 'chassis_number', 'vehicle_body_type', 'seats'
    )

    missing = [k for k in required if not d.get(k)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    try:
        value = float(d.get('vehicle_value', 0) or 0)
    except (ValueError, TypeError):
        value = 0

    cover = d['type_of_cover']
    product = d['product']
    cert = d['type_of_certificate']

    if cover != 'third_party_only' and value <= 0 and product != 'psv':
        return jsonify({"error": "Vehicle value is required for this cover type"}), 400

    try:
        calc = calculate_premium(
            cover,
            product,
            value,
            cert,
            seats=int(d.get('seats', 0) or 0),
            company=d.get('company', ''),
            tonnage=float(d.get('tonnage', 0) or 0),
            sub_type=d.get('sub_type'),
            pax=int(d.get('pax', 0) or 0),
        )
    except UnsupportedInsurerProductError as e:
        return jsonify({"error": str(e)}), 400

    quote_id = f"WL-{date.today().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"

    agent = {
        "name": session.get("name", session.get("username")),
        "email": session.get("email", ""),
        "id": session["user_id"],
    }

    try:
        query("""
            INSERT INTO quotations (
                id,
                agent_id,
                company,
                type_of_cover,
                type_of_certificate,
                product,
                sub_type,
                commencing_date,
                expiry_date,
                policy_holder_name,
                kra_pin,
                phone,
                email,
                id_number,
                postal_address,
                vehicle_reg,
                chassis_number,
                engine_number,
                vehicle_body_type,
                make,
                vehicle_make,
                vehicle_model,
                year_of_manufacture,
                year_of_registration,
                seats,
                vehicle_value,
                tonnage,
                base_premium,
                levies_and_taxes,
                total_payable,
                status
            )
            VALUES (
                %s,%s,%s,%s,%s,
                %s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s
            )
        """, (
            quote_id,
            session["user_id"],
            d["company"],
            cover,
            cert,
            product,
            d.get("sub_type") or None,
            d["commencing_date"],
            d["expiry_date"],
            d["policy_holder_name"],
            d["kra_pin"],
            d["phone"],
            d.get("email", ""),
            d.get("id_number", "").strip(),
            d.get("postal_address", "").strip(),
            d["vehicle_reg"].upper(),
            d["chassis_number"].upper(),
            d.get("engine_number", "").upper(),
            d["vehicle_body_type"],
            d["make"],
            d["vehicle_make"],
            d["vehicle_model"],
            d.get("year_of_manufacture") or None,
            d.get("year_of_registration") or None,  # <-- Added missing parameter here
            int(d["seats"]),
            value,
            float(d.get("tonnage", 0) or 0),
            calc["base_premium"],
            calc["levies_and_taxes"],
            calc["total_payable"],
            "pending"
        ), commit=True)
    except Exception as e:
        return safe_error_response(e, "Could not save this quotation.")

    d.update(calc)

    def _generate_and_notify():
        pdf_bytes = None

        if PDF_AVAILABLE:
            try:
                pdf_bytes = generate_quote_pdf(d, quote_id, agent)
            except Exception as e:
                log.error("PDF error: %s", type(e).__name__)

        try:
            notify_quotation(
                agent,
                {
                    k: d.get(k, "")
                    for k in (
                        "policy_holder_name",
                        "email",
                        "phone",
                        "kra_pin",
                    )
                },
                {
                    k: d.get(k, "")
                    for k in (
                        "vehicle_reg",
                        "chassis_number",
                        "make",
                        "vehicle_body_type",
                        "seats",
                    )
                },
                {
                    k: d.get(k, "")
                    for k in (
                        "company",
                        "type_of_cover",
                        "type_of_certificate",
                        "commencing_date",
                        "expiry_date",
                    )
                },
                {
                    "id": quote_id,
                    "generated_at": str(date.today()),
                    "base_premium": calc["base_premium"],
                    "levies_and_taxes": calc["levies_and_taxes"],
                    "total_payable": calc["total_payable"],
                },
                pdf_bytes,
            )
        except Exception as e:
            log.error("Email notification error: %s", type(e).__name__)

    enqueue("quotation_certificate_and_email", _generate_and_notify)
    cache_delete_prefix("cache:dashboard")

    return jsonify({
        "quote": {
            "id": quote_id,
            "base_premium": calc["base_premium"],
            "levies_and_taxes": calc["levies_and_taxes"],
            "total_payable": calc["total_payable"],
        },
        "period_breakdown": calc["period_breakdown"],
        "pdf_available": PDF_AVAILABLE,
    })


@app.route('/api/quotations/pdf/<quote_id>')
@login_required
def download_quotation_pdf(quote_id):
    if not PDF_AVAILABLE:
        return jsonify({"error": "Quotation PDF generation is unavailable."}), 503

    quote = query("SELECT * FROM quotations WHERE id=%s", (quote_id,), fetchone=True)
    if not quote:
        return jsonify({"error": "Quotation not found."}), 404
    if session['role'] != 'admin' and quote.get('agent_id') != session['user_id']:
        return jsonify({"error": "You do not have access to this quotation."}), 403

    agent = query("SELECT full_name, email FROM users WHERE id=%s",
                  (quote.get('agent_id'),), fetchone=True) or {}
    try:
        pdf_bytes = generate_quote_pdf(
            quote,
            quote_id,
            {"name": agent.get('full_name', ''), "email": agent.get('email', '')},
        )
    except Exception:
        log.exception("Quotation PDF generation failed for %s", quote_id)
        return jsonify({"error": "Could not generate this quotation PDF."}), 500

    if not pdf_bytes:
        return jsonify({"error": "Quotation PDF generation is unavailable."}), 503

    filename = secure_filename(f"quotation-{quote_id}.pdf") or "quotation.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename,
        max_age=0,
    )


@app.route('/api/quotations/list')
@login_required
def list_quotations():
    if session['role'] == 'admin':
        rows = query("""
            SELECT q.*, u.full_name AS agent_name
            FROM   quotations q
            LEFT JOIN users u ON u.id = q.agent_id
            ORDER BY q.created_at DESC LIMIT 200
        """)
    else:
        rows = query("""
            SELECT * FROM quotations
            WHERE  agent_id = %s
            ORDER BY created_at DESC LIMIT 200
        """, (session['user_id'],))
    return jsonify({"quotations": rows})


# ─────────────────────────────────────────────────────────────────────────────
# BUY COVER → POLICY
#
# FIXED in v2.5: this is the function that actually has policy_no and q
# in scope — DMVIC issuance is now correctly wired here (see the
# enqueue(...) call near the end), not in add_client() where it was
# mistakenly pasted before.
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/quotations/buy', methods=['POST'])
@login_required
@approved_required
def buy_cover():
    d        = request.get_json() or {}
    quote_id = d.get('quote_id', '').strip()
    if not quote_id:
        return jsonify({"error": "quote_id required"}), 400

    q = query("SELECT * FROM quotations WHERE id=%s", (quote_id,), fetchone=True)
    if not q:
        return jsonify({"error": "Quotation not found"}), 404
    if session['role'] != 'admin' and q['agent_id'] != session['user_id']:
        return jsonify({"error": "You do not have access to this quotation"}), 403
    if q['status'] == 'converted':
        return jsonify({"error": "Policy already created for this quotation"}), 409

    client = query("""
        SELECT id FROM clients
        WHERE  phone=%s AND agent_id=%s
    """, (q['phone'], q['agent_id']), fetchone=True)

    if not client:
        client_id = query("""
            INSERT INTO clients
                (agent_id, first_name, last_name, phone, kra_pin, vehicle_reg, email)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            q['agent_id'],
            q['policy_holder_name'], '',
            q['phone'], q.get('kra_pin',''),
            q.get('vehicle_reg',''), q.get('email',''),
        ), commit=True)
    else:
        client_id = client['id']

    if (q.get('company') or '').lower() == 'monarch':
        m_class = monarch_policy_class(q.get('product'))
        if m_class:
            policy_no = next_monarch_policy_no(m_class)
        else:
            policy_no = f"POL-{date.today().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"
            log.warning("Monarch product '%s' has no assigned policy-number series "
                        "(private/commercial only) — falling back to internal format for %s",
                        q.get('product'), policy_no)
    else:
        policy_no = f"POL-{date.today().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"
    query("""
        INSERT INTO policies
            (policy_no, quote_id, agent_id, client_id,
             vehicle_reg, type_of_cover, commencing_date, expiry_date,
             total_payable, status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'active')
    """, (
        policy_no, quote_id, q['agent_id'], client_id,
        q.get('vehicle_reg',''), q['type_of_cover'],
        q['commencing_date'], q['expiry_date'],
        q['total_payable'],
    ), commit=True)

    query("UPDATE quotations SET status='converted' WHERE id=%s",
          (quote_id,), commit=True)

    query("""
        INSERT INTO payments (policy_no, amount, status, method)
        VALUES (%s, %s, 'pending', 'manual')
    """, (policy_no, q['total_payable']), commit=True)

    # DMVIC certificate issuance — async on the background worker pool,
    # same pattern as PDF/email in generate_quotation() above. `q` is the
    # full quotations row, which has everything issue_dmvic_certificate
    # needs (product, cover, dates, vehicle + policyholder details).
    enqueue("dmvic_issue_certificate", issue_dmvic_certificate, policy_no, q)

    cache_delete_prefix("cache:dashboard")
    cache_delete_prefix("cache:reports_summary")

    return jsonify({
        "success":       True,
        "policy_no":     policy_no,
        "total_payable": float(q['total_payable']),
        "message":       "Policy created. Payment pending confirmation."
    })


@app.route('/api/policies/<policy_no>/status')
@login_required
def policy_dmvic_status(policy_no):
    """Lightweight poll target for the quotation wizard's post-purchase
    screen. DMVIC issuance runs on a background worker (see buy_cover()
    above) so the agent isn't blocked waiting on DMVIC's response — the
    frontend polls this every few seconds until dmvic_status settles into
    'issued' / 'failed' / 'pending_manual' / 'pending_confirmation' /
    'unsupported'."""
    row = query("""SELECT policy_no, agent_id, dmvic_status, dmvic_certificate_no,
                          dmvic_transaction_no, dmvic_issuance_request_id, dmvic_error
                   FROM policies WHERE policy_no=%s""", (policy_no,), fetchone=True)
    if not row:
        return jsonify({"error": "Policy not found"}), 404
    if session['role'] != 'admin' and row['agent_id'] != session['user_id']:
        return jsonify({"error": "You do not have access to this policy"}), 403

    return jsonify({
        "policy_no":       row.get('policy_no'),
        "dmvic_status":    row.get('dmvic_status'),
        "certificate_no":  row.get('dmvic_certificate_no'),
        "transaction_no":  row.get('dmvic_transaction_no'),
        "issuance_request_id": row.get('dmvic_issuance_request_id'),
        "error":           row.get('dmvic_error'),
        "can_confirm":     bool(row.get('dmvic_status') == 'pending_manual' and row.get('dmvic_issuance_request_id')),
    })

# ─────────────────────────────────────────────────────────────────────────────
# POLICIES
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/policies/list')
@login_required
def list_policies():
    if session['role'] == 'admin':
        rows = query("""
            SELECT p.*, c.first_name, c.last_name, c.phone,
                   u.full_name AS agent_name,
                   DATEDIFF(p.expiry_date, CURDATE()) AS days_remaining
            FROM   policies p
            LEFT JOIN clients c ON c.id = p.client_id
            LEFT JOIN users   u ON u.id = p.agent_id
            ORDER BY p.created_at DESC
        """)
    else:
        rows = query("""
            SELECT p.*, c.first_name, c.last_name, c.phone,
                   DATEDIFF(p.expiry_date, CURDATE()) AS days_remaining
            FROM   policies p
            LEFT JOIN clients c ON c.id = p.client_id
            WHERE  p.agent_id = %s
            ORDER BY p.created_at DESC
        """, (session['user_id'],))
    return jsonify({"policies": rows})


@app.route('/api/policies/check-double-insurance')
@login_required
@limiter.limit("20 per minute")
def check_double_insurance():
    """Search by vehicle registration number OR chassis number to see if a
    vehicle already has an active (non-expired, non-cancelled) policy —
    the same check DMVIC itself does server-side and returns as ER005
    'Double Insurance'. This lets an agent catch it BEFORE submitting a
    quote/policy, rather than finding out only after DMVIC rejects the
    certificate issuance.

    Chassis number isn't stored on `policies` directly, so this joins to
    `quotations` (via policies.quote_id) to search both fields at once. A
    best-effort DMVIC lookup is returned separately, so agents can catch an
    overlap created by another insurer before they create a new policy.
    """
    q = (request.args.get('query') or '').strip().upper()
    if not q or len(q) < 3:
        return jsonify({"error": "Enter at least 3 characters of the registration or chassis number"}), 400

    commencing_date = (request.args.get('commencing_date') or '').strip()
    expiry_date = (request.args.get('expiry_date') or '').strip()
    today = date.today()
    today_str = today.strftime('%d/%m/%Y')
    dmvic_start = _dmvic_fmt_date(commencing_date) if commencing_date else today_str
    # DMVIC rejects a zero-day period. The dashboard has no period controls,
    # so use the smallest valid range unless the caller supplies dates.
    dmvic_end = _dmvic_fmt_date(expiry_date) if expiry_date else (today + timedelta(days=1)).strftime('%d/%m/%Y')
    # The quotation wizard supplies the two identifiers separately. For the
    # dashboard's one generic search field, never call DMVIC with a partial
    # value and then incorrectly report the vehicle clear.
    dmvic_reg = (request.args.get('vehicle_reg') or '').strip().upper()
    dmvic_chassis = (request.args.get('chassis_number') or '').strip().upper()
    if not dmvic_reg and not dmvic_chassis:
        if re.fullmatch(r"[A-Z]{2,3}\d{3,4}[A-Z]?", q):
            dmvic_reg = q
        elif len(q) >= 6:
            dmvic_chassis = q

    like = f"%{q}%"
    rows = query("""
        SELECT p.policy_no, p.vehicle_reg, p.type_of_cover, p.status,
               p.commencing_date, p.expiry_date,
               DATEDIFF(p.expiry_date, CURDATE()) AS days_remaining,
               qq.chassis_number,
               CONCAT(c.first_name,' ',c.last_name) AS client_name,
               u.full_name AS agent_name
        FROM   policies p
        LEFT JOIN quotations qq ON qq.id = p.quote_id
        LEFT JOIN clients    c  ON c.id  = p.client_id
        LEFT JOIN users      u  ON u.id  = p.agent_id
        WHERE  (p.vehicle_reg LIKE %s OR qq.chassis_number LIKE %s)
        ORDER BY p.expiry_date DESC
        LIMIT 20
    """, (like, like))

    for r in rows:
        active_today = (
            r['status'] == 'active'
            and r['days_remaining'] is not None
            and r['days_remaining'] >= 0
        )
        r['is_double_insurance_risk'] = active_today

    dmvic = {"checked": False, "matches": [], "error": None}
    token = dmvic_get_token() if (dmvic_reg or dmvic_chassis) else None
    if token:
        live = dmvic_validate_double_insurance(
            token,
            policy_start_date=dmvic_start,
            policy_end_date=dmvic_end,
            vehicle_reg=dmvic_reg or None,
            chassis_number=dmvic_chassis or None,
        )
        if _dmvic_is_token_error(live):
            log.warning("DMVIC pre-check rejected its token; refreshing once before reporting failure")
            fresh_token = dmvic_get_token(force_refresh=True)
            if fresh_token:
                live = dmvic_validate_double_insurance(
                    fresh_token,
                    policy_start_date=dmvic_start,
                    policy_end_date=dmvic_end,
                    vehicle_reg=dmvic_reg or None,
                    chassis_number=dmvic_chassis or None,
                )
        if live.get("success"):
            dmvic["checked"] = True
            dmvic["matches"] = live.get("matches") or []
        else:
            dmvic["error"] = live.get("error")
    elif dmvic_reg or dmvic_chassis:
        dmvic["error"] = "Could not obtain a DMVIC auth token."
    else:
        dmvic["error"] = "Enter a full registration or chassis number for the live DMVIC check."

    return jsonify({
        "results": rows,
        "has_active_match": any(r['is_double_insurance_risk'] for r in rows),
        "dmvic": dmvic,
        "has_dmvic_match": bool(dmvic["matches"]),
    })


# ─────────────────────────────────────────────────────────────────────────────
# CLAIMS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/claims', methods=['GET'])
@login_required
def list_claims():
    try:
        if session['role'] == 'admin':
            rows = query("""
                SELECT c.*, u.full_name AS agent_name
                FROM   claims c
                LEFT JOIN users u ON u.id = c.agent_id
                ORDER BY c.created_at DESC
            """)
        else:
            rows = query("""
                SELECT c.*, u.full_name AS agent_name
                FROM   claims c
                LEFT JOIN users u ON u.id = c.agent_id
                WHERE  c.agent_id = %s
                ORDER BY c.created_at DESC
            """, (session['user_id'],))
        return jsonify({"claims": rows})
    except Exception as e:
        return safe_error_response(e, "Could not load claims.")


@app.route('/api/claims', methods=['POST'])
@login_required
@approved_required
def submit_claim():
    d = request.get_json() or {}

    claim_form             = bool(d.get("claim_form"))
    police_abstract        = bool(d.get("police_abstract"))
    inspection_report       = bool(d.get("inspection_report"))
    national_id             = bool(d.get("national_id"))
    driving_licence          = bool(d.get("driving_licence"))
    logbook                 = bool(d.get("logbook"))
    kra_pin                 = bool(d.get("kra_pin"))
    premium_paid             = bool(d.get("premium_paid"))

    try:
        excess_fee = float(d.get("excess_fee", 0) or 0)
    except (TypeError, ValueError):
        excess_fee = 0.0

    excess_paid              = bool(d.get("excess_paid"))
    liability_acknowledged   = bool(d.get("liability_acknowledged"))

    required = (
        'claim_policy',
        'incident_date',
        'incident_type',
        'incident_desc'
    )

    if not all(str(d.get(k, '')).strip() for k in required):
        return jsonify({"error": "All fields are required"}), 400

    try:
        existing = query("""
            SELECT id
            FROM claims
            WHERE claim_policy=%s
            AND incident_date=%s
            AND agent_id=%s
        """, (
            d['claim_policy'],
            d['incident_date'],
            session['user_id']
        ), fetchone=True)
    except Exception as e:
        return safe_error_response(e, "Could not submit claim.")

    if existing:
        return jsonify({
            "error": "A claim with this policy and incident date already exists"
        }), 409

    claim_id = f"CLM-{str(uuid.uuid4())[:8].upper()}"

    try:
        query("""
            INSERT INTO claims (
                id,
                agent_id,
                claim_policy,
                incident_date,
                incident_type,
                incident_desc,
                claim_form,
                police_abstract,
                inspection_report,
                national_id,
                driving_licence,
                logbook,
                kra_pin,
                premium_paid,
                excess_fee,
                excess_paid,
                liability_acknowledged,
                status
            )
            VALUES (
                %s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,
                'pending'
            )
        """, (
            claim_id,
            session['user_id'],
            d['claim_policy'].strip().upper(),
            d['incident_date'],
            d['incident_type'],
            d['incident_desc'].strip(),
            claim_form,
            police_abstract,
            inspection_report,
            national_id,
            driving_licence,
            logbook,
            kra_pin,
            premium_paid,
            excess_fee,
            excess_paid,
            liability_acknowledged
        ), commit=True)
    except Exception as e:
        return safe_error_response(e, "Could not save this claim.")

    cache_delete_prefix("cache:dashboard")

    return jsonify({
        "success": True,
        "claim_id": claim_id,
        "claim": {
            "id": claim_id
        }
    })


@app.route('/api/claims/<claim_id>/status', methods=['POST'])
@login_required
@admin_required
def update_claim_status(claim_id):
    d      = request.get_json() or {}
    status = d.get('status')
    if status not in ('pending', 'approved', 'rejected'):
        return jsonify({"error": "Invalid status"}), 400

    claim = query("SELECT id FROM claims WHERE id=%s", (claim_id,), fetchone=True)
    if not claim:
        return jsonify({"error": "Claim not found"}), 404

    try:
        query("UPDATE claims SET status=%s WHERE id=%s", (status, claim_id), commit=True)
        query("INSERT INTO audit_log (user_id, action, detail) VALUES (%s,%s,%s)",
              (session['user_id'], 'claim_status_change',
               f"claim_id={claim_id} status={status}"), commit=True)
    except Exception as e:
        return safe_error_response(e, "Could not update claim status.")

    cache_delete_prefix("cache:dashboard")
    return jsonify({"success": True})


# ─────────────────────────────────────────────────────────────────────────────
# RENEWALS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/renewals/list')
@login_required
def list_renewals():
    if session['role'] == 'admin':
        rows = query("""
            SELECT p.id, p.policy_no, p.vehicle_reg, p.type_of_cover,
                   p.expiry_date, p.total_payable AS premium,
                   DATEDIFF(p.expiry_date, CURDATE()) AS days_remaining,
                   CONCAT(c.first_name,' ',c.last_name) AS client_name,
                   c.phone,
                   COALESCE(q.make, '') AS vehicle_make,
                   u.full_name AS agent_name
            FROM   policies p
            LEFT JOIN clients c ON c.id = p.client_id
            LEFT JOIN quotations q ON q.id = p.quote_id
            LEFT JOIN users u ON u.id = p.agent_id
            WHERE  DATEDIFF(p.expiry_date, CURDATE()) <= 30
               OR  p.expiry_date < CURDATE()
            ORDER BY p.expiry_date ASC
        """)
    else:
        rows = query("""
            SELECT p.id, p.policy_no, p.vehicle_reg, p.type_of_cover,
                   p.expiry_date, p.total_payable AS premium,
                   DATEDIFF(p.expiry_date, CURDATE()) AS days_remaining,
                   CONCAT(c.first_name,' ',c.last_name) AS client_name,
                   c.phone,
                   COALESCE(q.make, '') AS vehicle_make
            FROM   policies p
            LEFT JOIN clients c ON c.id = p.client_id
            LEFT JOIN quotations q ON q.id = p.quote_id
            WHERE  p.agent_id = %s
              AND (DATEDIFF(p.expiry_date, CURDATE()) <= 30
               OR  p.expiry_date < CURDATE())
            ORDER BY p.expiry_date ASC
        """, (session['user_id'],))
    return jsonify({"renewals": rows})


@app.route('/api/renewals/renew', methods=['POST'])
@login_required
@approved_required
def renew_policy():
    d  = request.get_json() or {}
    pid = d.get('id')
    if not pid:
        return jsonify({"error": "Policy ID required"}), 400

    p = query("SELECT * FROM policies WHERE id=%s", (pid,), fetchone=True)
    if not p:
        return jsonify({"error": "Policy not found"}), 404
    if session['role'] != 'admin' and p['agent_id'] != session['user_id']:
        return jsonify({"error": "You do not have access to this policy"}), 403

    old_expiry = p['expiry_date']
    if isinstance(old_expiry, str):
        old_expiry = datetime.strptime(old_expiry, '%Y-%m-%d').date()
    new_expiry = old_expiry.replace(year=old_expiry.year + 1)

    query("""
        UPDATE policies
        SET expiry_date=%s, status='active', updated_at=NOW()
        WHERE id=%s
    """, (str(new_expiry), pid), commit=True)

    return jsonify({"success": True, "new_expiry": str(new_expiry)})


# ─────────────────────────────────────────────────────────────────────────────
# REPORTS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/reports/download')
@login_required
@admin_required
def download_report():
    period = request.args.get('period', 'monthly')
    fmt    = request.args.get('format', 'pdf')

    if period not in ('daily','weekly','monthly','yearly'):
        return jsonify({"error": "Invalid period"}), 400
    if fmt not in ('pdf','excel'):
        return jsonify({"error": "Invalid format"}), 400

    rdata = build_report_data(period)

    if fmt == 'pdf':
        data = generate_report_pdf(rdata)
        if not data:
            return jsonify({"error": "PDF generation unavailable — install reportlab"}), 503
        return send_file(
            io.BytesIO(data),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"Westlake_{period}_report_{date.today()}.pdf"
        )
    else:
        data = generate_report_excel(rdata)
        if not data:
            return jsonify({"error": "Excel generation unavailable — install openpyxl"}), 503
        return send_file(
            io.BytesIO(data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Westlake_{period}_report_{date.today()}.xlsx"
        )


@app.route('/api/reports/summary')
@login_required
@admin_required
@cached_response("cache:reports_summary", ttl=120)
def report_summary():
    period = request.args.get('period', 'monthly')
    rdata  = build_report_data(period)
    return jsonify({
        "period":   rdata['period'],
        "start":    rdata['start_date'],
        "end":      rdata['end_date'],
        "totals":   rdata['totals'],
        "agents":   rdata['agents'],
    })


# ─────────────────────────────────────────────────────────────────────────────
# AUDIT LOG (admin only)
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/audit/log')
@login_required
@admin_required
def audit_log():
    rows = query("""
        SELECT a.*, u.full_name, u.username
        FROM   audit_log a
        LEFT JOIN users u ON u.id = a.user_id
        ORDER BY a.created_at DESC
        LIMIT 500
    """)
    return jsonify({"log": rows})


# ─────────────────────────────────────────────────────────────────────────────
# M-PESA ROUTES
# ─────────────────────────────────────────────────────────────────────────────

MAX_UNDERPAYMENT_ALLOWED = 800

@app.route('/api/mpesa/stk', methods=['POST'])
@login_required
@approved_required
@limiter.limit("6 per minute")
def mpesa_stk():
    d         = request.get_json() or {}
    policy_no = d.get('policy_no','').strip()
    phone     = d.get('phone','').strip()
    amount    = d.get('amount', 0)

    if not policy_no or not phone or not amount:
        return jsonify({"error": "policy_no, phone and amount are required"}), 400

    policy = query("SELECT * FROM policies WHERE policy_no=%s", (policy_no,), fetchone=True)
    if not policy:
        return jsonify({"error": "Policy not found"}), 404
    if session['role'] != 'admin' and policy['agent_id'] != session['user_id']:
        return jsonify({"error": "You do not have access to this policy"}), 403

    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid amount"}), 400

    if amount <= 0:
        return jsonify({"error": "Amount must be greater than zero"}), 400

    quoted_amount = float(policy['total_payable'])
    paid_row = query("""
        SELECT COALESCE(SUM(amount),0) AS paid FROM payments
        WHERE policy_no=%s AND status='completed'
    """, (policy_no,), fetchone=True) or {}
    already_paid = float(paid_row.get('paid', 0) or 0)
    balance = round(quoted_amount - already_paid, 2)

    if balance <= 0:
        return jsonify({"error": "This policy has already been paid in full."}), 400

    # Underpayment no longer blocks the STK push — the amount is fully
    # agent-customizable, with no floor or ceiling. Any amount below the
    # outstanding balance still triggers a warning email + account flag
    # for admin visibility, but the payment always proceeds.
    if amount < balance:
        agent = {
            'name':  session.get('name', session.get('username')),
            'email': session.get('email', ''),
        }
        shortfall = balance - amount
        flag_reason = (f"Paid KES {amount:,.0f} against policy {policy_no}, "
                       f"KES {shortfall:,.0f} below the outstanding balance of KES {balance:,.0f} "
                       f"(on {datetime.now().strftime('%Y-%m-%d %H:%M')}).")

        # Flag the agent's account: increment their running underpayment-attempt
        # counter and mark flagged=1 so admins can see/filter this in the agents
        # list, in addition to the warning email below. This is informational
        # only now — it never blocks the request.
        try:
            query("""UPDATE users
                     SET flagged=1, flagged_reason=%s, flagged_at=NOW(),
                         underpayment_attempts = underpayment_attempts + 1
                     WHERE id=%s""",
                  (flag_reason, session['user_id']), commit=True)
        except Exception as e:
            # Don't let a missing migration silently block the whole request —
            # log it loudly so it gets noticed and fixed.
            log.error("Could not flag account for underpayment (has "
                      "migrations_add_underpayment_flag.sql been run?): %s", type(e).__name__)

        enqueue("underpayment_alert", notify_underpayment_attempt,
                agent, policy_no, balance, amount, phone)

        query("INSERT INTO audit_log (user_id, action, detail) VALUES (%s,%s,%s)",
              (session['user_id'], 'underpayment_warning',
               f"policy={policy_no} balance={balance} attempted={amount}"), commit=True)

    result = mpesa_stk_push(phone=phone, amount=amount,
                            account_ref=policy_no, description='Insurance Premium')
    if not result['success']:
        return jsonify({"error": result.get('error','STK push failed')}), 400

    existing = query("SELECT id FROM payments WHERE policy_no=%s AND method='mpesa'",
                     (policy_no,), fetchone=True)
    if existing:
        query("""UPDATE payments SET amount=%s, reference=%s, status='pending', paid_at=NULL
                 WHERE policy_no=%s AND method='mpesa'""",
              (amount, result['checkout_request_id'], policy_no), commit=True)
    else:
        query("INSERT INTO payments (policy_no, amount, status, method, reference) VALUES (%s,%s,'pending','mpesa',%s)",
              (policy_no, amount, result['checkout_request_id']), commit=True)

    return jsonify({
        "success":             True,
        "checkout_request_id": result['checkout_request_id'],
        "customer_message":    result.get('customer_message','Check your phone for the M-Pesa prompt.'),
    })


@app.route('/api/mpesa/query', methods=['POST'])
@login_required
def mpesa_query():
    d                   = request.get_json() or {}
    checkout_request_id = d.get('checkout_request_id','').strip()
    policy_no           = d.get('policy_no','').strip()
    if not checkout_request_id:
        return jsonify({"error": "checkout_request_id required"}), 400

    if policy_no:
        policy = query("SELECT agent_id FROM policies WHERE policy_no=%s", (policy_no,), fetchone=True)
        if not policy:
            return jsonify({"error": "Policy not found"}), 404
        if session['role'] != 'admin' and policy['agent_id'] != session['user_id']:
            return jsonify({"error": "You do not have access to this policy"}), 403
    elif session['role'] != 'admin':
        return jsonify({"error": "policy_no required"}), 400

    result = mpesa_query_status(checkout_request_id)
    if result.get('success') and policy_no:
        query("UPDATE payments SET status='completed', paid_at=NOW() WHERE reference=%s",
              (checkout_request_id,), commit=True)
        query("UPDATE policies SET status='active', updated_at=NOW() WHERE policy_no=%s",
              (policy_no,), commit=True)
        query("INSERT INTO audit_log (user_id, action, detail) VALUES (%s,%s,%s)",
              (session.get('user_id'), 'mpesa_payment_confirmed',
               f"policy={policy_no} ref={checkout_request_id}"), commit=True)
    return jsonify(result)


def _request_ip():
    fwd = request.headers.get('X-Forwarded-For', '')
    if fwd:
        return fwd.split(',')[0].strip()
    return request.remote_addr or ''


def _is_safaricom_ip(ip_str):
    try:
        ip = ipaddress.ip_address(ip_str)
    except ValueError:
        return False
    return False


@app.route(f'/api/mpesa/callback/<secret>', methods=['POST'])
@csrf.exempt
def mpesa_callback(secret):
    # CSRF-exempt: this endpoint is called directly by ArchPay's servers over a
    # server-to-server POST with no session cookie. Security comes from the
    # secret path segment and matching callbacks to transactions we initiated.
    if not hmac_compare(secret, MPESA_CALLBACK_SECRET):
        log.warning("ArchPay callback rejected: bad secret path from %s", _request_ip())
        abort(404)

    try:
        data = request.get_json(force=True) or {}
        callback_data = data.get('data') if isinstance(data.get('data'), dict) else data
        ref = (callback_data.get('checkoutRequestId')
               or callback_data.get('checkout_request_id')
               or callback_data.get('CheckoutRequestID')
               or '').strip()
        status = (callback_data.get('status') or '').strip().lower()
        receipt = callback_data.get('mpesaReceiptNumber')
        if not ref:
            return jsonify({"received": False, "error": "Missing checkoutRequestId"}), 400

        pmt = query("SELECT policy_no FROM payments WHERE reference=%s", (ref,), fetchone=True)
        if not pmt:
            log.warning("ArchPay callback ignored: unknown checkoutRequestId=%s", ref)
            return jsonify({"received": True})

        terminal_failures = ('cancelled', 'canceled', 'timeout', 'reversed')
        if status in ('completed', 'success', 'successful', 'paid') or (receipt and status not in terminal_failures):
            mpesa_ref = receipt or ref
            query("UPDATE payments SET status='completed', paid_at=NOW() WHERE reference=%s",
                  (ref,), commit=True)
            query("UPDATE policies SET status='active', updated_at=NOW() WHERE policy_no=%s",
                  (pmt['policy_no'],), commit=True)
            query("INSERT INTO audit_log (action, detail) VALUES (%s,%s)",
                  ('archpay_callback_confirmed', f"policy={pmt['policy_no']} ref={mpesa_ref or ref}"), commit=True)
        elif status in ('failed', 'cancelled', 'timeout'):
            query("UPDATE payments SET status=%s WHERE reference=%s", (status, ref), commit=True)
            query("INSERT INTO audit_log (action, detail) VALUES (%s,%s)",
                  ('archpay_callback_not_completed', f"policy={pmt['policy_no']} ref={ref} status={status}"), commit=True)
        else:
            log.info("ArchPay callback received non-final status=%s ref=%s", status, ref)
    except Exception as e:
        log.error("ArchPay callback processing error: %s", type(e).__name__)
    return jsonify({"received": True})


def hmac_compare(a, b):
    import hmac as _hmac
    return _hmac.compare_digest(str(a), str(b))


@app.route('/api/admin/dmvic/pending-confirmations')
@login_required
@admin_required
def list_pending_dmvic_confirmations():
    """Admin review queue for DMVIC policy alerts held with an issuance ID."""
    policies = query("""SELECT policy_no, quote_id, agent_id, vehicle_reg, total_payable,
                               dmvic_issuance_request_id, dmvic_error, created_at
                        FROM policies
                        WHERE dmvic_status=%s
                        ORDER BY created_at DESC""", ('pending_confirmation',))
    pending = []
    for policy in policies:
        quote = query("""SELECT policy_holder_name, chassis_number, company
                         FROM quotations WHERE id=%s""", (policy.get('quote_id'),), fetchone=True) or {}
        agent = query("SELECT full_name FROM users WHERE id=%s", (policy.get('agent_id'),), fetchone=True) or {}
        pending.append({
            "policy_no": policy.get('policy_no'),
            "vehicle_reg": policy.get('vehicle_reg'),
            "chassis_number": quote.get('chassis_number'),
            "policy_holder_name": quote.get('policy_holder_name'),
            "company": quote.get('company'),
            "agent_name": agent.get('full_name'),
            "total_payable": policy.get('total_payable'),
            "dmvic_error": policy.get('dmvic_error'),
            "created_at": policy.get('created_at'),
        })
    return jsonify({"pending": pending})


@app.route('/api/dmvic/confirm-issuance', methods=['POST'])
@login_required
@approved_required
def dmvic_confirm_issuance_route():
    data = request.get_json() or {}
    policy_no = (data.get('policy_no') or '').strip()
    is_approved = data.get('is_approved')
    is_logbook_verified = data.get('is_logbook_verified')
    is_vehicle_inspected = data.get('is_vehicle_inspected')
    additional_comments = (data.get('additional_comments') or '').strip()

    if not policy_no:
        return jsonify({"error": "policy_no is required"}), 400
    if not isinstance(is_approved, bool):
        return jsonify({"error": "is_approved must be true or false"}), 400
    if not isinstance(is_logbook_verified, bool) or not isinstance(is_vehicle_inspected, bool):
        return jsonify({"error": "is_logbook_verified and is_vehicle_inspected must be explicitly true or false"}), 400
    if is_approved and not (is_logbook_verified and is_vehicle_inspected):
        return jsonify({"error": "Both logbook verification and vehicle inspection must be "
                                  "confirmed before approving issuance."}), 400

    policy = query("""SELECT policy_no, agent_id, dmvic_issuance_request_id, dmvic_status
                      FROM policies WHERE policy_no=%s""", (policy_no,), fetchone=True)
    if not policy:
        return jsonify({"error": "Policy not found"}), 404
    if session['role'] != 'admin' and policy.get('agent_id') != session['user_id']:
        return jsonify({"error": "You do not have access to this policy"}), 403
    if policy.get('dmvic_status') != 'pending_manual':
        return jsonify({"error": "This policy has no pending DMVIC issuance alert to act on."}), 409
    issuance_request_id = policy.get('dmvic_issuance_request_id')
    if not issuance_request_id:
        return jsonify({"error": "No IssuanceRequestID stored for this policy — cannot confirm. "
                                  "Check dmvic_error for details or re-attempt issuance."}), 409

    token = dmvic_get_token()
    if not token:
        return jsonify({"error": "Could not obtain DMVIC auth token. Please try again."}), 503

    result = dmvic_issue_with_retry(
        dmvic_confirm_certificate_issuance,
        token,
        issuance_request_id=issuance_request_id,
        is_approved=is_approved,
        is_logbook_verified=is_logbook_verified,
        is_vehicle_inspected=is_vehicle_inspected,
        additional_comments=additional_comments,
        usernames=session.get('username', ''),
    )

    action = 'dmvic_confirm_issuance' if is_approved else 'dmvic_reject_issuance'
    query("INSERT INTO audit_log (user_id, action, detail) VALUES (%s,%s,%s)",
          (session['user_id'], action,
           f"policy={policy_no} issuance_request_id={issuance_request_id} approved={is_approved}"), commit=True)

    if not is_approved:
        query("""UPDATE policies SET dmvic_status='failed',
                  dmvic_error='Issuance rejected by agent after manual review.'
                  WHERE policy_no=%s""", (policy_no,), commit=True)
        cache_delete_prefix("cache:dashboard")
        return jsonify({"success": True, "approved": False})

    if result.get('success'):
        query("""UPDATE policies
                  SET dmvic_status='issued',
                      dmvic_transaction_no=%s,
                      dmvic_certificate_no=%s,
                      dmvic_api_request_no=%s,
                      dmvic_issued_at=NOW(),
                      dmvic_error=NULL
                  WHERE policy_no=%s""",
              (result.get('transaction_no'), result.get('certificate_no'),
               result.get('api_request_number'), policy_no), commit=True)
        cache_delete_prefix("cache:dashboard")
        log.info("DMVIC certificate confirmed for %s: %s", policy_no, result.get('certificate_no'))
        return jsonify({
            "success": True,
            "approved": True,
            "certificate_no": result.get('certificate_no'),
            "transaction_no": result.get('transaction_no'),
        })

    error_msg = result.get('error', 'Confirmation failed.')
    query("""UPDATE policies SET dmvic_error=%s WHERE policy_no=%s""",
          (error_msg, policy_no), commit=True)
    log.warning("DMVIC confirm-issuance failed for %s: %s", policy_no, error_msg)
    return jsonify({"error": error_msg}), 502

@app.route('/api/dmvic/record-policy-alert', methods=['POST'])
@login_required
@admin_required
def record_dmvic_policy_alert():
    """Reject the obsolete recovery flow without sending DMVIC a bad request."""
    data = request.get_json() or {}
    policy_no = (data.get('policy_no') or '').strip()
    issuance_request_id = (data.get('issuance_request_id') or '').strip()

    if not policy_no:
        return jsonify({"error": "policy_no is required"}), 400
    if not re.fullmatch(r"[A-Za-z0-9-]{3,80}", issuance_request_id):
        return jsonify({"error": "Enter the Issuance Request ID supplied by DMVIC Support."}), 400

    policy = query("SELECT policy_no, dmvic_status FROM policies WHERE policy_no=%s",
                   (policy_no,), fetchone=True)
    if not policy:
        return jsonify({"error": "Policy not found"}), 404
    if policy.get('dmvic_status') == 'issued':
        return jsonify({"error": "This policy already has an issued DMVIC certificate."}), 409

    reason = ("DMVIC policy alert recorded from Support. DMVIC has confirmed that "
              "manual confirmation is unavailable; verify the logbook and insurer "
              "record, then create a corrected quotation.")
    query("""UPDATE policies
             SET dmvic_status='pending_manual',
                 dmvic_issuance_request_id=%s,
                 dmvic_error=%s
             WHERE policy_no=%s""",
          (issuance_request_id, reason, policy_no), commit=True)
    query("INSERT INTO audit_log (user_id, action, detail) VALUES (%s,%s,%s)",
          (session['user_id'], 'dmvic_record_policy_alert',
           f"policy={policy_no} issuance_request_id={issuance_request_id}"), commit=True)
    return jsonify({"success": True, "message": "DMVIC alert recorded for data review; no manual confirmation request was sent."})


@app.route('/api/mpesa/status')
@login_required
def mpesa_config_status():
    configured = bool(ARCHPAY_API_KEY)
    return jsonify({
        "configured": configured,
        "provider": "archpay",
        "env": ARCHPAY_MODE,
        "channel_id": ARCHPAY_CHANNEL_ID or None,
    })


if __name__ == '__main__':
    if IS_PRODUCTION:
        log.warning("Running via `python app.py` in a production environment is not "
                    "recommended — use a WSGI server (gunicorn/uwsgi) behind a reverse proxy instead.")
    app.run(debug=DEBUG_MODE, host=os.environ.get('HOST', '127.0.0.1'),
            port=int(os.environ.get('PORT', 8080)))
